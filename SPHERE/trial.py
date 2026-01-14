from fastapi import FastAPI, HTTPException, Query #http is used for get,post,put/delete-communication btw dart and python. #fastapi-talks from python to flutter using http
from pydantic import BaseModel #input quality checker which makes sure no required fields r missing, is structured correctly and has the right types.
import openpyxl #bridge that lets Python understand and extract data from Excel spreadsheets.
import json #used to store text files like dictionaries or lists
import os #lets python interact with operation systems. e.g.windows,macOS,Linux
import tempfile
from typing import List, Dict, Optional, Any
from geopy.distance import geodesic #used to calculate the shortest distance btw two point on the earth, using latitude and longitude
from geopy.geocoders import Nominatim #converting a place name like "Germany" into latitude and longtiude coordinates
import requests
import re
import hashlib # for hashing passwords
from fastapi import Request
from fastapi.middleware.cors import CORSMiddleware
import portalocker
from openpyxl import load_workbook


BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # folder of trial.py (SPHERE)
USERS_FILE = os.path.join(BASE_DIR, "users.xlsx")

#1.File Storage Setup
#Function to locate where user data is stored on their system
def get_user_data_path(filename):
    if os.name == "nt":  # Windows
        base = os.path.join(os.environ.get("LOCALAPPDATA", os.path.expanduser("~")), "SPHERE")
    else:  # macOS/Linux
        base = os.path.join(os.path.expanduser("~"), ".sphere")
    os.makedirs(base, exist_ok=True)
    return os.path.join(base, filename)

# Define file names
PROFILE_FILE = os.path.join(BASE_DIR, "profiles.json")
print("Using profile file:", PROFILE_FILE)

EMISSION_FACTORS_FILE = get_user_data_path("custom_emission_factors.json")

print("Profile file path:", PROFILE_FILE)
# Make sure required files exist    
def ensure_default_files():
    if not os.path.exists(EMISSION_FACTORS_FILE):
        with open(EMISSION_FACTORS_FILE, "w", encoding="utf-8") as f:
            json.dump({}, f)
    if not os.path.exists(PROFILE_FILE):
        with open(PROFILE_FILE, "w", encoding="utf-8") as f:
            json.dump({}, f)

ensure_default_files()

def load_profiles() -> dict:
    if not os.path.exists(PROFILE_FILE):
        return {}
    try:
        with open(PROFILE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except json.JSONDecodeError:
        with open(PROFILE_FILE, "w", encoding="utf-8") as f:
            json.dump({}, f)
        return {}


def ensure_users_excel():
    """
    Make sure the users.xlsx file exists with a 'Users' sheet and headers.
    """
    if not os.path.exists(USERS_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["username", "password_hash"])
        wb.save(USERS_FILE)


def load_users_sheet():
    """
    Return workbook + sheet for users.xlsx, ensuring it exists.
    """
    ensure_users_excel()
    wb = openpyxl.load_workbook(USERS_FILE)
    if "Users" not in wb.sheetnames:
        ws = wb.create_sheet("Users")
        ws.append(["username", "password_hash"])
    else:
        ws = wb["Users"]
    return wb, ws


def find_user_row(ws, username: str):
    """
    Find the row index (1-based) of a given username in the Users sheet.
    Returns None if not found.
    """
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value == username:
            return row
    return None


def hash_password(password: str) -> str:
    """
    Simple SHA-256 hash for passwords
    """
    return hashlib.sha256(password.encode("utf-8")).hexdigest()


def save_profiles(profiles: dict) -> None:
    with open(PROFILE_FILE, "w", encoding="utf-8") as f:
        json.dump(profiles, f, indent=2, ensure_ascii=False)

def get_user_profiles_bucket(all_profiles: dict, username: str) -> dict:
    """
    Returns the dict of profiles for this user, creating it if missing.
    profiles.json becomes:
      { "username": { "profile_name": {...}, ... }, ... }
    """
    if username not in all_profiles or not isinstance(all_profiles.get(username), dict):
        all_profiles[username] = {}
    return all_profiles[username]


def load_custom_emission_factors():
    if os.path.exists(EMISSION_FACTORS_FILE):
        with open(EMISSION_FACTORS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_custom_emission_factors(data):
    with open(EMISSION_FACTORS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

custom_emission_factors = load_custom_emission_factors()

# 2. LOAD EXCEL DATA

# Open the Excel workbook
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "Emission Data.xlsx")
LOCK_PATH = EXCEL_PATH + ".lock"

def _atomic_save_workbook(wb, excel_path: str):
    folder = os.path.dirname(excel_path)
    fd, tmp_path = tempfile.mkstemp(prefix="excel_", suffix=".xlsx", dir=folder)
    os.close(fd)
    try:
        wb.save(tmp_path)
        os.replace(tmp_path, excel_path)
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except:
                pass

book = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# Load each sheet that SPHERE uses
sheet  = book['Data Sheet']
sheet2 = book['Transport']
sheet3 = book['Transport WTT']
sheet4 = book['Machine Types']
sheet5 = book['Recycling']
sheet6 = book['Packaging']
sheet7 = book["Materials"]
sheet8 = book["GWP of GHG_gases"]
sheet9 = book["Disassembly by Industry"]
sheet10 = book["Usage Types"]
sheet11 = book["Facilities"]
sheet12 = book["Process"]
sheet13 = book['Machine_Process']
sheet14 = book['YCM']
sheet15 = book['Amada']
sheet16 = book['Mazak']
sheet17 = book['Van']
sheet18 = book['HGV']
sheet19 = book['Freight Flight']
sheet20 = book['Rail']
sheet21 = book['Ship']
sheet22 = book['Usage']
sheet23 = book['End-of-Life']
sheet24 = book['Assembly']
sheet25 = book['Waste']

# Helper functions to read values from Excel columns
def extract_selection_list(cells):
    """Stops reading when it finds an empty cell."""
    values = []
    for cell in cells:
        value = cell[0].value if cell[0].value is not None else ""
        value = value.strip() if isinstance(value, str) else value
        if value == "" or value is None:
            break
        values.append(value)
    return values

def extract_list(cells):
    """Reads all values regardless of blanks."""
    values = []
    for cell in cells:
        value = cell[0].value if cell[0].value is not None else ""
        values.append(value)
    return values
def extract_material_list(cells):
    values=[]
    for cell in cells:
        value = cell[0].value if cell[0].value is not None else ""
        if value == "":
            break
        values.append(value)
    return values

def extract_transport_list(cells):
    # SPHERE stops when it hits blank for transport types :contentReference[oaicite:2]{index=2}
    values = []
    for cell in cells:
        value = cell[0].value if cell[0].value is not None else ""
        if value == "":
            break
        values.append(value)
    return values

def extract_emission_list(cells):
    values = []
    for cell in cells:
        value = cell[0].value
        if value is None or str(value).strip() == "":
            break
        try:
            values.append(float(value))
        except Exception:
            values.append("N/A")
    return values

# pull columns from workbook (exact same cells used in SPHERE.py) :contentReference[oaicite:3]{index=3}
country_cells        = sheet['A2':'A999']
distance_cells       = sheet['B2':'B999']
steel_cells          = sheet['C2':'C999']
aluminium_cells      = sheet['D2':'D999']
cement_cells         = sheet['E2':'E999']
electricity_cells    = sheet['F2':'F999']
plastic_cells        = sheet['G2':'G999']
carbon_fiber_cells   = sheet['H2':'H999']
materials_cells      = sheet7['A2':'A999']

Transport_cells      = sheet2['A2':'A999']

van_cells            = sheet3['B2':'B5']
van_diesel_cells     = sheet3['D2':'D5']
van_petrol_cells     = sheet3['E2':'E5']
van_battery_cells    = sheet3['F2':'F5']

HGV_diesel_cells     = sheet3['B8':'B15']
HGV_0_diesel_cells   = sheet3['D8':'D15']
HGV_50_diesel_cells  = sheet3['E8':'E15']
HGV_100_diesel_cells = sheet3['F8':'F15']

HGV_R_diesel_cells     = sheet3['B17':'B24']
HGV_R_0_diesel_cells   = sheet3['D17':'D24']
HGV_R_50_diesel_cells  = sheet3['E17':'E24']
HGV_R_100_diesel_cells = sheet3['F17':'F24']

flight_cells           = sheet3['B27':'B29']
Int_flight_cells       = sheet3['D27':'D27']
L_1500_flight_cells    = sheet3['E27':'E29']
M_1500_flight_cells    = sheet3['F27':'F29']

rail_cells             = sheet3['B31':'B31']
freight_train_cells    = sheet3['D31':'D31']

sea_tanker_cells       = sheet3['G34':'G38']
crude_tanker_cells     = sheet3['C34':'C40']
products_tanker_cells  = sheet3['C41':'C46']
chemical_tanker_cells  = sheet3['C47':'C51']
LNG_tanker_cells       = sheet3['C52':'C54']
LPG_tanker_cells       = sheet3['C55':'C57']

crude_value_cells      = sheet3['E34':'E40']
products_value_cells   = sheet3['E41':'E46']
chemical_value_cells   = sheet3['E47':'E51']
LNG_value_cells        = sheet3['E52':'E54']
LPG_value_cells        = sheet3['E55':'E57']

cargo_cells            = sheet3['G60':'G64']
bulk_value_cells       = sheet3['E59':'E65']
general_value_cells    = sheet3['E66':'E72']
container_value_cells  = sheet3['E73':'E79']
vehicle_value_cells    = sheet3['E80':'E82']
roro_value_cells       = sheet3['E83':'E85']

van_mode_cells         = sheet17['A11':'A22']
van_emission_cells     = sheet17['C11':'C22']

HGV_mode_cells         = sheet18['A21':'A52']
HGV_emission_cells     = sheet18['C21':'C52']
HGV_r_mode_cells       = sheet18['E21':'E52']
HGV_r_emission_cells   = sheet18['G21':'G52']

Freight_Flight_mode_cells       = sheet19['A10':'A15']
Freight_Flight_emission_cells   = sheet19['C10':'C15']

Rail_mode_cells      = sheet20['A7':'A8']
Rail_emission_cells  = sheet20['C7':'C8']

Sea_Tanker_mode_cells    = sheet21['L2':'L25']
Sea_Tanker_emission_cells = sheet21['N2':'N25']

Cargo_ship_mode_cells       = sheet21['L26':'L54']
Cargo_ship_emission_cells   = sheet21['N26':'N54']

machine_value_cells                 = sheet4['A2':'A999']
specific_machine_energy_use_cells   = sheet4['B2':'B999']
machine_types_cells                 = sheet13['A2':'A999']
Machine_brands_cells                = sheet13['C2':'C999']

YCM_machine_model_cells             = sheet14['A2':'A999']
YCM_main_spindle_cells              = sheet14['B2':'B999']
YCM_sub_spindle_cells               = sheet14['C2':'C999']

Amada_machine_model_cells           = sheet15['A2':'A999']
Amada_main_spindle_cells            = sheet15['B2':'B999']
Amada_sub_spindle_cells             = sheet15['C2':'C999']

Mazak_machine_model_cells           = sheet16['A2':'A999']
Mazak_main_spindle_cells            = sheet16['B2':'B999']
Mazak_secondary_spindle_cells       = sheet16['C2':'C999']
Mazak_second_spindle_cells          = sheet16['D2':'D999']

metal_recycling_types_cells    = sheet5['A2':'A999']
metal_recycling_emission_cells = sheet5['B2':'B999']

packaging_types_cells    = sheet6['A2':'A999']
packaging_box_frame_cells = sheet6['B2':'B999']

Indicator_GHG_cells    = sheet8["A2":"A999"]
GHG_values_cells       = sheet8["B2":"B999"]
GWP_for_GHG_cells      = sheet8["C2":"C999"]

Process_cells   = sheet12["A2":"A999"]
Facilities_cells    = sheet11["A2":"A999"]
Usage_type_cells     = sheet10["A2":"A999"]
Disassembly_by_Industry_cells    = sheet9["A2":"A999"]

Usage_category_cells    = sheet22["A2":"A5"]
Usage_electronic_cells  = sheet22["C2":"C5"]
Usage_electronic_ef_cells  = sheet22["D2":"D5"]
Usage_energy_cells      = sheet22["E2":"E5"]
Usage_energy_ef_cells      = sheet22["F2":"F5"]
Usage_Consumables_cells = sheet22["G2":"G5"]
Usage_Consumables_ef_cells = sheet22["H2":"H5"]
Usage_services_cells    = sheet22["I2":"I5"]
Usage_services_ef_cells    = sheet22["J2":"J5"]

End_of_Life_Activities_cells  = sheet23["A2":"A5"]
End_of_Life_ef_cells          = sheet23["B2":"B5"]

Assembly_modes_cells  = sheet24["A2":"A999"]
Assembly_ef_cells     = sheet24["B2":"B999"]

Waste_mode_cells    = sheet25["A2":"A999"]
Waste_ef_cells      = sheet25["B2":"B999"]
# turn into lists
country_list      = extract_selection_list(country_cells)
distance_list     = extract_list(distance_cells)
steel_list        = extract_list(steel_cells)
aluminium_list    = extract_list(aluminium_cells)
cement_list       = extract_list(cement_cells)
electricity_list  = extract_selection_list(electricity_cells)
plastic_list      = extract_list(plastic_cells)
carbon_fiber_list = extract_list(carbon_fiber_cells)
material_list     = extract_material_list(materials_cells)

transport_list    = extract_transport_list(Transport_cells)

van_list          = extract_list(van_cells)
van_diesel_list   = extract_list(van_diesel_cells)
van_petrol_list   = extract_list(van_petrol_cells)
van_battery_list  = extract_list(van_battery_cells)

# (note: we’re loading the HGV/ship/flight/etc. lists too in case you expand modes later)
HGV_diesel_list     = extract_list(HGV_diesel_cells)
HGV_0_diesel_list   = extract_list(HGV_0_diesel_cells)
HGV_50_diesel_list  = extract_list(HGV_50_diesel_cells)
HGV_100_diesel_list = extract_list(HGV_100_diesel_cells)

HGV_R_diesel_list     = extract_list(HGV_R_diesel_cells)
HGV_R_0_diesel_list   = extract_list(HGV_R_0_diesel_cells)
HGV_R_50_diesel_list  = extract_list(HGV_R_50_diesel_cells)
HGV_R_100_diesel_list = extract_list(HGV_R_100_diesel_cells)

van_mode_list       = extract_list(van_mode_cells)
van_emission_list   = extract_list(van_emission_cells)
HGV_mode_list       = extract_list(HGV_mode_cells)
HGV_emission_list   = extract_list(HGV_emission_cells)
HGV_r_mode_list     = extract_list(HGV_r_mode_cells)
HGV_r_emission_list = extract_list(HGV_r_emission_cells)
Freight_Flight_mode_list    = extract_list(Freight_Flight_mode_cells)
Freight_Flight_emission_list    = extract_emission_list(Freight_Flight_emission_cells)
Rail_mode_list      = extract_list(Rail_mode_cells)
Rail_emission_list = extract_emission_list(Rail_emission_cells)
Sea_Tanker_mode_list = extract_list(Sea_Tanker_mode_cells)
Sea_Tanker_emission_list = extract_emission_list(Sea_Tanker_emission_cells)
Cargo_ship_mode_list     = extract_list(Cargo_ship_mode_cells)
Cargo_ship_emission_list = extract_emission_list(Cargo_ship_emission_cells)

flight_list        = extract_list(flight_cells)
Int_flight_list    = extract_list(Int_flight_cells)
L_1500_flight_list = extract_list(L_1500_flight_cells)
M_1500_flight_list = extract_list(M_1500_flight_cells)

rail_list              = extract_list(rail_cells)
freight_train_rail_list = extract_list(freight_train_cells)

sea_tanker_list     = extract_list(sea_tanker_cells)
crude_tanker_list   = extract_list(crude_tanker_cells)
products_tanker_list = extract_list(products_tanker_cells)
chemical_tanker_list = extract_list(chemical_tanker_cells)
LNG_tanker_list     = extract_list(LNG_tanker_cells)
LPG_tanker_list     = extract_list(LPG_tanker_cells)

crude_value_list     = extract_list(crude_value_cells)
products_value_list  = extract_list(products_value_cells)
chemical_value_list  = extract_list(chemical_value_cells)
LNG_value_list       = extract_list(LNG_value_cells)
LPG_value_list       = extract_list(LPG_value_cells)

cargo_list          = extract_list(cargo_cells)
bulk_value_list     = extract_list(bulk_value_cells)
general_value_list  = extract_list(general_value_cells)
container_value_list = extract_list(container_value_cells)
vehicle_value_list   = extract_list(vehicle_value_cells)
roro_value_list      = extract_list(roro_value_cells)

machine_value_list               = extract_selection_list(machine_value_cells)
specific_machine_energy_use_list = extract_list(specific_machine_energy_use_cells)
machine_types                    = extract_selection_list(machine_types_cells)
Machine_brands                   = extract_selection_list(Machine_brands_cells)

YCM_machine_model                = extract_selection_list(YCM_machine_model_cells)
YCM_main_spindle                 = extract_selection_list(YCM_main_spindle_cells)
YCM_sub_spindle                  = extract_selection_list(YCM_sub_spindle_cells)

Amada_machine_model              = extract_selection_list(Amada_machine_model_cells)
Amada_main_spindle               = extract_selection_list(Amada_main_spindle_cells)
Amada_sub_spindle                = extract_selection_list(Amada_sub_spindle_cells)

Mazak_machine_model              = extract_selection_list(Mazak_machine_model_cells)
Mazak_main_spindle               = extract_selection_list(Mazak_main_spindle_cells)
Mazak_secondary_spindle          = extract_selection_list(Mazak_secondary_spindle_cells)
Mazak_second_spindle             = extract_selection_list(Mazak_second_spindle_cells)

metal_recycling_types_list    = extract_selection_list(metal_recycling_types_cells)
metal_recycling_emission_list = extract_emission_list(metal_recycling_emission_cells)

packaging_types_list          = extract_selection_list(packaging_types_cells)
packaging_box_frame_list      = extract_emission_list(packaging_box_frame_cells)

Indicator_GHG   = extract_selection_list(Indicator_GHG_cells)
GHG_values      = extract_selection_list(GHG_values_cells)
GWP_for_GHG     = extract_selection_list(GWP_for_GHG_cells)

Disassembly_by_Industry = extract_selection_list(Disassembly_by_Industry_cells)
Usage_type = extract_selection_list(Usage_type_cells)
Facilities = extract_selection_list(Facilities_cells)
Process = extract_selection_list(Process_cells)

Usage_category = extract_selection_list(Usage_category_cells)
Usage_electronic = extract_selection_list(Usage_electronic_cells)
Usage_electronic_ef = extract_selection_list(Usage_electronic_ef_cells)
Usage_energy = extract_selection_list(Usage_energy_cells)
Usage_energy_ef = extract_selection_list(Usage_energy_ef_cells)
Usage_Consumables = extract_selection_list(Usage_Consumables_cells)
Usage_Consumables_ef = extract_selection_list(Usage_Consumables_ef_cells)
Usage_services = extract_selection_list(Usage_services_cells)
Usage_services_ef = extract_selection_list(Usage_services_ef_cells)

End_of_Life_Activities = extract_selection_list(End_of_Life_Activities_cells)
End_of_Life_ef         = extract_selection_list(End_of_Life_ef_cells)

Assembly_modes  = extract_selection_list(Assembly_modes_cells)
Assembly_ef     = extract_selection_list(Assembly_ef_cells)

Waste_mode  = extract_selection_list(Waste_mode_cells)
Waste_ef    = extract_selection_list(Waste_ef_cells)
# --- LOOKUP DICTIONARIES FOR FAST CALCULATION ---
van_lookup = dict(zip(van_mode_list, van_emission_list))
hgv_lookup = dict(zip(HGV_mode_list, HGV_emission_list))
hgv_refrig_lookup = dict(zip(HGV_r_mode_list, HGV_r_emission_list))
freight_flight_lookup = dict(zip(Freight_Flight_mode_list, Freight_Flight_emission_list))
rail_lookup           = dict(zip(Rail_mode_list, Rail_emission_list))
sea_tanker_lookup     =dict(zip(Sea_Tanker_mode_list,Sea_Tanker_emission_list))
cargo_ship_lookup     =dict(zip(Cargo_ship_mode_list,Cargo_ship_emission_list))
waste_lookup = dict(zip(Waste_mode,Waste_ef))
Assembly_lookup = dict(zip(Assembly_modes,Assembly_ef))

# ---- Transport configuration for all modes ----

TRANSPORT_CONFIG = {
    # ROAD – VANS
    "van": {
        "classes": van_list,  # B2:B5
        "variants": {
            "diesel": van_diesel_list,   # D2:D5
            "petrol": van_petrol_list,   # E2:E5
            "battery": van_battery_list  # F2:F5
        }
    },

    # ROAD – HGV (articulated or main HGV group)
    "hgv": {
        "classes": HGV_diesel_list,  # B8:B15
        "variants": {
            "0% laden":  HGV_0_diesel_list,   # D8:D15
            "50% laden": HGV_50_diesel_list,  # E8:E15
            "100% laden": HGV_100_diesel_list # F8:F15
        }
    },

    # ROAD – HGV (rigid) – if you want a separate mode
    "hgv_rigid": {
        "classes": HGV_R_diesel_list,  # B17:B24
        "variants": {
            "0% laden":  HGV_R_0_diesel_list,
            "50% laden": HGV_R_50_diesel_list,
            "100% laden": HGV_R_100_diesel_list,
        }
    },

    # AIR
    "flight": {
        "classes": flight_list,  # B27:B29
        "variants": {
            "international": Int_flight_list,   # D27 (long-haul int.)
            "less_1500km":  L_1500_flight_list, # E27:E29
            "more_1500km":  M_1500_flight_list  # F27:F29
        }
    },

    # RAIL
    "rail": {
        "classes": rail_list,  # B31
        "variants": {
            "freight": freight_train_rail_list  # D31
        }
    },

    # SEA / CARGO SHIPS EXAMPLE – you can extend these patterns
    # One example using cargo ships with different cargo types
    "sea_cargo": {
        "classes": cargo_list,  # G60:G64
        "variants": {
            "bulk":      bulk_value_list,
            "general":   general_value_list,
            "container": container_value_list,
            "vehicle":   vehicle_value_list,
            "roro":      roro_value_list,
        }
    },

    # You can add more modes here, e.g. "sea_crude_tanker", "sea_products_tanker"
    # mapping crude_tanker_list -> crude_value_list, etc.
}

def tonkm1000_to_kgkm_emission(ef_per_1000kg_km: float, mass_kg: float, distance_km: float) -> float:
    # Excel EF is kgCO2e per 1000kg·km
    ef_per_kg_km = float(ef_per_1000kg_km) / 1000.0
    return ef_per_kg_km * mass_kg * distance_km


# ---------- Machine Optimization ("Pareto suggestion") ----------

# SPHERE builds `allowed_machine_substitutions` so CNC machines can swap
# between similar CNCs, others can't, then checks which emits less for the
# same machining time and weight. :contentReference[oaicite:17]{index=17}
non_substitutable = [
    "Surface Grinder (S.G)",
    "Wire Cut (W/C)",
    "Rokou Roku Micro Fine Precision Machine Center"
]
cnc_names = [
    "CNC (RB8)",
    "CNC (YCM4)",
    "CNC (YCM1)"
]
cnc_machines = [m for m in machine_value_list if m in cnc_names]

allowed_machine_substitutions = {}
for m in machine_value_list:
    if m in non_substitutable:
        allowed_machine_substitutions[m] = []
    elif m in cnc_machines:
        allowed_machine_substitutions[m] = [x for x in cnc_machines if x != m]
    else:
        allowed_machine_substitutions[m] = []

def suggest_optimal_machine(
    current_machine: str,
    machining_time_min: float,
    weight_after_kg: float,
    weight_after_qty: int,
    manufacturing_country: str,
):
    """
    Mirrors SPHERE.suggest_optimal_machine():
      - compute emission of current machine
      - test allowed substitutes
      - return best machine + its emission. :contentReference[oaicite:18]{index=18}
    """
    if manufacturing_country not in country_list:
        raise ValueError("Unknown manufacturing_country")
    cidx = country_list.index(manufacturing_country)

    def machine_emission(machine_name: str) -> float:
        if machine_name not in machine_value_list:
            return float("inf")
        midx = machine_value_list.index(machine_name)
        energy = float(specific_machine_energy_use_list[midx])
        electricity_emission_factor = float(electricity_list[cidx])
        return (
            energy
            * weight_after_kg
            * weight_after_qty
            * (machining_time_min / 60.0)
            * electricity_emission_factor
        )

    current_em = machine_emission(current_machine)
    best_machine = current_machine
    best_emission = current_em

    for alt in allowed_machine_substitutions.get(current_machine, []):
        alt_em = machine_emission(alt)
        if alt_em < best_emission:
            best_emission = alt_em
            best_machine = alt

    return {
        "current_machine": current_machine,
        "current_emission": current_em,
        "best_machine": best_machine,
        "best_emission": best_emission,
        "reduction": current_em - best_emission
    }


# ---------- Sourcing Optimization (material+country) ----------

def get_best_material_and_country(
    total_weight_kg: float,
    user_country: str,
    current_material: Optional[str] = None,
):
    """
    SPHERE iterates over every [material, source_country] pair and calculates:
      material_emission + transport_emission,
    where transport is modeled as shipping material from source_country
    to user_country using a 0.01 CO2e per ton-km factor. :contentReference[oaicite:19]{index=19}

    We expose this for Flutter to show "best source recommendation".
    """
    best_choice = None
    min_total_emission = float("inf")

    # Which materials to consider:
    if current_material:
        test_materials = [current_material] if any(
            current_material.strip().lower() == m.strip().lower() for m in material_list
        )else material_list
    else:
        test_materials = material_list

    for mat in test_materials:
        for cidx, src_country in enumerate(country_list):

            # material emission factor for (mat, src_country)
            override = custom_emission_factors.get("materials", {}).get(mat)
            if override is not None:
                material_factor = float(override)
            else:
                material_lists = {
                    "Steel": steel_list,
                    "Aluminium": aluminium_list,
                    "Cement": cement_list,
                    "Plastic": plastic_list,
                    "Carbon Fiber": carbon_fiber_list,
                }
                base_list = material_lists[mat]
                try:
                    material_factor = float(base_list[cidx])
                except Exception:
                    continue

            #skip if can't geocode either end
            if (
                user_country not in country_coords_cache
                or src_country not in country_coords_cache
                or country_coords_cache[user_country] is None
                or country_coords_cache[src_country] is None
            ):
                continue

            distance_km = geodesic(
                country_coords_cache[src_country],
                country_coords_cache[user_country]
            ).km

            transport_emission_factor = 0.01  # ton-km factor (same as SPHERE) :contentReference[oaicite:20]{index=20}
            transport_emission = total_weight_kg * distance_km * transport_emission_factor / 1000.0

            # total for this combo
            total_emission = material_factor * total_weight_kg + transport_emission

            if total_emission < min_total_emission:
                min_total_emission = total_emission
                best_choice = {
                    "material": mat,
                    "source_country": src_country,
                    "material_emission_factor": material_factor,
                    "material_emission_total": material_factor * total_weight_kg,
                    "transport_emission_total": transport_emission,
                    "distance_km": distance_km,
                    "total_emission": total_emission,
                }

    return best_choice

#Transportation calculation helper
def calculate_transport_single(mode, vehicle_class, variant, distance_km, mass_tonnes=None):
    if mode not in TRANSPORT_CONFIG:
        raise HTTPException(status_code=400, detail="Unknown transport mode")

    config = TRANSPORT_CONFIG[mode]

    # Validate class
    if vehicle_class not in config["classes"]:
        raise HTTPException(status_code=400, detail=f"Unknown class for mode '{mode}'")

    class_index = config["classes"].index(vehicle_class)

    # Validate variant
    if variant not in config["variants"]:
        raise HTTPException(status_code=400, detail=f"Unknown variant for mode '{mode}'")

    ef_list = config["variants"][variant]
    ef = float(ef_list[class_index])     # EF for that class & variant

    # HGV mass factor
    if mode.startswith("hgv") and mass_tonnes is not None:
        ef *= mass_tonnes

    return ef * distance_km



# --------- 5. API REQUEST / RESPONSE MODELS ---------------------------------#


class MaterialEmissionReq(BaseModel):
    material: str  #e.g.Steel,Aluminum,etc.
    country: str   #e.g.Singapore,Malaysia,China,etc.
    mass_kg: float #mass from flutter ui

class MachineEmissionsReq(BaseModel):
    machine_model: str
    country: str
    time_operated_hr: float

class MachineEmissionsAdvancedReq(BaseModel):
    machine_model: str
    country: str
    time_operated_main_hr: float = 0.0
    time_operated_second_hr: float = 0.0
    time_operated_secondary_hr: float = 0.0


class TransportCalcRequest(BaseModel):
    mode: str            # "van", "hgv", "flight", "rail", "sea_cargo", ...
    vehicle_class: str   # one entry from the relevant classes list
    variant: str         # fuel/load/distance-band: e.g. "diesel", "0% laden", "international"
    distance_km: float   # distance
    transport_type: str
    # mass_tonnes is optional – only needed if your EF is in kgCO2e per ton-km
    mass_tonnes: Optional[float] = None

class TonKmRequest(BaseModel):
    transport_type: str
    distance_km: float
    mass_kg: float

class TransportRow(BaseModel):
    mode: str
    vehicle_class: str
    variant: str
    distance_km: float
    mass_tonnes: Optional[float] = None

class TransportTableRequest(BaseModel):
    rows: List[TransportRow]

class FugitiveEmissionFromExcelRequest(BaseModel):
    ghg_name: str               # GHG name from Excel (e.g. "CO2", "R134a")
    total_charged_amount_kg: float
    current_charge_amount_kg: float

class ProfileSaveRequest(BaseModel):
    username: str
    profile_name: str
    description: Optional[str] = ""
    data: dict  # full UI state

class ProfileRenameRequest(BaseModel):
    username: str
    old_name: str
    new_name: str

class NewsArticle(BaseModel):
    title: str
    description: Optional[str]
    url: str
    source: str
    published_at: Optional[str]
    image_url: Optional[str]


class NewsResponse(BaseModel):
    total_results: int
    articles: List[NewsArticle]

class UserSignupRequest(BaseModel):
    username: str
    password: str

class UserLoginRequest(BaseModel):
    username: str
    password: str

class DistanceOnlyRequest(BaseModel):
    transport_type: str
    distance_km: float

class MaterialEmissionAdvancedReq(BaseModel):
    material: str
    country: str

    # masses (kg)
    total_material_purchased_kg: float   # total purchased for that material type
    mass_of_material_recycled_kg: float = 0.0  # in-house recycled mass (optional)

    # optional overrides (if you don’t send them, backend uses Excel EF for regular,
    # and 0 for custom/recycled/internal unless you provide)
    custom_ef_of_material: float         # EF for recycled material (kgCO2e/kg)
    custom_internal_ef: float          # EF for in-house recycling (kgCO2e/kg)

class ExcelUpdateCellRequest(BaseModel):
    sheet: str
    row: int
    col: int
    value: Any

class ExcelUpdateCellsRequest(BaseModel):
    sheet: str
    updates: List[ExcelUpdateCellRequest]
    
class WasteRequest(BaseModel):
    waste_mode: str
    mass_kg: float

class AssemblyRequest(BaseModel):
    Assembly_mode: str
    Power: float
# --------- 6. FASTAPI APP + ENDPOINTS ---------------------------------------#

app = FastAPI(title="SPHERE Backend API (Flutter)")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- NewsAPI configuration ---
NEWSAPI_KEY = os.getenv("NEWSAPI_KEY", "adfd0f78d5304adea0d4490623f32aec")
NEWSAPI_ENDPOINT = "https://newsapi.org/v2/everything"

@app.get("/meta/machines")
def get_machinedata():
    """
    machine information
    """
    return{
        "machines":machine_value_list
    } 
    
@app.get("/meta/material type")
def get_materialdata():
    """
    material data 
    """
    return {
        "material types":material_list
    }

@app.get("/meta/Grid_intensity_of_all_countries")
def Grid_intensity_of_all_countries():
    """
    Grid intensity for different countries
    """
    return {
        "countries": country_list,
        "grid_intensity": electricity_list,
    }

@app.get("/meta/transport(cargotype)")
def get_transport_types():
    """
    different transport types
    """
    return {
        "transport_types": transport_list
    }

@app.get("/meta/GWP of GHG_gases")
def getGWPvalues():
    """
    Values of the Indicator, GHG, and GWP values
    """
    return{
        "indicator": Indicator_GHG,
        "GHG": GHG_values,
        "GWP": GWP_for_GHG
        
    } 

@app.get("/meta/options")
def get_options():
    """
    Flutter calls this to populate dropdowns:
      - countries
      - materials
      - machine types
      - packaging types
      - recycling types
      - Grid intensity
      - transport types
      - Values of the Indicator, GHG, and GWP values
      - Process
      - Facilities
      - Usage Types
      - Disassembly by Industry
      - Machine Types
      - Transport types
      - Usage Categories
      - Assembly
    """
    return {
        "countries": country_list,
        "materials": material_list,
        "machines": machine_value_list,
        "packaging_types": packaging_types_list,
        "recycling_types": metal_recycling_types_list,
        "grid_intensity": electricity_list,
        "transport_types": transport_list,
        "indicator": Indicator_GHG,
        "GHG": GHG_values,
        "GWP": GWP_for_GHG,
        "process": Process,
        "facilities": Facilities,
        "usage_types": Usage_type,
        "disassembly_by_industry": Disassembly_by_Industry,
        "machine_type": machine_types,
        "YCM_types":YCM_machine_model,
        "Amada_types":Amada_machine_model,
        "Mazak_types":Mazak_machine_model,
        "Van_mode":van_mode_list,
        "Van_emissions":van_emission_list,
        "HGV_mode":HGV_mode_list,
        "HGV_emissions":HGV_emission_list,
        "HGV_r_mode":HGV_r_mode_list,
        "HGV_r_emissions":HGV_r_emission_list,
        "Freight_flight_modes":Freight_Flight_mode_list,
        "Freight_flight_emissions":Freight_Flight_emission_list,
        "Rail_mode":Rail_mode_list,
        "Rail_emissions":Rail_emission_list,
        "Sea_Tanker_mode":Sea_Tanker_mode_list,
        "Sea_Tanker_emissions":Sea_Tanker_emission_list,
        "Cargo_ship_mode":Cargo_ship_mode_list,
        "Cargo_ship_emissions":Cargo_ship_emission_list,
        "Usage_categories": Usage_category,
        "Usage_electronics": Usage_electronic,
        "Usage_electronic_ef": Usage_electronic_ef,
        "Usage_energy": Usage_energy,
        "Usage_energy_ef":Usage_energy_ef,
        "Usage_consumables": Usage_Consumables,
        "Usage_consumables_ef": Usage_Consumables_ef,
        "Usage_services": Usage_services,
        "Usage_services_ef": Usage_services_ef,
        "End_of_Life_Activities": End_of_Life_Activities,
        "End_of_Life_ef": End_of_Life_ef,
        "Machine_brands": Machine_brands,
        "Assembly_modes": Assembly_modes,
        "Assembly_ef"   : Assembly_ef,
        "Waste_mode"    : Waste_mode,
        "Waste_ef"      : Waste_ef
    }

@app.get("/meta/transport/config")
def get_transport_config():
    """
    Return all transport modes, and for each mode:
      - classes
      - variants (fuel/load/band)
    """
    modes = list(TRANSPORT_CONFIG.keys())
    classes_by_mode = {
        mode: cfg["classes"]
        for mode, cfg in TRANSPORT_CONFIG.items()
    }
    variants_by_mode = {
        mode: list(cfg["variants"].keys())
        for mode, cfg in TRANSPORT_CONFIG.items()
    }

    return {
        "modes": modes,
        "classes_by_mode": classes_by_mode,
        "variants_by_mode": variants_by_mode
    }

@app.get("/meta/machines_type")
def get_machinetypes():
    '''
    -machine info like milling etc.
    '''
    return {
        "machine_type": machine_types,
    }

@app.get("/meta/YCM_model")
def get_machinetypes_YCM():
    '''
    -YCM machine models
    -main spindle(kW)
    -Sub Spindle(kW)
    '''
    return{
        "Machine Model": YCM_machine_model,
        "main_spindle": YCM_main_spindle,
        "Sub Spindle": YCM_sub_spindle
    }

@app.get("/meta/Amada_model")
def get_machinetypes_YCM():
    '''
    -Amada machine models
    -main spindle(kW)
    -Sub Spindle(kW)
    '''
    return{
        "Machine Model": Amada_machine_model,
        "main_spindle": Amada_main_spindle,
        "Sub Spindle": Amada_sub_spindle
    }

@app.get("/meta/Mazak_model")
def get_machinetypes_YCM():
    '''
    -Mazak machine models
    -main spindle(kW)
    -Secondary Spindle(kW)
    '''
    return{
        "Machine Model":Mazak_machine_model,
        "main_spindle":Mazak_main_spindle,
        "Secondary Spindle":Mazak_secondary_spindle,
        "Second Spindle":Mazak_second_spindle,
    }

@app.get("/news/sustainability", response_model=NewsResponse)
def get_sustainability_news():
    """
    Returns sustainability & carbon-emissions–related news articles
    for the Flutter UI.
    """
    if not NEWSAPI_KEY or NEWSAPI_KEY == "YOUR_NEWSAPI_KEY_HERE":
        raise HTTPException(
            status_code=500,
            detail="NEWSAPI_KEY is not set in environment variables."
        )

    params = {
        "q": 'sustainability OR "carbon emissions" OR "climate change"',
        "apiKey": NEWSAPI_KEY,
        "language": "en",
        "pageSize": 50,
        "sortBy": "publishedAt",
    }

    try:
        resp = requests.get(NEWSAPI_ENDPOINT, params=params, timeout=10)
    except requests.RequestException:
        raise HTTPException(status_code=502, detail="Unable to reach NewsAPI.")

    if resp.status_code != 200:
        raise HTTPException(
            status_code=resp.status_code,
            detail=resp.json().get("message", "NewsAPI error")
        )

    data = resp.json()
    articles = []

    for art in data.get("articles", []):
        articles.append(NewsArticle(
            title=art.get("title") or "",
            description=art.get("description"),
            url=art.get("url") or "",
            source=(art.get("source") or {}).get("name", ""),
            published_at=art.get("publishedAt"),
            image_url=art.get("urlToImage"),
        ))

    return NewsResponse(
        total_results=data.get("totalResults", 0),
        articles=articles
    )

@app.get("/profiles/{profile_name}")
def get_profile(profile_name: str, username: str):
    all_profiles = load_profiles()
    bucket = all_profiles.get(username, {})

    if profile_name not in bucket:
        raise HTTPException(status_code=404, detail="Profile not found")

    return bucket[profile_name]


@app.get("/profiles")
def list_profiles(username: str):
    all_profiles = load_profiles()
    bucket = all_profiles.get(username, {})
    return {"username": username, "profiles": list(bucket.keys())}

@app.post("/calculate/material_emission")
def calculate_material_emissions(req:MaterialEmissionReq): #req: is the name of the input the fastapi endpoint receives.
    if req.country not in country_list:
        raise HTTPException(status_code=400,detail="Country not found in Data Sheet") #find the row for the given country Data Sheet.
    cidx=country_list.index(req.country)
    materials={
        "Steel":steel_list,
        "Aluminum":aluminium_list,
        "Cement":cement_list,
        "Plastic":plastic_list,
        "Carbon Fiber":carbon_fiber_list
    }
    if req.material not in materials:
        raise HTTPException(status_code=400,detail="Material not supported for calculation")
    ef_list = materials[req.material]

    try:
        emisson_factor=float(ef_list[cidx])#emission factor for that country and the material.
    except:
        raise HTTPException(status_code=500,detail="Emission factor missing in Data Sheet.")
    
    calculated_emission=emisson_factor*req.mass_kg

    return{
        "country":req.country, "material":req.material, "mass_kg":req.mass_kg, "material_emission_factor":emisson_factor,
        "materialacq_emission":calculated_emission
    }

@app.post("/calculate/material_emission_recycled")
def calculate_material_emissions_advanced(req: MaterialEmissionAdvancedReq):

    # --- validate country ---
    if req.country not in country_list:
        raise HTTPException(status_code=400, detail="Country not found in Data Sheet")
    cidx = country_list.index(req.country)

    # --- validate material + get regular EF from Excel ---
    materials = {
        "Steel": steel_list,
        "Aluminum": aluminium_list,
        "Cement": cement_list,
        "Plastic": plastic_list,
        "Carbon Fiber": carbon_fiber_list
    }
    if req.material not in materials:
        raise HTTPException(status_code=400, detail="Material not supported for calculation")

    # --- validate masses ---
    total_purchased = float(req.total_material_purchased_kg)
    recycled_inhouse_mass = float(req.mass_of_material_recycled_kg or 0.0)

    if total_purchased <= 0:
        raise HTTPException(status_code=400, detail="total_material_purchased_kg must be > 0")

    if recycled_inhouse_mass < 0:
        raise HTTPException(status_code=400, detail="mass_of_material_recycled_kg cannot be negative")

    # --- custom factors (optional) ---
    recycled_ef = float(req.custom_ef_of_material) 
    internal_ef = float(req.custom_internal_ef)

    # recycled materials emissions = custom_ef_material × total_mass
    recycled_materials_emissions = recycled_ef * total_purchased

    # in-house recycling emissions = custom_internal_ef × mass_recycled
    in_house_recycling_emissions = internal_ef * recycled_inhouse_mass

    # total emissions = recycled materials emissions + normal material emissions
    total_material_emissions = recycled_materials_emissions + in_house_recycling_emissions

    return {
        "country": req.country,
        "material": req.material,

        "total_material_purchased_kg": total_purchased,
        "mass_of_material_recycled_kg": recycled_inhouse_mass,

        "custom_ef_of_material_kgco2e_per_kg": recycled_ef,
        "custom_internal_ef_kgco2e_per_kg": internal_ef,

        "recycled_materials_emissions": recycled_materials_emissions,
        "in_house_recycling_emissions": in_house_recycling_emissions,

        "total_material_emissions": total_material_emissions,
    }

@app.get("/meta/machining/mazak")
def get_mazak_list():
    return {
        "Mazak_machine_model": Mazak_machine_model
    }

@app.get("/meta/machining/countries")
def get_countries():
    return {
        "countries": country_list
    }

@app.post("/calculate/machining_advanced")
def calculate_machine(req:MachineEmissionsAdvancedReq):
    if req.country not in country_list:
        raise HTTPException(
            status_code=400,
            detail=f"Country '{req.country}' not found in grid intensity list."
        )
    cidx = country_list.index(req.country)
    grid_intensity = float(electricity_list[cidx])  # kg CO2e per kWh 

    # 2) Get Mazak main spindle power from selected machine
    if req.machine_model not in Mazak_machine_model:
        raise HTTPException(
            status_code=400,
            detail=f"Mazak machine '{req.machine_model}' not found."
        )
    midx = Mazak_machine_model.index(req.machine_model)
    main_spindle_kw = float(Mazak_main_spindle[midx])  # kW
    secondary_spindle_kw = float(Mazak_secondary_spindle[midx])
    second_spindle_kw = float(Mazak_second_spindle[midx])

    t_main = float(req.time_operated_main_hr)
    t_second = float(req.time_operated_second_hr)
    t_secondary = float(req.time_operated_secondary_hr)

    if t_main < 0 or t_second < 0 or t_secondary < 0:
        raise HTTPException(status_code=400, detail="Time operated cannot be negative.")

    emissions_main = main_spindle_kw * grid_intensity * t_main
    emissions_second = second_spindle_kw * grid_intensity * t_second
    emissions_secondary = secondary_spindle_kw * grid_intensity * t_secondary

    total_emissions = emissions_main + emissions_second + emissions_secondary

    return {
        "country": req.country,
        "machine_model": req.machine_model,
        "grid_intensity": grid_intensity,

        "power_drawed_main_kw": main_spindle_kw,
        "power_drawed_second_kw": second_spindle_kw,
        "power_drawed_secondary_kw": secondary_spindle_kw,

        "time_operated_main_hr": t_main,
        "time_operated_second_hr": t_second,
        "time_operated_secondary_hr": t_secondary,

        "emissions_main": emissions_main,
        "emissions_second": emissions_second,
        "emissions_secondary": emissions_secondary,
        "total_emissions": total_emissions
    }

@app.post("/calculate/machine_power_emission")
def calculate_machine_power_emission(req:MachineEmissionsReq):
    if req.country not in country_list:
        raise HTTPException(
            status_code=400,
            detail=f"Country '{req.country}' not found in grid intensity list."
        )
    cidx = country_list.index(req.country)
    grid_intensity = float(electricity_list[cidx])  # kg CO2e per kWh 

    # 2) Get Mazak main spindle power from selected machine
    if req.machine_model not in Mazak_machine_model:
        raise HTTPException(
            status_code=400,
            detail=f"Mazak machine '{req.machine_model}' not found."
        )
    midx = Mazak_machine_model.index(req.machine_model)
    main_spindle_kw = float(Mazak_main_spindle[midx])  # kW

    power_drawed = main_spindle_kw                 # only main spindle for now
    time_operated = req.time_operated_hr          # hours
    emissions = power_drawed * grid_intensity * time_operated

    return {
        "country": req.country,
        "machine_model": req.machine_model,
        "time_operated_hr": time_operated,
        "power_drawed_kw": power_drawed,
        "grid_intensity": grid_intensity,
        "emissions": emissions,       # kg CO2e
    }

@app.post("/calculate/waste")
def calculate_waste(req: WasteRequest):

    if req.waste_mode not in waste_lookup:
        raise HTTPException(status_code=400, detail="Invalid waste mode")

    ef = waste_lookup[req.waste_mode]  # kgCO2e per kg
    emissions = req.mass_kg * ef

    return {
        "category": "Waste",
        "waste_mode": req.waste_mode,
        "mass_kg": req.mass_kg,
        "ef_kgco2e_per_kg": float(ef),
        "emissions_kgco2e": round(emissions, 4)
    }

@app.post("/calculate/assembly")
def calculate_waste(req: AssemblyRequest):

    if req.Assembly_mode not in Assembly_lookup:
        raise HTTPException(status_code=400, detail="Invalid assembly mode")

    ef = Assembly_lookup[req.Assembly_mode]  # kgCO2e per kg
    emissions = req.Power * ef

    return {
        "category": "Assembly",
        "waste_mode": req.Assembly_mode,
        "Power": req.Power,
        "ef_kgco2e_per_kg": float(ef),
        "emissions_kgco2e": round(emissions, 4)
    }

@app.post("/calculate/transport_table") #for tables
def calculate_transport_table(req: TransportTableRequest):
    total = 0.0
    row_results = []

    for row in req.rows:
        emission = calculate_transport_single(
            row.mode,
            row.vehicle_class,
            row.variant,
            row.distance_km,
            row.mass_tonnes
        )
        total += emission

        row_results.append({
            "mode": row.mode,
            "vehicle_class": row.vehicle_class,
            "variant": row.variant,
            "distance_km": row.distance_km,
            "mass_tonnes": row.mass_tonnes,
            "emission_kgco2e": emission
        })

    return {
        "total_emissions": total,
        "rows": row_results
    }

@app.post("/calculate_transport_emission")
def calculate_transport_emission(data: TransportCalcRequest):

    mode = data.transport_type
    distance = data.distance_km

    # VAN
    if mode in van_lookup:
        ef = van_lookup[mode]
        return {
            "category": "Van",
            "mode": mode,
            "distance_km": distance,
            "emission_factor": ef,
            "total_emission": ef * distance    
        }

    # HGV NORMAL
    if mode in hgv_lookup:
        ef = hgv_lookup[mode]
        return {
            "category": "HGV",
            "mode": mode,
            "distance_km": distance,
            "emission_factor": ef,
            "total_emission": ef * distance
        }

    # HGV REFRIGERATED
    if mode in hgv_refrig_lookup:
        ef = hgv_refrig_lookup[mode]
        return {
            "category": "HGV Refrigerated",
            "mode": mode,
            "distance_km": distance,
            "emission_factor": ef,
            "total_emission": ef * distance
        }

    return {"error": "Invalid transport type."}

@app.post("/calculate/hgv")
def calculate_hgv(req: DistanceOnlyRequest):
    transport_type = req.transport_type
    distance = req.distance_km

    if transport_type not in hgv_lookup:
        raise HTTPException(status_code=400, detail="Invalid HGV transport type")

    ef = float(hgv_lookup[transport_type])

    return {
        "category": "HGV",
        "transport_type": transport_type,
        "distance_km": distance,
        "emission_factor": ef,
        "total_transport_type_emission": ef * distance
    }

@app.post("/calculate/hgv_r")
def calculate_hgv_r(req: DistanceOnlyRequest):
    transport_type = req.transport_type
    distance = req.distance_km

    # HGV R (your HGV_r_mode_list / HGV_r_emission_list)
    if transport_type not in hgv_refrig_lookup:  # this dict is built from HGV_r_mode_list + HGV_r_emission_list
        raise HTTPException(status_code=400, detail="Invalid HGV R transport type")

    ef = float(hgv_refrig_lookup[transport_type])

    return {
        "category": "HGV R",
        "transport_type": transport_type,
        "distance_km": distance,
        "emission_factor": ef,
        "total_transport_type_emission": ef * distance
    }

@app.post("/calculate/freight_flight")
def calculate_freight_flight(req: TonKmRequest):
    if req.transport_type not in freight_flight_lookup:
        raise HTTPException(status_code=400, detail="Invalid freight flight type")

    ef_1000 = freight_flight_lookup[req.transport_type]
    total = tonkm1000_to_kgkm_emission(ef_1000, req.mass_kg, req.distance_km)

    return {
        "category": "Freight Flight",
        "transport_type": req.transport_type,
        "distance_km": req.distance_km,
        "mass_kg": req.mass_kg,
        "ef_kgco2e_per_1000kg_km": float(ef_1000),
        "ef_kgco2e_per_kg_km": float(ef_1000) / 1000.0,
        "total_transport_type_emission": total
    }

@app.post("/calculate/rail_sheet")
def calculate_rail_sheet(req: TonKmRequest):
    if req.transport_type not in rail_lookup:
        raise HTTPException(status_code=400, detail="Invalid rail type")

    ef_1000 = rail_lookup[req.transport_type]
    total = tonkm1000_to_kgkm_emission(ef_1000, req.mass_kg, req.distance_km)

    return {
        "category": "Rail",
        "transport_type": req.transport_type,
        "distance_km": req.distance_km,
        "mass_kg": req.mass_kg,
        "ef_kgco2e_per_1000kg_km": float(ef_1000),
        "ef_kgco2e_per_kg_km": float(ef_1000) / 1000.0,
        "total_transport_type_emission": total
    }

@app.post("/calculate/sea_tanker")
def calculate_sea_tanker(req: TonKmRequest):
    if req.transport_type not in sea_tanker_lookup:
        raise HTTPException(status_code=400, detail="Invalid sea tanker type")

    ef_1000 = sea_tanker_lookup[req.transport_type]  # kgCO2e / 1000kg·km
    total = tonkm1000_to_kgkm_emission(ef_1000, req.mass_kg, req.distance_km)

    return {
        "category": "Sea Tanker",
        "transport_type": req.transport_type,
        "distance_km": req.distance_km,
        "mass_kg": req.mass_kg,
        "ef_kgco2e_per_1000kg_km": float(ef_1000),
        "ef_kgco2e_per_kg_km": float(ef_1000) / 1000.0,
        "total_transport_type_emission": total
    }

@app.post("/calculate/cargo_ship") 
def calculate_cargo_ship(req: TonKmRequest):
    if req.transport_type not in cargo_ship_lookup:
            raise HTTPException(status_code=400, detail="Invalid cargo ship type")
         
    ef_1000 = cargo_ship_lookup[req.transport_type] 
    total = tonkm1000_to_kgkm_emission(ef_1000, req.mass_kg, req.distance_km)
    return { 
        "category": "Cargo Ship",
        "transport_type": req.transport_type, 
        "distance_km": req.distance_km, 
        "mass_kg": req.mass_kg, 
        "ef_kgco2e_per_1000kg_km": float(ef_1000), 
        "ef_kgco2e_per_kg_km": float(ef_1000) / 1000.0, 
        "total_transport_type_emission": total }

@app.post("/calculate/van")
def calculate_van(req: dict):

    transport_type = req["transport_type"]       # dropdown selection
    distance = req["distance_km"]                # entered by user

    if transport_type not in van_lookup:
        return {"error": "Invalid van transport type"}

    ef = van_lookup[transport_type]

    return {
        "transport_type": transport_type,
        "distance_km": distance,
        "emission_factor": ef,
        "total_transport_type_emission": ef * distance
    }

@app.post("/calculate/fugitive_emissions")
def calculate_fugitive_emissions(req: FugitiveEmissionFromExcelRequest):

    if req.ghg_name not in GHG_values:
        raise HTTPException(status_code=400, detail="GHG value is not found in GWP sheet")
    ghgidx = GHG_values.index(req.ghg_name)
        
    try:
        gwp = float(GWP_for_GHG[ghgidx])
    except Exception:
        raise HTTPException(status_code=500, detail="GHG value is missing.")

    mass_released_kg = req.total_charged_amount_kg - req.current_charge_amount_kg
    if mass_released_kg < 0:
        mass_released_kg = 0

    emissions_kgco2e = gwp * mass_released_kg

    return {
        "ghg_name": req.ghg_name,
        "total_charged_amount_kg": req.total_charged_amount_kg,
        "current_charge_amount_kg": req.current_charge_amount_kg,
        "mass_of_ghg_released_kg": mass_released_kg,
        "gwp": gwp,
        "emissions_kgco2e": emissions_kgco2e
    }

@app.post("/profiles/save")
def save_profile(req: ProfileSaveRequest):
    all_profiles = load_profiles()
    bucket = get_user_profiles_bucket(all_profiles, req.username)

    bucket[req.profile_name] = {
        "description": req.description,
        "data": req.data
    }

    save_profiles(all_profiles)
    return {"status": "ok", "saved_profile": req.profile_name, "username": req.username}
    
@app.delete("/profiles/delete/{username}/{profile_name}")
def delete_profile(username: str, profile_name: str):
    all_profiles = load_profiles()
    bucket = all_profiles.get(username, {})

    if profile_name not in bucket:
        raise HTTPException(status_code=404, detail="Profile not found")

    del bucket[profile_name]
    all_profiles[username] = bucket
    save_profiles(all_profiles)

    return {"status": "deleted", "username": username, "profile_name": profile_name}


@app.post("/profiles/rename")
def rename_profile(req: ProfileRenameRequest):
    all_profiles = load_profiles()
    bucket = all_profiles.get(req.username, {})

    if req.old_name not in bucket:
        raise HTTPException(status_code=404, detail="Old profile name not found")

    if req.new_name in bucket:
        raise HTTPException(status_code=400, detail="New profile name already exists")

    bucket[req.new_name] = bucket[req.old_name]
    del bucket[req.old_name]

    all_profiles[req.username] = bucket
    save_profiles(all_profiles)

    return {"status": "renamed","username": req.username,"old_name": req.old_name,"new_name": req.new_name}


@app.post("/auth/signup")
def signup(req: UserSignupRequest):
    wb, ws = load_users_sheet()

    # prevent duplicate usernames
    if find_user_row(ws, req.username) is not None:
        raise HTTPException(status_code=409, detail="Username already exists")

    # append new user
    ws.append([req.username, hash_password(req.password)])
    wb.save(USERS_FILE)

    return {"status": "ok", "username": req.username}


@app.post("/auth/login")
def login_user(req: UserLoginRequest):
    wb, ws = load_users_sheet()

    row = find_user_row(ws, req.username)
    if row is None:
        raise HTTPException(status_code=401, detail="Invalid username or password")

    stored_hash = ws.cell(row=row, column=2).value
    if stored_hash != hash_password(req.password):
        raise HTTPException(status_code=401, detail="Invalid username or password")

    all_profiles = load_profiles()
    bucket = all_profiles.get(req.username, {})
    return {"status": "ok","username": req.username,"profiles": list(bucket.keys())}

@app.post("/excel/update_cell")
def excel_update_cell(req: ExcelUpdateCellRequest):
    if req.row < 1 or req.col < 1:
        raise HTTPException(status_code=400, detail="row/col must be >= 1")

    with portalocker.Lock(LOCK_PATH, timeout=10):
        wb = load_workbook(EXCEL_PATH, data_only=False)  # must be False for writing
        if req.sheet not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Sheet '{req.sheet}' not found")
        ws = wb[req.sheet]
        ws.cell(row=req.row, column=req.col).value = req.value
        _atomic_save_workbook(wb, EXCEL_PATH)

    return {"status": "ok"}

@app.post("/excel/update_cells")
def excel_update_cells(req: ExcelUpdateCellsRequest):
    with portalocker.Lock(LOCK_PATH, timeout=10):
        wb = load_workbook(EXCEL_PATH, data_only=False)
        if req.sheet not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Sheet '{req.sheet}' not found")
        ws = wb[req.sheet]

        for u in req.updates:
            if u.sheet != req.sheet:
                raise HTTPException(status_code=400, detail="All updates must use the same sheet")
            if u.row < 1 or u.col < 1:
                raise HTTPException(status_code=400, detail="row/col must be >= 1")
            ws.cell(row=u.row, column=u.col).value = u.value

        _atomic_save_workbook(wb, EXCEL_PATH)

    return {"status": "ok", "count": len(req.updates)}
