from itertools import filterfalse
import requests
import openpyxl
import tkinter as tk
from tkinter import ttk, messagebox
import sv_ttk
from pprint import pprint
import googlemaps
from geopy.distance import geodesic 
import matplotlib.pyplot as plt
import webbrowser
import folium
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut
import searoute as sr
import cartopy.crs as ccrs
import cartopy.feature as cfeature
import threading
from threading import Timer
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import sys
import subprocess
import os
import tempfile
import json
from tkinter import simpledialog
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from tkinter import filedialog
import tkinter.simpledialog
import csv
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas as pdf_canvas
import threading
import platform
import struct
import sys
import re

print(f"Python version: {sys.version}")
print(f"Architecture reported by platform: {platform.architecture()}")
print(f"Architecture based on struct size: {'64bit' if struct.calcsize('P') == 8 else '32bit'}")
print(f"Is Python 64bit: {sys.maxsize > 2**32}")
print(f"Executable path: {sys.executable}")

class ToolTip(object):
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tipwindow = None
        self.id = None
        self.x = self.y = 0
        widget.bind("<Enter>", self.show_tip)
        widget.bind("<Leave>", self.hide_tip)

    def show_tip(self, event=None):
        if self.tipwindow or not self.text:
            return
        x, y, cx, cy = self.widget.bbox("insert") if hasattr(self.widget, "bbox") else (0, 0, 0, 0)
        x = x + self.widget.winfo_rootx() + 25
        y = y + self.widget.winfo_rooty() + 20
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry("+%d+%d" % (x, y))
        label = tk.Label(tw, text=self.text, justify='left',
                         background="#ffffe0", relief='solid', borderwidth=1,
                         font=("tahoma", "8", "normal"))
        label.pack(ipadx=1)

    def hide_tip(self, event=None):
        tw = self.tipwindow
        self.tipwindow = None
        if tw:
            tw.destroy()

def get_user_data_path(filename):
    if sys.platform == "win32":
        base = os.path.join(os.environ.get("LOCALAPPDATA", os.path.expanduser("~")), "SPHERE")
    else:
        base = os.path.join(os.path.expanduser("~"), ".sphere")
    os.makedirs(base, exist_ok=True)
    return os.path.join(base, filename)

PROFILE_FILE = get_user_data_path("profiles.json")
EMISSION_FACTORS_FILE = get_user_data_path("custom_emission_factors.json")

if os.path.dirname(sys.executable) == tempfile.gettempdir():
    import tkinter as tk
    from tkinter import messagebox
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror("Update Error", "Please run the application from its original location, not from the temp folder.")
    sys.exit(1)

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

#https://github.com/Spmae-vrar/update-Version/blob/main/Instruction [Raw Link to View more Detailed Instruction on Auto Update on APP]
APP_VERSION = "1.0.2" #The Logic behind is always -1 less than version.txt so it will Prompt for update if the version.txt is +1 more than this number
API_KEY = 'AIzaSyBgXa2swWuYosJ2LQQRvpMWCmja_CqdGnA' #google maps API key, replace with your own
current_unit = "kg"

def ensure_default_files():
    try:
        if not os.path.exists(EMISSION_FACTORS_FILE):
            with open(EMISSION_FACTORS_FILE, "w", encoding="utf-8") as f:
                json.dump({}, f)
        if not os.path.exists(PROFILE_FILE):
            with open(PROFILE_FILE, "w", encoding="utf-8") as f:
                json.dump({}, f)
    except PermissionError as e:
        from tkinter import messagebox
        messagebox.showerror("Permission Error", f"Cannot create data files:\n{e}\n\n"
                                                "Please ensure you have write access to this location.")
        sys.exit(1)

# Global variables
total_emission_labels = []  # Global list to track total emission labels
transport_row_widgets = []  # Global list to track transport row widgets
business_travel_row_widgets=[] #Global list to track business travel row in UI
machine_process_row_widgets = []  # Global List to track machine process row widgets
machine_process_rows_state = {}  # Dictionary to save the state of machine process rows
pie_chart_debounce_id = None
pie_chart_canvas = None
transport_rows_state = {}
debounce_timer = None
parts_debounce_id = None
debounce_id = None
get_weight_after_total_global = None
restoring_profile = False
is_navigating_back = False
is_importing = False
previous_widget_values = {}
previous_transport_rows = []
previous_machine_parts_data = []
materials = ["Steel", "Aluminium", "Cement", "Plastic", "Carbon Fiber"] #Extend more as material increases!
current_part_index = 0
machine_parts_data = []

UPDATE_VERSION_URL = "https://raw.githubusercontent.com/Spmae-vrar/update-Version/refs/heads/main/version.txt"
UPDATE_EXE_URL = "https://github.com/Spmae-vrar/Upload-new-EXE/releases/download/1.0.2/SPHERE.zip" #Always +1 more than the version.txt file (Same for UPDATE_EXE Github)

def check_for_update_and_prompt(root):
    try:
        response = requests.get(UPDATE_VERSION_URL, timeout=5)
        if response.status_code == 200:
            latest_version = response.text.strip()
            if latest_version != APP_VERSION:
                if messagebox.askyesno("Update Available", f"A new version ({latest_version}) is available. Update now?"):
                    download_and_replace_exe(root)
    except Exception as e:
        print(f"Update check failed: {e}")

def download_and_replace_exe(root):
    import tempfile
    import shutil
    import zipfile
    import struct
    
    try:
        response = requests.get(UPDATE_EXE_URL, stream=True)
        print("Update download status code:", response.status_code)
        if response.status_code == 200:
            # Download to a zip file
            tmp_zip = os.path.join(tempfile.gettempdir(), "SPHERE_update.zip")
            with open(tmp_zip, "wb") as f:
                for chunk in response.iter_content(1024 * 1024):
                    f.write(chunk)
            
            # Extract the zip file
            tmp_extract_dir = os.path.join(tempfile.gettempdir(), "SPHERE_extract")
            if os.path.exists(tmp_extract_dir):
                shutil.rmtree(tmp_extract_dir)
            os.makedirs(tmp_extract_dir)
            
            with zipfile.ZipFile(tmp_zip, 'r') as zip_ref:
                zip_ref.extractall(tmp_extract_dir)
            
            # Find the .exe file in the extracted directory
            exe_files = []
            for root_dir, dirs, files in os.walk(tmp_extract_dir):
                for file in files:
                    if file.lower().endswith('.exe'):
                        exe_files.append(os.path.join(root_dir, file))
            
            if not exe_files:
                messagebox.showerror("Update Error", "No executable found in update package")
                return
                
            tmp_exe = exe_files[0] 
            
            # Verify the executable is 64-bit (PE header check)
            try:
                with open(tmp_exe, 'rb') as f:
                    # Skip DOS header
                    f.seek(60)
                    # Get offset to PE header
                    offset = struct.unpack('<I', f.read(4))[0]
                    # Go to PE header
                    f.seek(offset)
                    # Check PE signature
                    if f.read(4) != b'PE\x00\x00':
                        messagebox.showerror("Update Error", "Invalid executable format")
                        return
                    # Skip COFF header
                    f.seek(offset + 4 + 20)
                    # Read optional header magic to determine architecture
                    magic = struct.unpack('<H', f.read(2))[0]
                    if magic == 0x20b:  # PE32+ (64-bit)
                        print("64-bit executable detected, proceeding with update")
                    else:
                        messagebox.showerror("Update Error", 
                            "The update is not compatible with your system architecture.\n"
                            "Please download the correct version manually.")
                        return
            except Exception as e:
                print(f"Error verifying executable: {e}")
            
            # Create the batch file with admin elevation and retry logic
            batch_path = os.path.join(tempfile.gettempdir(), "SPHERE_updater.bat")
            
            # Get the parent directory of the executable to update the whole app directory
            app_dir = os.path.dirname(sys.executable)
            
            with open(batch_path, "w") as batch:
                batch.write(f'''@echo off
echo Updating SPHERE application...
echo Current app directory: {app_dir}
echo New executable: {tmp_exe}
echo Extract directory: {tmp_extract_dir}

:: Request admin privileges
>nul 2>&1 "%SYSTEMROOT%\\system32\\cacls.exe" "%SYSTEMROOT%\\system32\\config\\system"
if '%errorlevel%' NEQ '0' (
    echo Requesting administrative privileges...
    goto UACPrompt
) else (
    goto gotAdmin
)

:UACPrompt
echo Set UAC = CreateObject^("Shell.Application"^) > "%temp%\\getadmin.vbs"
echo UAC.ShellExecute "%~s0", "", "", "runas", 1 >> "%temp%\\getadmin.vbs"
"%temp%\\getadmin.vbs"
exit /B

:gotAdmin
if exist "%temp%\\getadmin.vbs" del "%temp%\\getadmin.vbs"

:: Kill all running SPHERE.exe processes
taskkill /f /im SPHERE.exe >nul 2>&1

:: Wait for the original process to close
echo Waiting for the application to close...
:WAITLOOP
tasklist | find /i "SPHERE.exe" >nul 2>&1
if not errorlevel 1 (
    echo SPHERE.exe is still running. Waiting...
    timeout /t 2 > nul
    goto WAITLOOP
)

:: Try to copy with retries
set /a attempts=0
set /a max_attempts=10

:RETRY
set /a attempts+=1
echo Attempt %attempts% of %max_attempts%...

:: Copy the entire directory contents
xcopy /E /I /Y "{tmp_extract_dir}\\*" "{app_dir}" > nul 2>&1
if %errorlevel% equ 0 (
    echo Update successful!
    start "" "{sys.executable}"
    goto CLEANUP
) else (
    echo Copy failed, retrying in 2 seconds...
    timeout /t 2 > nul
    if %attempts% lss %max_attempts% goto RETRY
    echo Update failed after %max_attempts% attempts.
    echo Please close all instances of SPHERE and try again.
    pause
)

:CLEANUP
del "{tmp_zip}"
rmdir /s /q "{tmp_extract_dir}"
del "%~f0"
''')
            
            # Inform user and start updater
            messagebox.showinfo("Update", "The application will close and update itself.\nPlease approve any administrator permission requests.")
            subprocess.Popen([batch_path], shell=True)
            root.destroy()
            sys.exit(0)
        else:
            # Improved error message with status code
            error_message = f"Failed to download the update (Status code: {response.status_code}).\n\n"
            
            # Add common error explanations
            if response.status_code == 404:
                error_message += f"The update file could not be found at:\n{UPDATE_EXE_URL}\n\nPlease check that the URL is correct and the file exists."
            elif response.status_code == 403:
                error_message += "Access to the update file was denied. Please check your network connection and permissions."
            elif response.status_code >= 500:
                error_message += "The server encountered an error. Please try again later."
            
            # Print the exact URL for debugging
            print(f"Update download failed. URL: {UPDATE_EXE_URL}, Status: {response.status_code}")
            
            messagebox.showerror("Update Error", error_message)
    except Exception as e:
        messagebox.showerror("Update Error", f"Update failed: {e}")

def extract_float(text):
    match = re.search(r"[-+]?\d*\.\d+|\d+", text)
    return float(match.group()) if match else 0.0

def extract_selection_list(cells):
        values = []
        for cell in cells:
            value = cell[0].value if cell[0].value is not None else ""
            value = value.strip()  
            if value == "":
                break
            values.append(value)
        return values

def extract_list(cells):
        values = []
        for cell in cells:
            value = cell[0].value if cell[0].value is not None else ""
            values.append(value)
        return values  # Do not stop; extract all values
    
def extract_transport_list(cells):
        values = []
        for cell in cells:
            value = cell[0].value if cell[0].value is not None else ""
            if value == "":  # Stop when a blank cell is encountered
                break
            values.append(value)
        return values

def extract_emission_list(cells):
        values = []
        for cell in cells:
            value = cell[0].value
            if value is None or str(value).strip() == "":
                break
            try:
                values.append(float(value))
            except Exception:
                values.append("N/A")
        return values

def store_all_widget_values(widgets):
    global previous_widget_values
    previous_widget_values = {}
    for key, widget in widgets.items():
        if hasattr(widget, 'get'):
            try:
                previous_widget_values[key] = widget.get()
            except Exception:
                previous_widget_values[key] = None

def restore_all_widget_values(widgets):
    global restoring_profile
    restoring_profile = True
    for key, value in previous_widget_values.items():
        widget = widgets.get(key)
        if widget:
            try:
                if hasattr(widget, 'set'):
                    widget.set(value)
                elif hasattr(widget, 'delete') and hasattr(widget, 'insert'):
                    widget.delete(0, tk.END)
                    widget.insert(0, value)
            except Exception:
                pass
    restoring_profile = False

def store_full_state(widgets):
    global previous_widget_values, previous_transport_rows, previous_machine_parts_data
    previous_widget_values = {}
    previous_transport_rows = []
    previous_machine_parts_data = []

    for key, widget in widgets.items():
        if hasattr(widget, 'get'):
            try:
                previous_widget_values[key] = widget.get()
            except Exception:
                previous_widget_values[key] = None

    # Store transport rows
    for row in transport_row_widgets:
        row_data = {
            'transport': row['transport_combo'].get(),
            'variation': row['variation_combo'].get(),
            'fuel': row['fuel_combo'].get(),
            'origin': row['origin_entry'].get(),
            'destination': row['destination_entry'].get(),
            'mode': row['transport_combo'].mode_var.get() if hasattr(row['transport_combo'], 'mode_var') else '',
            'value_label': row['value_label'].cget('text'),
        }
        previous_transport_rows.append(row_data)

    # Store machine process data (deep copy)
    import copy
    previous_machine_parts_data = copy.deepcopy(machine_parts_data)

def restore_full_state(widgets, generate_transport_rows, update_machine_process_page):
    global restoring_profile
    print("[DEBUG][restore_full_state] Restoring all widget values and dynamic rows to previous state.")
    restoring_profile = True

    tkroot = widgets['transport_country_combo'].winfo_toplevel()

    # Restore static widgets
    for key, value in previous_widget_values.items():
        widget = widgets.get(key)
        if widget:
            try:
                if hasattr(widget, 'set'):
                    widget.set(value)
                elif hasattr(widget, 'delete') and hasattr(widget, 'insert'):
                    widget.delete(0, tk.END)
                    widget.insert(0, value)
            except Exception:
                pass

    # Restore transport rows
    quantity = len(previous_transport_rows)
    print(f"[DEBUG][restore_full_state] Restoring {quantity} transport rows.")
    widgets['quantity_entry'].delete(0, tk.END)
    widgets['quantity_entry'].insert(0, str(quantity))

    try:
        generate_transport_rows(quantity)
    except tk.TclError:
        print("[DEBUG] Error generating transport rows - widget may not exist")
    
    # Check if actually have transport rows to restore
    if len(transport_row_widgets) > 0 and len(previous_transport_rows) > 0:
        for i, row_data in enumerate(previous_transport_rows):
            if i < len(transport_row_widgets):
                row = transport_row_widgets[i]
                try:
                    row['transport_combo'].set(row_data['transport'])
                    row['variation_combo'].set(row_data['variation'])
                    row['fuel_combo'].set(row_data['fuel'])
                    if 'origin_entry' in row and row['origin_entry']:
                        row['origin_entry'].delete(0, tk.END)
                        row['origin_entry'].insert(0, row_data['origin'])
                    if 'destination_entry' in row and row['destination_entry']:
                        row['destination_entry'].delete(0, tk.END)
                        row['destination_entry'].insert(0, row_data['destination'])
                    if hasattr(row['transport_combo'], 'mode_var'):
                        row['transport_combo'].mode_var.set(row_data['mode'])
                    if 'value_label' in row and row['value_label']:
                        row['value_label'].config(text=row_data['value_label'])
                except (tk.TclError, KeyError) as e:
                    print(f"[DEBUG] Error restoring transport row {i}: {e}")
    else:
        print(f"[DEBUG] Cannot restore transport rows: available={len(transport_row_widgets)}, previous={len(previous_transport_rows)}")

    # Restore machine process data
    import copy
    global machine_parts_data
    machine_parts_data = copy.deepcopy(previous_machine_parts_data)
    print(f"[DEBUG][restore_full_state] Restored machine_parts_data: {machine_parts_data}")
    update_machine_process_page()

    # Don't reset restoring_profile flag immediately
    def reset_restoring_flag():
        global restoring_profile
        restoring_profile = False
    
    tkroot.after(500, reset_restoring_flag)

# Data Extraction from Excel
book = openpyxl.load_workbook(resource_path('Emission Data.xlsx'))
sheet = book['Data Sheet']
sheet2 = book['Transport']
sheet3 = book['Transport WTT']
sheet4 = book['Machine Types']
sheet5 = book['Recycling']
sheet6 = book['Packaging']

country_cells = sheet['A2':'A999']
distance_cells = sheet['B2':'B999']
steel_cells = sheet['C2':'C999']
aluminium_cells = sheet['D2':'D999']
cement_cells = sheet['E2':'E999']
electricity_cells = sheet['F2':'F999']
plastic_cells = sheet['G2':'G999']
carbon_fiber_cells = sheet['H2':'H999']
Transport_cells = sheet2['A2':'A999']  # Different Transport Types
van_cells = sheet3['B2':'B5']  # Different Van Types
van_diesel_cells = sheet3['D2':'D5']
van_petrol_cells = sheet3['E2':'E5']
van_battery_cells = sheet3['F2':'F5']
HGV_diesel_cells = sheet3['B8':'B15']  # Different HGV Types
HGV_0_diesel_cells = sheet3['D8':'D15']
HGV_50_diesel_cells = sheet3['E8':'E15']
HGV_100_diesel_cells = sheet3['F8':'F15']
HGV_R_diesel_cells = sheet3['B17':'B24']  # Different HGV Refrigerated Types
HGV_R_0_diesel_cells = sheet3['D17':'D24']
HGV_R_50_diesel_cells = sheet3['E17':'E24']
HGV_R_100_diesel_cells = sheet3['F17':'F24']
flight_cells = sheet3['B27':'B29']  # Different Flight Types
Int_flight_cells = sheet3['D27':'D27']
L_1500_flight_cells = sheet3['E27':'E29']
M_1500_flight_cells = sheet3['F27':'F29']
rail_cells = sheet3['B31':'B31']  # Different Rail Types
freight_train_rail_cells = sheet3['D31':'D31']
sea_tanker_cells = sheet3['G34':'G38']  # Different Sea Tanker Types
crude_tanker_cells = sheet3['C34':'C40']  # Different Sea Tanker Types Sizes
products_tanker_cells = sheet3['C41':'C46']
chemical_tanker_cells = sheet3['C47':'C51']
LNG_tanker_cells = sheet3['C52':'C54']
LPG_tanker_cells = sheet3['C55':'C57']
crude_value_cells = sheet3['E34':'E40']  # Different Sea Tanker Type Sizes Values
products_value_cells = sheet3['E41':'E46']
chemical_value_cells = sheet3['E47':'E51']
LNG_value_cells = sheet3['E52':'E54']
LPG_value_cells = sheet3['E55':'E57']
cargo_cells = sheet3['G60':'G64']  # Different Cargo Ship Types
Bulk_carrier_cells = sheet3['C59':'C65']  # Different Cargo Ship Types Sizes
General_cargo_cells = sheet3['C66':'C72']
Container_ship_cells = sheet3['C73':'C79']
Vehicle_transport_cells = sheet3['C80':'C82']
Roro_ferry_cells = sheet3['C83':'C85']
bulk_value_cells = sheet3['E59':'E65']  # Different Cargo Ship Type Sizes Values
general_value_cells = sheet3['E66':'E72']
container_value_cells = sheet3['E73':'E79']
vehicle_value_cells = sheet3['E80':'E82']
roro_value_cells = sheet3['E83':'E85']
machine_value_cells = sheet4['A2':'A999']  # Different Machine Types
specific_machine_energy_use_cells = sheet4['B2':'B999']  # Different Machine Types Energy Use
metal_recyling_types = sheet5['A2':'A999']  # Different Metal Recycling Types
metal_recyling_emission = sheet5['B2':'B999']  # Different Metal Recycling Types Emission
packaging_types = sheet6['A2':'A999']  # Different Packaging Types
packaging_box_frame = sheet6['B2':'B999']  # Different Packaging Types Box Frame

country_list = extract_selection_list(country_cells) 
steel_list = extract_list(steel_cells)
distance_list = extract_list(distance_cells)
aluminium_list = extract_list(aluminium_cells)
cement_list = extract_list(cement_cells)
electricity_list = extract_list(electricity_cells)
plastic_list = extract_list(plastic_cells)
carbon_fiber_list = extract_list(carbon_fiber_cells)
transport_list = extract_transport_list(Transport_cells)
van_list = extract_list(van_cells)
van_diesel_list = extract_list(van_diesel_cells)
van_petrol_list = extract_list(van_petrol_cells)
van_battery_list = extract_list(van_battery_cells)
HGV_diesel_list = extract_list(HGV_diesel_cells)
HGV_0_diesel_list = extract_list(HGV_0_diesel_cells)
HGV_50_diesel_list = extract_list(HGV_50_diesel_cells)
HGV_100_diesel_list = extract_list(HGV_100_diesel_cells)
HGV_R_diesel_list = extract_list(HGV_R_diesel_cells)
HGV_R_0_diesel_list = extract_list(HGV_R_0_diesel_cells)
HGV_R_50_diesel_list = extract_list(HGV_R_50_diesel_cells)
HGV_R_100_diesel_list = extract_list(HGV_R_100_diesel_cells)
flight_list = extract_list(flight_cells)
Int_flight_list = extract_list(Int_flight_cells)
L_1500_flight_list = extract_list(L_1500_flight_cells)
M_1500_flight_list = extract_list(M_1500_flight_cells)
rail_list = extract_list(rail_cells)
freight_train_rail_list = extract_list(freight_train_rail_cells)
sea_tanker_list = extract_list(sea_tanker_cells)
crude_tanker_list =extract_list(crude_tanker_cells)
products_tanker_list = extract_list(products_tanker_cells)
chemical_tanker_list = extract_list(chemical_tanker_cells)
LNG_tanker_list = extract_list(LNG_tanker_cells)
LPG_tanker_list = extract_list(LPG_tanker_cells)
crude_value_list = extract_list(crude_value_cells)
products_value_list = extract_list(products_value_cells)
chemical_value_list = extract_list(chemical_value_cells)
LNG_value_list = extract_list(LNG_value_cells)
LPG_value_list = extract_list(LPG_value_cells)
cargo_list = extract_list(cargo_cells)
bulk_value_list = extract_list(bulk_value_cells)
general_value_list = extract_list(general_value_cells)
container_value_list = extract_list(container_value_cells)
vehicle_value_list = extract_list(vehicle_value_cells)
roro_value_list = extract_list(roro_value_cells)
machine_value_list = extract_selection_list(machine_value_cells)
specific_machine_energy_use_list = extract_list(specific_machine_energy_use_cells)
metal_recyling_types_list = extract_selection_list(metal_recyling_types)
metal_recyling_emission_list = extract_emission_list(metal_recyling_emission)
packaging_types_list = extract_selection_list(packaging_types)
packaging_box_frame_list = extract_emission_list(packaging_box_frame)

from geopy.geocoders import Nominatim

def build_country_coords_cache(country_list):
    geolocator = Nominatim(user_agent="emission_app")
    cache = {}
    for country in country_list:
        try:
            loc = geolocator.geocode(country, timeout=5)
            if loc:
                cache[country] = (loc.latitude, loc.longitude)
            else:
                cache[country] = None
        except Exception:
            cache[country] = None
    return cache

allowed_machine_substitutions = {}

non_substitutable = [
    "Surface Grinder (S.G)",
    "Wire Cut (W/C)",
    "Rokou Roku Micro Fine Precision Machine Center"
] # These machines cannot be substituted with others
cnc_names = [
    "CNC (RB8)",
    "CNC (YCM4)",
    "CNC (YCM1)"
] # These machines are CNC machines and can be substituted with each other
cnc_machines = [m for m in machine_value_list if m in cnc_names]

for m in machine_value_list:
    if m in non_substitutable:
        allowed_machine_substitutions[m] = []
    elif m in cnc_machines:
        allowed_machine_substitutions[m] = [x for x in cnc_machines if x != m]
    else:
        allowed_machine_substitutions[m] = []

def suggest_optimal_machine(current_machine, weight_after, weight_after_quantity, machining_time, country_index):
    idx = machine_value_list.index(current_machine)
    current_energy = float(specific_machine_energy_use_list[idx])
    electricity_emission = float(electricity_list[country_index])
    current_emission = current_energy * weight_after * weight_after_quantity * (machining_time / 60) * electricity_emission

    best_machine = current_machine
    best_emission = current_emission

    for alt_machine in allowed_machine_substitutions[current_machine]:
        alt_idx = machine_value_list.index(alt_machine)
        alt_energy = float(specific_machine_energy_use_list[alt_idx])
        alt_emission = alt_energy * weight_after * weight_after_quantity * (machining_time / 60) * electricity_emission
        if alt_emission < best_emission:
            best_machine = alt_machine
            best_emission = alt_emission

    return best_machine, best_emission #default to current machine if no better alternative is found
country_coords_cache = build_country_coords_cache(country_list)

def get_best_material_and_country(
    total_weight, country_coords_cache, country_list, materials, material_emission_lists, custom_factors, user_country, current_material=None
):
    best = None
    min_total_emission = float('ivxcnf')
    
    # Filter materials based on current selection if provided
    target_materials = [current_material] if current_material in materials else materials
    
    for mat_idx, mat in enumerate(materials):
        # Skip materials not in our target list
        if mat not in target_materials:
            continue
            
        mat_idx = materials.index(mat) 
        for country_idx, country in enumerate(country_list):
            # Material emission
            emission = custom_factors.get("materials", {}).get(mat)
            if emission is None:
                value = material_emission_lists[mat_idx][country_idx]
                try:
                    emission = float(value)
                except (ValueError, TypeError):
                    continue 
            if not user_country or user_country not in country_coords_cache or country not in country_coords_cache:
                continue
            distance_km = geodesic(country_coords_cache[country], country_coords_cache[user_country]).km
            transport_emission_factor = 0.01  # CO2e per ton-km
            transport_emission = total_weight * distance_km * transport_emission_factor / 1000
            total_emission = emission * total_weight + transport_emission
            if total_emission < min_total_emission:
                min_total_emission = total_emission
                best = {
                    "material": mat,
                    "source_country": country,
                    "material_emission": emission,
                    "transport_emission": transport_emission,
                    "total_emission": total_emission,
                    "distance_km": distance_km
                }
    return best # If no material is found, return None

def load_custom_emission_factors():
    if os.path.exists(EMISSION_FACTORS_FILE):
        with open(EMISSION_FACTORS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}
custom_emission_factors = load_custom_emission_factors()

def save_custom_emission_factors(data):
    with open(EMISSION_FACTORS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

def show_custom_emission_factors_dialog(root, material_list, on_update=None):
    factors = load_custom_emission_factors()
    dialog = tk.Toplevel(root)
    dialog.title("Custom Emission Factors")
    dialog.geometry("400x500")
    dialog.grab_set()
    tk.Label(dialog, text="Edit Emission Factors (kg CO2e/unit)").pack(pady=5)

    entries = {}

    tk.Label(dialog, text="Materials:").pack(anchor="w")
    for mat in material_list:
        frame = tk.Frame(dialog)
        frame.pack(fill="x", padx=10)
        tk.Label(frame, text=mat, width=22, anchor="w").pack(side="left")
        val = tk.StringVar(value=str(factors.get("materials", {}).get(mat, "")))
        entries[f"mat_{mat}"] = val
        entry = tk.Entry(frame, textvariable=val, width=10)
        entry.pack(side="left")

    def save_and_close():
        new_factors = {"materials": {}}
        for k, v in entries.items():
            if v.get().strip():
                try:
                    val = float(v.get())
                except Exception:
                    continue
                if k.startswith("mat_"):
                    new_factors["materials"][k[4:]] = val
                elif k == "electricity":
                    new_factors["electricity"] = val
        save_custom_emission_factors(new_factors)
        global custom_emission_factors
        custom_emission_factors = load_custom_emission_factors()
        if on_update:
            on_update()
        dialog.destroy()
        messagebox.showinfo("Saved", "Custom emission factors saved and applied.")

    def reset_to_default():
        # Clear all entry fields
        for k, v in entries.items():
            v.set("")
        # Remove custom emission factors from file
        save_custom_emission_factors({})
        global custom_emission_factors
        custom_emission_factors = load_custom_emission_factors()
        if on_update:
            on_update()
        messagebox.showinfo("Reset", "All emission factors reset to default.\nOverview updated.")

    tk.Button(dialog, text="Reset to Default", command=reset_to_default, bg="#FF6666", fg="white").pack(pady=(10, 0))
    tk.Button(dialog, text="Save", command=save_and_close).pack(pady=10)
    tk.Button(dialog, text="Cancel", command=dialog.destroy).pack()

def sync_current_part_ui_to_data():
    if not machine_parts_data or not machine_process_row_widgets:
        return
    part_data = machine_parts_data[current_part_index]
    for row_idx, row_widgets in enumerate(machine_process_row_widgets):
        if row_idx < len(part_data["rows"]):
            row = part_data["rows"][row_idx]
            row["machine_type"] = row_widgets["machine_type_combo"].get()
            row["energy_use"] = row_widgets["energy_use_value_label"].cget("text")
            row["machining_time"] = row_widgets["machining_time_entry"].get()
            row["carbon_emission"] = row_widgets["carbon_emission_value_label"].cget("text")

def save_profile_inputs(profile_data, widgets):
    sync_current_part_ui_to_data()
    # Save static widgets
    for key in [
        'transport_country_combo', 'material_combo', 'weight_entry', 'quantity_entry_weight',
        'weight_after_entry', 'weight_after_quantity_entry', 'packaging_dropdown',
        'metal_recycling_dropdown', 'machine_process_quantity_entry', 'manufacturing_country_combo', 'quantity_entry'
    ]:
        widget = widgets.get(key)
        if widget and widget.winfo_exists() and hasattr(widget, 'get'):
            profile_data[key] = widget.get()
    print("Saved static widget values:", {k: profile_data[k] for k in profile_data if k not in ['transport_rows', 'machine_process_rows']})

    # Save dynamic transport rows
    profile_data['transport_rows'] = []
    for idx, row in enumerate(transport_row_widgets):
        transport_combo = row.get('transport_combo')
        if transport_combo and transport_combo.winfo_exists():
            row_data = {
                'transport': transport_combo.get(),
                'variation': row['variation_combo'].get() if row['variation_combo'].winfo_exists() else '',
                'fuel': row['fuel_combo'].get() if row['fuel_combo'].winfo_exists() else '',
                'origin': row['origin_entry'].get() if row.get('origin_entry') and row['origin_entry'].winfo_exists() else '',
                'destination': row['destination_entry'].get() if row.get('destination_entry') and row['destination_entry'].winfo_exists() else '',
                'mode': transport_combo.mode_var.get() if hasattr(transport_combo, 'mode_var') else '',
                'value_label': row['value_label'].cget('text') if row.get('value_label') and row['value_label'].winfo_exists() else '',
            }
            print(f"[DEBUG][save_profile_inputs] Row {idx}: transport={row_data['transport']}, variation={row_data['variation']}, fuel={row_data['fuel']}, origin={row_data['origin']}, destination={row_data['destination']}, mode={row_data['mode']}, value_label={row_data['value_label']}")
            profile_data['transport_rows'].append(row_data)
    # Save dynamic business rows (Business Travel)
    profile_data['business_travel_rows'] = []
    for idx, row in enumerate(business_travel_row_widgets):
        # matches the keys created in the Business Travel UI block
        type_combo = row.get('type_combo')
        variation_combo = row.get('variation_combo')
        fuel_combo = row.get('fuel_combo')
        origin_entry = row.get('origin_entry')
        destination_entry = row.get('destination_entry')
        mode_combo = row.get('mode_combo')

        if type_combo and type_combo.winfo_exists():
            row_data = {
                'transport': type_combo.get(),
                'variation': variation_combo.get() if variation_combo and variation_combo.winfo_exists() else '',
                'fuel':      fuel_combo.get() if fuel_combo and fuel_combo.winfo_exists() else '',
                'origin':    origin_entry.get() if origin_entry and origin_entry.winfo_exists() else '',
                'destination': destination_entry.get() if destination_entry and destination_entry.winfo_exists() else '',
                'mode':      mode_combo.get() if mode_combo and mode_combo.winfo_exists() else '',
                'value_label': '',  # (no value label in Business Travel rows)
            }
            print(f"[DEBUG][save_profile_inputs][BT] Row {idx}: "
                f"transport={row_data['transport']}, variation={row_data['variation']}, "
                f"fuel={row_data['fuel']}, origin={row_data['origin']}, "
                f"destination={row_data['destination']}, mode={row_data['mode']}")
            profile_data['business_travel_rows'].append(row_data)


    # Only update the current part (book-style navigation)
    if machine_parts_data and machine_process_row_widgets:
        part_data = machine_parts_data[current_part_index]
        for row_idx, row_widgets in enumerate(machine_process_row_widgets):
            if row_idx < len(part_data["rows"]):
                row = part_data["rows"][row_idx]
                row["machine_type"] = row_widgets["machine_type_combo"].get()
                row["energy_use"] = row_widgets["energy_use_value_label"].cget("text")
                row["machining_time"] = row_widgets["machining_time_entry"].get()
                row["carbon_emission"] = row_widgets["carbon_emission_value_label"].cget("text")

    # Save dynamic machine process rows (book-style)
    profile_data['machine_process_rows'] = []
    for part_idx, part_data in enumerate(machine_parts_data, 1):
        if "rows" in part_data:
            for row_idx, row in enumerate(part_data["rows"], 1):
                profile_data['machine_process_rows'].append({
                    "part": f"Part {part_idx}",
                    "row": row_idx,
                    "machine_type": row.get("machine_type", ""),
                    "energy_use": row.get("energy_use", "0.0"),
                    "machining_time": row.get("machining_time", ""),
                    "carbon_emission": row.get("carbon_emission", "0.0")
                })
    print(f"[DEBUG][save_profile_inputs] Saved {len(profile_data['machine_process_rows'])} machine process rows (book-style).")

def load_profiles():
    if os.path.exists(PROFILE_FILE):
        try:
            with open(PROFILE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            messagebox.showerror("Error", "Profile file is corrupted. Starting with an empty profile list.")
            return {}
    return {}

def save_profiles(profiles):
    with open(PROFILE_FILE, "w", encoding="utf-8") as f:
        json.dump(profiles, f, indent=2)

def compare_profiles(profiles, selected_names):
    import io
    from PIL import Image, ImageTk

    compare_win = tk.Toplevel()
    compare_win.title("Profile Comparison")

    # Define fields to compare
    fields = [
        ("Material", "material_combo"),
        ("Raw Weight (kg)", "weight_entry"),
        ("Packaging Type", "packaging_dropdown"),
        ("Total Emission (CO2eKg)", "total_emission"),
        # Add more fields as needed
    ]

    tree = ttk.Treeview(compare_win, columns=["Field"] + selected_names, show="headings")
    tree.heading("Field", text="Field")
    for name in selected_names:
        tree.heading(name, text=name)
    tree.pack(fill="x", expand=False)

    # Insert rows for each field, formatting numbers to 2dp
    for label, key in fields:
        row = [label]
        for name in selected_names:
            val = profiles[name].get(key, "")
            # Format floats to 2dp, else show as string
            try:
                fval = float(val)
                val = f"{fval:.2f}"
            except (ValueError, TypeError):
                pass
            row.append(val)
        tree.insert("", "end", values=row)

    # Show chart images for each profile
    chart_frame = tk.Frame(compare_win)
    chart_frame.pack(fill="both", expand=True)
    for name in selected_names:
        # Generate chart for this profile (reuse your chart code)
        try:
            labels = [
                'Raw Material Emission',
                'Transport Emission',
                'Machine Emission',
                'Packaging Emission',
                'Recycled Emission'
            ]
            values = [
                float(profiles[name].get("material_emission", 0)),
                float(profiles[name].get("transport_emission", 0)),
                float(profiles[name].get("machine_emission", 0)),
                float(profiles[name].get("packaging_emission", 0)),
                float(profiles[name].get("recycled_emission", 0)),
            ]
            # Format values to 2dp for chart labels
            values = [round(v, 2) for v in values]
            colors = ['#FF9999', '#66B3FF', '#99FF99', '#FFB347', '#FFD700']
            filtered = [(label, value, color) for label, value, color in zip(labels, values, colors) if value > 0]

            if filtered:
                import matplotlib.pyplot as plt
                fig, ax = plt.subplots(figsize=(3, 2))
                filtered_labels, filtered_values, filtered_colors = zip(*filtered)
                # Format autopct to 2dp
                ax.pie(filtered_values, labels=[f"{l} ({v:.2f})" for l, v in zip(filtered_labels, filtered_values)],
                       colors=filtered_colors, autopct=lambda pct: f"{pct:.2f}%", startangle=90)
                ax.axis('equal')
                buf = io.BytesIO()
                fig.savefig(buf, format='png')
                buf.seek(0)
                img = Image.open(buf)
                photo = ImageTk.PhotoImage(img)
                label = tk.Label(chart_frame, text=name, image=photo, compound="top")
                label.image = photo  
                label.pack(side="left", padx=10)
                plt.close(fig)
        except Exception as e:
            label = tk.Label(chart_frame, text=f"{name}\n(No chart data)", compound="top")
            label.pack(side="left", padx=10)

class CoverPage(tk.Frame):
    def __init__(self, parent, show_main_page_callback):
        super().__init__(parent)
        self.configure(bg="#A8D5BA")
        self.show_main_page_callback = show_main_page_callback

        main_frame = tk.Frame(self, bg="#A8D5BA")
        main_frame.pack(fill="both", expand=True, padx=30, pady=30)

        # Left: Instructions
        left_frame = tk.Frame(main_frame, bg="#E0F2E9", width=350)
        left_frame.pack(side="left", fill="y", padx=(0, 30), pady=0)
        left_frame.pack_propagate(False)

        # Add SPHERE logo at the top of the left frame
        try:
            logo_path = resource_path("SPHERE_logo.png")
            logo_image = tk.PhotoImage(file=logo_path)
            logo_image = logo_image.subsample(4, 4)  # Adjust the subsample values to resize
            logo_label = tk.Label(left_frame, image=logo_image, bg="#E0F2E9")
            logo_label.image = logo_image 
            logo_label.pack(pady=(10, 5), anchor="center")
        except Exception as e:
            print(f"Error loading SPHERE logo: {e}")

        tk.Label(left_frame, text="How to Use", font=("Arial", 16, "bold"), bg="#E0F2E9", fg="#CC0000").pack(pady=(10, 5))
        instructions = (
            "1. To create a new product profile, enter a name and click 'Create Profile'.\n\n"
            "2. To load, delete, rename, or duplicate an existing profile, select it from the dropdown and use the buttons.\n\n"
            "3. Use 'Compare Profiles' to select and compare multiple profiles.\n\n"
            "4. All data is saved automatically. You can export or import profiles from the main page.\n\n"
            "5. For more help, Hover over info icons in the app for field-specific guidance."
        )
        tk.Label(left_frame, text=instructions, font=("Arial", 11), bg="#E0F2E9", justify="left", anchor="nw", wraplength=320).pack(fill="both", expand=True, padx=10, pady=10)

        # Right: Profile management 
        right_frame = tk.Frame(main_frame, bg="#A8D5BA")
        right_frame.pack(side="left", fill="both", expand=True)

        tk.Label(right_frame, text="Welcome to Carbon Emission Calculator", font=("Arial", 24, "bold"), bg="#A8D5BA", fg="#CC0000").pack(pady=20)
        tk.Label(right_frame, text="Singapore Polytechnic", font=("Arial", 18), bg="#A8D5BA", fg="black").pack(pady=5)

        self.profiles = load_profiles()
        self.selected_profile = tk.StringVar()
        self.profile_data = {}

        # Profile management section
        frame = tk.Frame(right_frame, bg="#A8D5BA")
        frame.pack(pady=10)
        tk.Label(frame, text="Select Product Profile:", bg="#A8D5BA").grid(row=0, column=0, sticky="e")
        self.profile_combo = ttk.Combobox(frame, textvariable=self.selected_profile, values=sorted(self.profiles.keys()), state="readonly")
        self.profile_combo.grid(row=0, column=1, padx=5)
        ttk.Button(frame, text="Load", command=self.load_profile).grid(row=0, column=2, padx=5)
        ttk.Button(frame, text="Delete", command=self.delete_profile).grid(row=0, column=3, padx=5)
        ttk.Button(frame, text="Rename", command=self.rename_profile).grid(row=0, column=4, padx=5)
        ttk.Button(frame, text="Duplicate", command=self.duplicate_profile).grid(row=0, column=5, padx=5)
        compare_button = ttk.Button(frame, text="Compare Profiles", command=self.open_comparison)
        compare_button.grid(row=2, column=0, columnspan=6, pady=5)

        # Or create new profile
        tk.Label(right_frame, text="Or Create New Product Profile", bg="#A8D5BA", font=("Arial", 12, "bold")).pack(pady=10)
        form = tk.Frame(right_frame, bg="#A8D5BA")
        form.pack(pady=5)
        tk.Label(form, text="Product Name:", bg="#A8D5BA").grid(row=0, column=0, sticky="e")
        self.product_entry = ttk.Entry(form)
        self.product_entry.grid(row=0, column=1, padx=5)

        # Info icon with tooltip
        info_icon_product = tk.Label(form, text="ℹ️", bg="#A8D5BA", font=("Arial", 12))
        info_icon_product.grid(row=0, column=2, sticky="w")
        ToolTip(info_icon_product, "Enter the product name for your profile. Example: Aluminum Gear Housing")

        def on_product_name_change(event):
            if self.product_entry.get().strip():
                self.create_profile_button.config(state="normal")
            else:
                self.create_profile_button.config(state="disabled")

        self.create_profile_button = ttk.Button(form, text="Create Profile", command=self.create_profile, state="disabled")
        self.create_profile_button.grid(row=1, column=0, columnspan=2, pady=5)
        self.product_entry.bind("<KeyRelease>", on_product_name_change)
        self.product_entry.focus_set()
       
        version_label = tk.Label(
            right_frame,
            text=f"Version: {APP_VERSION}",
            font=("Arial", 10, "italic"),
            bg="#A8D5BA",
            fg="#555"
        )
        version_label.pack(side="bottom", anchor="e", pady=(10, 5), padx=10)

    def create_profile(self):
        # Only save widget values if widgets exist (i.e., on the main page)
        if hasattr(self.master, 'widgets') and self.master.widgets:
            save_profile_inputs(self.profile_data, self.master.widgets)
            product_name = self.selected_profile.get() or self.product_entry.get().strip()
            if product_name:
                self.profiles[product_name] = dict(self.profile_data)
            save_profiles(self.profiles)
        # Always create a new profile with just the product name on the cover page
        product_name = self.product_entry.get().strip()
        if not product_name:
            messagebox.showerror("Error", "Product Name is required.")
            return
        if product_name in self.profiles:
            messagebox.showerror("Error", "Profile already exists.")
            return
        self.profiles[product_name] = {"product_name": product_name, "description": ""}
        save_profiles(self.profiles)
        self.profile_combo["values"] = sorted(self.profiles.keys())
        self.selected_profile.set(product_name)
        self.profile_data = self.profiles[product_name]
        self.show_main_page_callback(self.profile_data)

    def load_profile(self):
        product_name = self.selected_profile.get()
        if not product_name or product_name not in self.profiles:
            messagebox.showerror("Error", "Select a valid profile.")
            return
        self.profile_data = self.profiles[product_name]
        self.profile_data["product_name"] = product_name
        self.show_main_page_callback(self.profile_data)

    def delete_profile(self):
        product_name = self.selected_profile.get()
        if not product_name or product_name not in self.profiles:
            messagebox.showerror("Error", "Select a valid profile to delete.")
            return
        if messagebox.askyesno("Confirm Delete", f"Delete profile '{product_name}'?"):
            del self.profiles[product_name]
            save_profiles(self.profiles)
            self.profile_combo["values"] = sorted(self.profiles.keys())
            self.selected_profile.set("")
            self.profile_data = {}
            messagebox.showinfo("Deleted", f"Profile '{product_name}' deleted.")

    def rename_profile(self):
        old_name = self.selected_profile.get()
        if not old_name or old_name not in self.profiles:
            messagebox.showerror("Error", "Select a valid profile to rename.")
            return
        new_name = simpledialog.askstring("Rename Profile", "Enter new profile name:", initialvalue=old_name)
        if not new_name:
            return
        if new_name in self.profiles:
            messagebox.showerror("Error", "A profile with this name already exists.")
            return
        # Rename in the profiles dict
        self.profiles[new_name] = self.profiles.pop(old_name)
        self.profiles[new_name]["product_name"] = new_name
        save_profiles(self.profiles)
        self.profile_combo["values"] = sorted(self.profiles.keys())
        self.selected_profile.set(new_name)
        self.profile_data = self.profiles[new_name]
        messagebox.showinfo("Renamed", f"Profile renamed to '{new_name}'.")

    def duplicate_profile(self):
        old_name = self.selected_profile.get()
        if not old_name or old_name not in self.profiles:
            messagebox.showerror("Error", "Select a valid profile to duplicate.")
            return
        new_name = simpledialog.askstring("Duplicate Profile", "Enter new profile name:", initialvalue=f"{old_name}_copy")
        if not new_name:
            return
        if new_name in self.profiles:
            messagebox.showerror("Error", "A profile with this name already exists.")
            return
        # Deep copy the profile data
        import copy
        new_profile = copy.deepcopy(self.profiles[old_name])
        new_profile["product_name"] = new_name
        if "description" not in new_profile:
            new_profile["description"] = ""
        self.profiles[new_name] = new_profile
        save_profiles(self.profiles)
        self.profile_combo["values"] = sorted(self.profiles.keys())
        self.selected_profile.set(new_name)
        self.profile_data = self.profiles[new_name]
        messagebox.showinfo("Duplicated", f"Profile duplicated as '{new_name}'.")

    def open_comparison(self):
        if not self.profiles or len(self.profiles) < 2:
            messagebox.showinfo("Info", "At least two profiles are required to compare.")
            return

        dialog = tk.Toplevel(self)
        dialog.title("Select Profiles to Compare")
        dialog.geometry("350x350")
        dialog.grab_set()

        tk.Label(dialog, text="Select profiles to compare:", font=("Arial", 11, "bold")).pack(pady=10)

        listbox = tk.Listbox(dialog, selectmode=tk.MULTIPLE, width=40, height=12)
        for name in sorted(self.profiles.keys()):
            listbox.insert(tk.END, name)
        listbox.pack(padx=10, pady=10, fill="both", expand=True)

        def do_compare():
            selected_indices = listbox.curselection()
            selected = [listbox.get(i) for i in selected_indices]
            if len(selected) < 2:
                messagebox.showinfo("Info", "Select at least two profiles to compare.")
                return
            dialog.destroy()
            compare_profiles(self.profiles, selected)

        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="Compare", command=do_compare).pack(side="left", padx=5)
        tk.Button(btn_frame, text="Cancel", command=dialog.destroy).pack(side="left", padx=5)

def show_main_page(root, profile_data, back_to_cover_callback, cover_page, active_profile_name, settings_menu):
    global inner_box_frame, right_scrollable_frame, pie_chart_canvas
    parts_debounce_id = None
    pie_chart_canvas = None
    auto_save_after_id = None
    debounce_id = None 
    for widget in root.winfo_children():
        if widget is not cover_page:
            widget.destroy()
        else:
            widget.pack_forget()

    # Header
    header_text = f"Product Profile: {profile_data.get('product_name', '')}"
    header_label = tk.Label(root, text=header_text, font=("Arial", 14, "bold"), bg="#A8D5BA", fg="#333")
    header_label.pack(side="top", fill="x")

    desc_frame = tk.Frame(root, bg="#A8D5BA")
    desc_frame.pack(fill="x", padx=10, pady=5)

    tk.Label(desc_frame, text="Description:", bg="#A8D5BA").pack(side="left", anchor="n")

    desc_var = tk.StringVar(value=profile_data.get("description", ""))
    desc_entry = ttk.Entry(desc_frame, textvariable=desc_var, width=60)
    desc_entry.pack(side="left", padx=5)

    info_icon_desc = tk.Label(desc_frame, text="ℹ️", bg="#A8D5BA", font=("Arial", 12))
    info_icon_desc.pack(side="left")
    ToolTip(info_icon_desc, "Provide a brief description or notes about the product.")

    def save_desc_entry(event=None):
        profile_data["description"] = desc_var.get()
        if active_profile_name and hasattr(cover_page, "profiles"):
            cover_page.profiles[active_profile_name] = dict(profile_data)
            save_profiles(cover_page.profiles)

    desc_entry.bind("<FocusOut>", save_desc_entry)
    desc_entry.bind("<KeyRelease>", lambda e: save_desc_entry())

    # Header frame
    header_frame = tk.Frame(root, bg="#A8D5BA", height=100)
    header_frame.pack(side="top", fill="x")

    import copy
    last_saved_state = {}

    def save_profile_inputs_with_undo(profile_data, widgets):
        nonlocal last_saved_state
        save_profile_inputs(profile_data, widgets)
        last_saved_state = copy.deepcopy(profile_data)

    def undo_to_last_saved():
        if last_saved_state:
            restore_profile_inputs(last_saved_state, widgets, generate_transport_rows, recalculate_all)
            messagebox.showinfo("Undo", "Reverted to last saved state.")

    undo_button = tk.Button(header_frame, text="Undo", command=undo_to_last_saved, bg="#FFC107", fg="black")
    undo_button.pack(side="right", padx=5, pady=10)

    class LoadingIndicator:
        def __init__(self, parent, message="Loading..."):
            self.top = tk.Toplevel(parent)
            self.top.title("")
            self.top.geometry("250x80")
            self.top.transient(parent)
            self.top.grab_set()
            self.top.resizable(False, False)
            self.top.attributes("-topmost", True)
            self.label = tk.Label(self.top, text=message)
            self.label.pack(pady=10)
            self.progress = ttk.Progressbar(self.top, mode="indeterminate")
            self.progress.pack(fill="x", padx=20, pady=5)
            self.progress.start(10)
            self.top.protocol("WM_DELETE_WINDOW", lambda: None)  # Disable close

        def close(self):
            self.progress.stop()
            self.top.grab_release()
            self.top.destroy()

    import tempfile
    from openpyxl.drawing.image import Image as XLImage

    def save_current_chart_as_image(filtered, chart_type, filename):
        import matplotlib.pyplot as plt
        if not filtered:
            # Create a blank figure or show "No data"
            fig, ax = plt.subplots(figsize=(6, 3))
            ax.text(0.5, 0.5, "No emission data", ha='center', va='center', fontsize=14)
            ax.axis('off')
            fig.savefig(filename, bbox_inches='tight', dpi=150)
            plt.close(fig)
            return

        if chart_type == "Pie Chart":
            fig, ax = plt.subplots(figsize=(6, 3))
            labels, values, colors = zip(*filtered)
            total = sum(values)
            explode = [0.12 if v / total < 0.08 else 0.0 for v in values]
            ax.set_position([0.32, 0.08, 0.6, 0.84])
            wedges, _ = ax.pie(
                values,
                startangle=90,
                colors=colors,
                labels=None,
                explode=explode,
                autopct=None,
                pctdistance=0.8,
                textprops={'fontsize': 8, 'color': 'black', 'weight': 'bold'}
            )
            ax.axis('equal')
            ax.legend(
                wedges,
                labels,
                title="Emission Types",
                loc="lower left",
                bbox_to_anchor=(-0.35, 0),
                fontsize=7,
                title_fontsize=8,
                frameon=True
            )
        else:  # Pareto Chart
            fig, ax1 = plt.subplots(figsize=(6, 3))
            labels, values, colors = zip(*filtered)
            sorted_data = sorted(zip(values, labels, colors), reverse=True)
            sorted_values, sorted_labels, sorted_colors = zip(*sorted_data)
            cum_values = [sum(sorted_values[:i+1]) for i in range(len(sorted_values))]
            total = sum(sorted_values)
            cum_perc = [v / total * 100 for v in cum_values]
            x = range(len(sorted_labels))
            ax1.bar(x, sorted_values, color=sorted_colors)
            ax1.set_ylabel('Emission (CO2eKg)', fontsize=9)
            ax1.set_xticks(x)
            ax1.set_xticklabels(sorted_labels, rotation=30, ha='right', fontsize=7)
            fig.subplots_adjust(bottom=0.28)
            ax2 = ax1.twinx()
            ax2.plot(x, cum_perc, color='red', marker='o', linestyle='-', linewidth=2)
            ax2.set_ylabel('Cumulative %', fontsize=9)
            ax2.set_ylim(0, 110)
            ax2.grid(False)
            for i, v in enumerate(sorted_values):
                ax1.text(i, v, f"{v:.1f}", ha='center', va='bottom', fontsize=8)
        fig.savefig(filename, bbox_inches='tight', dpi=150)
        plt.close(fig)

    def get_optimal_material_emission(total_weight, country_index, custom_factors, material_lists, materials):
        min_emission = float('inf')
        best_material = None
        for i, mat in enumerate(materials):
            emission = custom_factors.get("materials", {}).get(mat)
            if emission is None:
                emission = float(material_lists[i][country_index])
            if emission < min_emission:
                min_emission = emission
                best_material = mat
        return best_material, min_emission * total_weight

    def get_optimal_transport_emission(total_weight, distance, transport_options):
        min_emission = float('inf')
        best_option = None
        for option in transport_options:
            emission = total_weight * (distance / 1000) * option[0]
            if emission < min_emission:
                min_emission = emission
                best_option = option[1]
        return best_option, min_emission

    def get_optimal_machine_emission(weight_after, weight_after_quantity, machining_time, country_index, machine_energy_list, electricity_list, machine_names):
        min_emission = float('inf')
        best_machine = None
        for i, machine in enumerate(machine_names):
            try:
                energy_use = float(machine_energy_list[i])
                electricity_emission = float(electricity_list[country_index])
                emission = energy_use * weight_after * weight_after_quantity * (machining_time / 60) * electricity_emission
                if emission < min_emission:
                    min_emission = emission
                    best_machine = machine
            except Exception:
                continue
        return best_machine, min_emission

    def get_emission_reduction_suggestions(labels, values, total):
        suggestions = []
        selected_material = material_combo.get()
        selected_transport_country = transport_country_combo.get()
        total_weight = get_total_weight()
        mat_idx = materials.index(selected_material)
        material_emission_lists = [steel_list, aluminium_list, cement_list, plastic_list, carbon_fiber_list]

        current_country_index = country_list.index(selected_transport_country)
        current_emission = custom_emission_factors.get("materials", {}).get(selected_material)
        if current_emission is None:
            current_emission = float(material_emission_lists[mat_idx][current_country_index])
        current_total_emission = current_emission * total_weight

        # Only suggest alternative countries for the same material
        best_option = None
        min_total_emission = float('inf')
        manuf_coords = country_coords_cache.get(selected_transport_country)
        for idx, country in enumerate(country_list):
            if country == selected_transport_country:
                continue
            country_coords = country_coords_cache.get(country)
            if not country_coords or not manuf_coords:
                continue
            distance_km = geodesic(manuf_coords, country_coords).km
            emission = custom_emission_factors.get("materials", {}).get(selected_material)
            if emission is None:
                value = material_emission_lists[mat_idx][idx]
                try:
                    emission = float(str(value).replace(" ", ""))
                except (ValueError, TypeError):
                    continue 
            transport_emission_factor = 0.01  
            transport_emission = total_weight * distance_km * transport_emission_factor / 1000
            total_emission = emission * total_weight + transport_emission
            if total_emission < min_total_emission:
                min_total_emission = total_emission
                best_option = {
                    "country": country,
                    "distance_km": distance_km,
                    "material_emission": emission,
                    "transport_emission": transport_emission,
                    "total_emission": total_emission,
                    "transport_mode": "Sea"
                }
        if best_option and best_option["total_emission"] < current_total_emission:
            abs_red = current_total_emission - best_option["total_emission"]
            pct_red = abs_red / current_total_emission * 100 if current_total_emission else 0
            suggestions.append((
                f"Source {selected_material} from {best_option['country']} ({best_option['transport_mode']})",
                current_total_emission, best_option["total_emission"], abs_red, pct_red
            ))

        # --- Transport: Suggest a more efficient mode among similar/available options (unchanged) ---
        try:
            if transport_row_widgets:
                row = transport_row_widgets[0]
                selected_transport = row['transport_combo'].get()
                selected_variation = row['variation_combo'].get()
                selected_fuel = row['fuel_combo'].get()
                idx = -1
                current_emission = None

                # Example for Van
                if selected_transport.lower() == "van":
                    idx = van_list.index(selected_variation) if selected_variation in van_list else -1
                    if selected_fuel == "Diesel" and idx != -1:
                        current_emission = float(van_diesel_list[idx])
                    elif selected_fuel == "Petrol" and idx != -1:
                        current_emission = float(van_petrol_list[idx])
                    elif selected_fuel == "Battery" and idx != -1:
                        current_emission = float(van_battery_list[idx])

                    # Suggest a lower-emission fuel for the same van type
                    alternatives = []
                    for fuel, lst in [("Diesel", van_diesel_list), ("Petrol", van_petrol_list), ("Battery", van_battery_list)]:
                        emission = float(lst[idx])
                        if emission < current_emission:
                            alternatives.append((fuel, emission))
                    if alternatives:
                        suggested_fuel, suggested_emission = min(alternatives, key=lambda x: abs(x[1] - current_emission))
                        total_weight = get_total_weight()
                        try:
                            distance_text = row['result_label'].cget("text")
                            distance = float(distance_text.split(":")[1].split("km")[0].strip().replace(",", ""))
                        except Exception:
                            distance = 1000  # fallback
                        scenario_emission = total_weight * (distance / 1000) * suggested_emission
                        abs_red = values[1] - scenario_emission
                        pct_red = abs_red / values[1] * 100 if values[1] else 0
                        suggestions.append((
                            f"Switch to Van ({suggested_fuel})",
                            values[1], scenario_emission, abs_red, pct_red
                        ))
                # Add similar logic for other transport types (HGV, Rail, etc.) as needed
        except Exception:
            pass

        # --- Machine: Suggest a more efficient machine of the same type (unchanged) ---
        try:
            if machine_parts_data:
                # Find the current machine and its energy use
                all_selected_machines = []
                all_current_energies = []
                for part_data in machine_parts_data:
                    if "rows" in part_data:
                        for row in part_data["rows"]:
                            selected_machine = row.get("machine_type", "")
                            if selected_machine in machine_value_list:
                                idx = machine_value_list.index(selected_machine)
                                current_energy = float(specific_machine_energy_use_list[idx])
                                all_selected_machines.append(selected_machine)
                                all_current_energies.append(current_energy)
                # Find the most common selected machine (or just use the first)
                if all_selected_machines:
                    selected_machine = all_selected_machines[0]
                    current_energy = all_current_energies[0]
                    # Find alternatives
                    alternatives = []
                    for i, machine in enumerate(machine_value_list):
                        if machine == selected_machine:
                            continue
                        energy = float(specific_machine_energy_use_list[i])
                        if energy < current_energy:
                            alternatives.append((machine, energy))
                    if alternatives:
                        suggested_machine, suggested_energy = min(alternatives, key=lambda x: abs(x[1] - current_energy))
                        selected_manufacturing_country = manufacturing_country_combo.get()
                        country_index = country_list.index(selected_manufacturing_country)
                        electricity_emission = float(electricity_list[country_index])
                        # Sum scenario emission for all rows
                        scenario_emission = 0.0
                        for part_data in machine_parts_data:
                            if "rows" in part_data:
                                for row in part_data["rows"]:
                                    weight_after = get_weight_after_total()
                                    weight_after_quantity = float(weight_after_quantity_entry.get()) if weight_after_quantity_entry.get() else 1.0
                                    machining_time = float(row.get("machining_time", "10")) if row.get("machining_time") else 10
                                    scenario_emission += suggested_energy * weight_after * weight_after_quantity * (machining_time / 60) * electricity_emission
                        abs_red = values[2] - scenario_emission
                        pct_red = abs_red / values[2] * 100 if values[2] else 0
                        suggestions.append((
                            f"Switch to {suggested_machine}",
                            values[2], scenario_emission, abs_red, pct_red
                        ))
        except Exception:
            pass

        # --- Packaging: Suggest lower-emission packaging of same class/size (unchanged) ---
        try:
            selected_packaging = packaging_dropdown.get()
            normalized_types = [t.strip().lower() for t in packaging_types_list]
            if selected_packaging.strip().lower() in normalized_types:
                idx = normalized_types.index(selected_packaging.strip().lower())
                current_emission = float(packaging_box_frame_list[idx]) if packaging_box_frame_list[idx] != "N/A" else float('inf')
                alternatives = []
                for i, emission in enumerate(packaging_box_frame_list):
                    if i == idx or emission == "N/A":
                        continue
                    emission = float(emission)
                    if emission < current_emission:
                        alternatives.append((packaging_types_list[i], emission))
                if alternatives:
                    suggested_pack, suggested_emission = min(alternatives, key=lambda x: abs(x[1] - current_emission))
                    weight_after = get_weight_after_total()
                    scenario_emission = suggested_emission * weight_after / 1000.0
                    abs_red = values[3] - scenario_emission
                    pct_red = abs_red / values[3] * 100 if values[3] else 0
                    suggestions.append((
                        f"Switch to {suggested_pack}",
                        values[3], scenario_emission, abs_red, pct_red
                    ))
        except Exception:
            pass

        # --- Recycled: Suggest increasing recycled content if low (unchanged) ---
        if "Recycled Emission" in labels:
            idx = labels.index("Recycled Emission")
            if idx < len(values) and values[idx] < 0.05 * total:
                suggestions.append((
                    "Increase recycled content in product",
                    values[idx], values[idx] * 2, values[idx], 100.0
                ))

        return suggestions

    def export_profile_to_excel():
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        file = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Export Profile to Excel"
        )
        if not file:
            return  # User cancelled

        wb = Workbook()
        ws = wb.active
        ws.title = "SPHERE Data"

        # --- Styling ---
        header_font = Font(bold=True, size=12)
        subheader_font = Font(bold=True, size=11)
        header_fill = PatternFill("solid", fgColor="A8D5BA")
        instruction_fill = PatternFill("solid", fgColor="FFFFCC")

        # ---- Instructions ----
        ws.merge_cells('A1:K1')
        ws['A1'] = "SPHERE Carbon Emission Calculator - Data Entry Template"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].fill = header_fill

        ws.merge_cells('A2:K2')
        ws['A2'] = "Instructions: Enter the number of rows needed in each quantity field, then fill in the data for each row."
        ws['A2'].font = Font(italic=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A2'].fill = instruction_fill

        # ---- Basic Info Section ----
        row = 4
        ws['A'+str(row)] = "Basic Information"
        ws['A'+str(row)].font = header_font
        ws['A'+str(row)].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')

        row += 1
        ws['A'+str(row)] = "Product Name"
        ws['B'+str(row)] = profile_data.get("product_name", "")

        row += 1
        ws['A'+str(row)] = "Description"
        ws['B'+str(row)] = profile_data.get("description", "")

        # ---- Material Information and Packaging & Recycling Side-by-Side ----
        row += 2

        # Left side: Material Information
        ws['A'+str(row)] = "Material Information"
        ws['A'+str(row)].font = header_font
        ws['A'+str(row)].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')

        # Right side: Packaging & Recycling
        ws['D'+str(row)] = "Packaging & Recycling"
        ws['D'+str(row)].font = header_font
        ws['D'+str(row)].fill = header_fill
        ws.merge_cells(f'D{row}:E{row}')

        material_start_row = row + 1

        ws['A'+str(material_start_row)] = "Transport Country"
        ws['B'+str(material_start_row)] = profile_data.get("transport_country_combo", "")

        material_start_row += 1
        ws['A'+str(material_start_row)] = "Manufacturing Country"
        ws['B'+str(material_start_row)] = profile_data.get("manufacturing_country_combo", "")

        material_start_row += 1
        ws['A'+str(material_start_row)] = "Material"
        ws['B'+str(material_start_row)] = profile_data.get("material_combo", "")

        material_start_row += 1
        ws['A'+str(material_start_row)] = "Raw Weight (kg)"
        ws['B'+str(material_start_row)] = profile_data.get("weight_entry", "")

        material_start_row += 1
        ws['A'+str(material_start_row)] = "Raw Weight Quantity"
        ws['B'+str(material_start_row)] = profile_data.get("quantity_entry_weight", "")

        material_start_row += 1
        ws['A'+str(material_start_row)] = "Weight After Manufacturing (kg)"
        ws['B'+str(material_start_row)] = profile_data.get("weight_after_entry", "")

        material_start_row += 1
        ws['A'+str(material_start_row)] = "Weight After Manufacturing Quantity"
        ws['B'+str(material_start_row)] = profile_data.get("weight_after_quantity_entry", "")

        # Packaging & Recycling fields (placed to the right of Material Information)
        packaging_row = row + 1
        ws['D'+str(packaging_row)] = "Packaging Type"
        ws['E'+str(packaging_row)] = profile_data.get("packaging_dropdown", "")

        packaging_row += 1
        ws['D'+str(packaging_row)] = "Recycled Metal Type"
        ws['E'+str(packaging_row)] = profile_data.get("metal_recycling_dropdown", "")

        # ---- TRANSPORTATION AND MACHINE SECTIONS ----
        row = max(material_start_row, packaging_row) + 2

        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "Note: When adding more rows, copy an existing row to maintain the validations."
        ws[f'A{row}'].font = Font(color="FF0000", italic=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        ws.merge_cells(f'H{row}:K{row}')
        ws[f'H{row}'] = "Note: 1 = machine part. Multiple rows of Part 1 = Process for that part. Same logic applies to other parts."
        ws[f'H{row}'].font = Font(color="FF0000", italic=True)
        ws[f'H{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws[f'H{row}'].fill = PatternFill("solid", fgColor="FFFFCC")

        row += 1
        ws['A'+str(row)] = "Transportation Options"
        ws['A'+str(row)].font = header_font
        ws['A'+str(row)].fill = header_fill
        ws.merge_cells(f'A{row}:F{row}')

        right_col = 8
        ws.cell(row=row, column=right_col).value = "Machine Processes"
        ws.cell(row=row, column=right_col).font = header_font
        ws.cell(row=row, column=right_col).fill = header_fill
        ws.merge_cells(f'{get_column_letter(right_col)}{row}:{get_column_letter(right_col+3)}{row}')

        # Transport quantity input
        transport_qty_row = row + 1
        ws['A'+str(transport_qty_row)] = "Number of Transport Rows"
        ws['B'+str(transport_qty_row)] = profile_data.get("quantity_entry", "1")
        ws['B'+str(transport_qty_row)].font = Font(bold=True)
        ws['C'+str(transport_qty_row)] = "← Enter quantity here"
        ws['C'+str(transport_qty_row)].font = Font(italic=True, color="FF0000")

        # Machine quantity input - same row as transport quantity
        machine_qty_row = transport_qty_row
        ws.cell(row=machine_qty_row, column=right_col).value = "Number of Machine Parts"
        ws.cell(row=machine_qty_row, column=right_col+1).value = profile_data.get("machine_process_quantity_entry", "1")
        ws.cell(row=machine_qty_row, column=right_col+1).font = Font(bold=True)
        ws.cell(row=machine_qty_row, column=right_col+2).value = "← Enter quantity here"
        ws.cell(row=machine_qty_row, column=right_col+2).font = Font(italic=True, color="FF0000")

        # Transport headers
        transport_header_row = transport_qty_row + 1
        for col, header in enumerate(['Transport Type', 'Variation', 'Fuel', 'Origin', 'Destination', 'Mode']):
            cell = ws.cell(row=transport_header_row, column=col+1)
            cell.value = header
            cell.font = subheader_font
            cell.fill = header_fill

        # Machine headers - same row as transport headers
        machine_header_row = transport_header_row
        for col, header in enumerate(['Part Number', 'Machine Type', 'Machining Time (min)', 'Notes']):
            cell = ws.cell(row=machine_header_row, column=right_col+col)
            cell.value = header
            cell.font = subheader_font
            cell.fill = header_fill
        # --- Business Travel Options (duplicate of Transport section) ---
        bt_row = transport_header_row + 15  # space below transport section
        ws['A'+str(bt_row)] = "Business Travel Options"
        ws['A'+str(bt_row)].font = header_font
        ws['A'+str(bt_row)].fill = header_fill
        ws.merge_cells(f'A{bt_row}:F{bt_row}')

        # Quantity input for Business Travel
        bt_qty_row = bt_row + 1
        ws['A'+str(bt_qty_row)] = "Number of Business Travel Rows"
        ws['B'+str(bt_qty_row)] = "1"
        ws['B'+str(bt_qty_row)].font = Font(bold=True)
        ws['C'+str(bt_qty_row)] = "← Enter quantity here"
        ws['C'+str(bt_qty_row)].font = Font(italic=True, color="FF0000")

        # Business Travel headers (same as Transport)
        bt_header_row = bt_qty_row + 1
        for col, header in enumerate(['Travel Type', 'Variation', 'Fuel', 'Origin', 'Destination', 'Mode']):
            cell = ws.cell(row=bt_header_row, column=col+1)
            cell.value = header
            cell.font = subheader_font
            cell.fill = header_fill


        # Transport data rows
        transport_data_start = transport_header_row + 1
        transport_rows = profile_data.get("transport_rows", [])
        for i in range(max(5, len(transport_rows))):
            row_idx = transport_data_start + i
            if i < len(transport_rows):
                row_data = transport_rows[i]
                ws.cell(row=row_idx, column=1).value = row_data.get("transport", "")
                ws.cell(row=row_idx, column=2).value = row_data.get("variation", "")
                ws.cell(row=row_idx, column=3).value = row_data.get("fuel", "")
                ws.cell(row=row_idx, column=4).value = row_data.get("origin", "")
                ws.cell(row=row_idx, column=5).value = row_data.get("destination", "")
                ws.cell(row=row_idx, column=6).value = row_data.get("mode", "")
            # else: leave blank for template


        # Machine data rows - aligned with transport rows
        machine_process_rows = profile_data.get("machine_process_rows", [])
        for i in range(max(5, len(machine_process_rows))):
            row_idx = transport_data_start + i
            if i < len(machine_process_rows):
                row_data = machine_process_rows[i]
                ws.cell(row=row_idx, column=right_col).value = row_data.get("part", "")
                ws.cell(row=row_idx, column=right_col+1).value = row_data.get("machine_type", "")
                ws.cell(row=row_idx, column=right_col+2).value = row_data.get("machining_time", "")
                ws.cell(row=row_idx, column=right_col+3).value = ""  
            # else: leave blank for template

        # Adjust column widths for readability
        column_widths = {
            1: 18,  # A
            2: 18,  # B
            3: 15,  # C
            4: 18,  # D
            5: 18,  # E
            6: 12,  # F
            7: 2,   # G (separator)
            8: 15,  # H
            9: 22,  # I
            10: 18, # J
            11: 15  # K
        }
        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        wb.save(file)
        messagebox.showinfo("Export Complete", f"Profile exported to {file}")

    def wrap_text(text, max_width, canvas, font_name="Helvetica", font_size=8):
        lines = []
        text = str(text)
        if canvas.stringWidth(text, font_name, font_size) > max_width:
            words = text.split()
            line = ""
            for word in words:
                test_line = f"{line} {word}".strip()
                if canvas.stringWidth(test_line, font_name, font_size) > max_width:
                    lines.append(line)
                    line = word
                else:
                    line = test_line
            lines.append(line)
        else:
            lines.append(text)
        return lines

    def export_profile_to_pdf():
        from datetime import datetime
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas as pdf_canvas
        from reportlab.lib.utils import ImageReader
        import tempfile
        import os

        static_field_map = {
            'transport_country_combo': 'Transport Country',
            'material_combo': 'Material',
            'weight_entry': 'Raw Weight (kg)',
            'quantity_entry_weight': 'Raw Weight Quantity',
            'weight_after_entry': 'Weight After Manufacturing (kg)',
            'weight_after_quantity_entry': 'Weight After Manufacturing Quantity',
            'packaging_dropdown': 'Packaging Type',
            'metal_recycling_dropdown': 'Recycled Metal Type',
            'manufacturing_country_combo': 'Manufacturing Country',
            'quantity_entry': 'Number of Transport Options'
        }

        file = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])
        if not file:
            return

        save_profile_inputs(profile_data, widgets)
        c = pdf_canvas.Canvas(file, pagesize=letter)
        y = 770

        # --- Draw logo at top right of page 1 ---
        logo_path = resource_path("SPHERE_logo.png")
        try:
            logo_width = 250
            logo_height = 250
            x = letter[0] - logo_width - 20
            y_logo = letter[1] - logo_height + 5
            c.drawImage(ImageReader(logo_path), x, y_logo, width=logo_width, height=logo_height, preserveAspectRatio=True, mask='auto')
        except Exception as e:
            print(f"[DEBUG] Could not add logo to PDF: {e}")

        # --- Title ---
        c.setFont("Helvetica-Bold", 16)
        c.setFillColorRGB(0.2, 0.4, 0.6)
        c.drawString(50, y, "Profile Report")
        c.setFillColorRGB(0, 0, 0)
        y -= 32

        # --- Profile Metadata ---
        c.setFont("Helvetica", 11)
        c.drawString(50, y, f"Product Name: {profile_data.get('product_name', '')}")
        y -= 18
        c.drawString(50, y, f"Description/Notes: {profile_data.get('description', '')}")
        y -= 18
        c.drawString(50, y, f"Date & Time of Export: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        y -= 18
        c.drawString(50, y, f"App Version: {APP_VERSION}")
        y -= 28

        # --- Threshold Info ---
        if threshold_enabled_var.get():
            threshold = threshold_value_var.get()
            total_emission = profile_data.get("total_emission", 0.0)
            try:
                threshold_val = float(threshold)
                status = "Below threshold" if float(total_emission) <= threshold_val else "Above threshold"
            except Exception:
                status = "Invalid threshold"
            c.drawString(50, y, f"Threshold Value: {threshold}")
            y -= 18
            c.drawString(50, y, f"Threshold Status: {status}")
            y -= 28

        # --- Custom Emission Factors Used ---
        if custom_emission_factors.get("materials"):
            c.setFont("Helvetica-Bold", 11)
            c.drawString(50, y, "Custom Emission Factors Used:")
            y -= 16
            c.setFont("Helvetica", 10)
            c.setFillColorRGB(0.93, 0.93, 0.93)
            c.rect(48, y-2, 300, 16, fill=1, stroke=0)
            c.setFillColorRGB(0, 0, 0)
            c.drawString(60, y, "Material")
            c.drawString(200, y, "Custom Factor (CO2e/kg)")
            y -= 14
            for mat, val in custom_emission_factors["materials"].items():
                c.drawString(60, y, str(mat))
                c.drawRightString(300, y, f"{val}")
                y -= 14
                if y < 100:
                    c.showPage()
                    y = 770
            y -= 18

        # --- Profile Fields (one per line) ---
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, y, "Profile Fields:")
        y -= 16
        c.setFont("Helvetica", 10)
        for k in static_field_map:
            v = profile_data.get(k, "")
            c.drawString(60, y, f"{static_field_map[k]}: {v}")
            y -= 14
        y -= 10  # Extra space before next section

        # --- Export chart as image (moved here) ---
        values = []
        labels = []
        colors = []
        filtered = []
        try:
            colors = ['#FF9999', '#66B3FF', '#99FF99', '#FFB347', '#FFD700']
            labels = [
                'Raw Material Emission',
                'Transport Emission',
                'Machine Emission',
                'Packaging Emission',
                'Recycled Emission'
            ]
            transport_total = round(
                sum(
                    round(extract_float(label.cget("text")), 2)
                    for label in total_emission_labels
                    if label and label.winfo_exists() and "Total Emission:" in label.cget("text")
                ),
                2
            )
            values = [
                extract_float(calculated_value_label.cget("text")),
                transport_total,
                sum(extract_float(part_data.get("carbon_emission", "0.0")) for part_data in machine_parts_data),
                extract_float(packaging_total_emission_label.cget("text")) if packaging_total_emission_label and packaging_total_emission_label.winfo_exists() else 0.0,
                -extract_float(recycled_emission_value_label.cget("text")) if recycled_emission_value_label and recycled_emission_value_label.winfo_exists() else 0.0
            ]
            filtered = [(label, value, color) for label, value, color in zip(labels, values, colors) if value != 0]
            if filtered:
                labels, values, colors = zip(*filtered)
                values = list(values)
            else:
                labels, values, colors = [], [], []
                values = []
            chart_type = chart_type_var.get()
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmpfile:
                save_current_chart_as_image(filtered, chart_type, tmpfile.name)
                c.setFont("Helvetica-Bold", 14)
                c.drawString(50, y, f"{chart_type} of Emission Contributions")
                y -= 20
                c.drawImage(ImageReader(tmpfile.name), 50, y-180, width=300, height=180, preserveAspectRatio=True, mask='auto')
            os.unlink(tmpfile.name)
            y -= 200 
        except Exception as e:
            print(f"[DEBUG] Failed to export chart image: {e}")

        # --- Emission Breakdown Section with highlight ---
        c.setFont("Helvetica-Bold", 13)
        c.setFillColorRGB(0.2, 0.4, 0.6)
        c.drawString(50, y, "Emission Breakdown")
        c.setFillColorRGB(0, 0, 0)
        y -= 22
        c.setFont("Helvetica-Bold", 10)
        c.setFillColorRGB(0.93, 0.93, 0.93)
        c.rect(48, y-2, 520, 16, fill=1, stroke=0)
        c.setFillColorRGB(0, 0, 0)

        category_x = 140
        value_x = 320
        percent_x = 500

        c.drawCentredString(category_x, y, "Category")
        c.drawCentredString(value_x, y, "Value (CO2eKg)")
        c.drawCentredString(percent_x, y, "Percentage (%)")
        y -= 16
        c.setFont("Helvetica", 10)

        total = sum(values) if sum(values) != 0 else 1  # Avoid division by zero
        max_idx = values.index(max(values))
        for i, (label, value) in enumerate(zip(labels, values)):
            percent = (value / total) * 100
            c.setFillColorRGB(0, 0, 0)
            c.drawCentredString(category_x, y, label)
            if i == max_idx:
                c.setFillColorRGB(1, 0, 0)  # Red for largest value
            else:
                c.setFillColorRGB(0, 0, 0)
            c.drawCentredString(value_x, y, f"{value:.2f}")
            c.setFillColorRGB(0, 0, 0)
            c.drawCentredString(percent_x, y, f"{percent:.1f}%")
            y -= 14
            if y < 100:
                c.showPage()
                y = 770
        c.setFont("Helvetica-Bold", 10)
        c.setFillColorRGB(0, 0, 0)
        total_emission_value = sum(values)
        total_emission_percent = 100.0  # Always 100% of itself
        c.drawCentredString(category_x, y, "Total Emission")
        c.drawCentredString(value_x, y, f"{total_emission_value:.2f}")
        c.drawCentredString(percent_x, y, f"{total_emission_percent:.1f}%")
        y -= 14
        c.setFont("Helvetica", 10) 
        if y < 100:
            c.showPage()
            y = 770
        c.setFillColorRGB(0, 0, 0)
        y -= 22

        # --- Highest Contributing Category ---
        if values:
            max_idx = values.index(max(values))
            if max_idx < len(labels):
                c.setFont("Helvetica-Bold", 11)
                c.drawString(50, y, f"Highest Contributing Category: {labels[max_idx]}")
            else:
                c.setFont("Helvetica-Bold", 11)
                c.drawString(50, y, "Highest Contributing Category: N/A")
        else:
            c.setFont("Helvetica-Bold", 11)
            c.drawString(50, y, "Highest Contributing Category: N/A")
        y -= 32

        # --- Machine Process Details Section ---
        c.setFont("Helvetica-Bold", 13)
        c.setFillColorRGB(0.2, 0.4, 0.6)
        c.drawString(50, y, "Machine Process Details")
        c.setFillColorRGB(0, 0, 0)
        y -= 22

        # Table header
        c.setFont("Helvetica-Bold", 8)
        c.setFillColorRGB(0.93, 0.93, 0.93)
        c.rect(48, y-2, 520, 14, fill=1, stroke=0)
        c.setFillColorRGB(0, 0, 0)
        part_x = 100
        machine_x = 200
        energy_x = 300
        time_x = 400
        emission_x = 500

        c.drawCentredString(part_x, y, "Part")
        c.drawCentredString(machine_x, y, "Machine Type")
        c.drawCentredString(energy_x, y, "Energy Use (kWh/kg)")
        c.drawCentredString(time_x, y, "Time (min)")
        c.drawCentredString(emission_x, y, "Emission (CO2eKg)")
        y -= 14

        c.setFont("Helvetica", 8)

        for part_idx, part_data in enumerate(machine_parts_data, 1):
            c.setFont("Helvetica-Bold", 8)
            c.drawCentredString(part_x, y, f"Part {part_idx}")
            c.setFont("Helvetica", 8)
            y -= 12  

            part_emission = float(part_data.get("carbon_emission", "0.0"))
            c.drawCentredString(part_x, y, f"Total Part Emission: {part_emission:.2f} CO2eKg")
            y -= 12

            if "rows" in part_data:
                for row_idx, row in enumerate(part_data["rows"], 1):
                    machine_type = row.get("machine_type", "")
                    energy_use = row.get("energy_use", "0.0")
                    machining_time = row.get("machining_time", "")
                    carbon_emission = row.get("carbon_emission", "0.0")

                    if c.stringWidth(machine_type, "Helvetica", 8) > 90:
                        lines = []
                        words = machine_type.split()
                        line = ""
                        for word in words:
                            test_line = f"{line} {word}".strip()
                            if c.stringWidth(test_line, "Helvetica", 8) > 90:
                                lines.append(line)
                                line = word
                            else:
                                line = test_line
                        lines.append(line)
                        c.drawCentredString(part_x, y, f"{part_idx}.{row_idx}")
                        for l in lines:
                            c.drawCentredString(machine_x, y, l)
                            y -= 10
                        c.drawCentredString(energy_x, y+10, str(energy_use))
                        c.drawCentredString(time_x, y+10, machining_time)
                        c.drawCentredString(emission_x, y+10, carbon_emission)
                        y -= 2
                    else:
                        c.drawCentredString(part_x, y, f"{part_idx}.{row_idx}")
                        c.drawCentredString(machine_x, y, machine_type)
                        c.drawCentredString(energy_x, y, str(energy_use))
                        c.drawCentredString(time_x, y, machining_time)
                        c.drawCentredString(emission_x, y, carbon_emission)
                        y -= 10

                    if y < 100:
                        c.showPage()
                        y = 770
                        c.setFont("Helvetica", 8)

            c.setLineWidth(0.5)
            c.line(60, y-3, 550, y-3)
            y -= 12  
            if y < 100:
                c.showPage()
                y = 770
                c.setFont("Helvetica", 8)

        y -= 8

        # --- Suggestions & Impact Section ---
        c.showPage()
        y = 770
        c.setFont("Helvetica-Bold", 13)
        c.setFillColorRGB(0.2, 0.4, 0.6)
        c.drawString(50, y, "Suggestions & Impact")
        c.setFillColorRGB(0, 0, 0)
        y -= 18
        c.setFont("Helvetica", 8)
        explanation_lines = [
            "Baseline: Current emission value.",
            "Scenario: Projected emission if the suggestion is applied.",
            "Abs. Red.: Absolute reduction (Baseline - Scenario).",
            "% Red.: Percent reduction compared to Baseline."
        ]
        for line in explanation_lines:
            c.drawString(50, y, line)
            y -= 12
        y -= 10
        c.setFont("Helvetica-Bold", 10)
        c.setFillColorRGB(0.93, 0.93, 0.93)
        c.rect(48, y-2, 520, 16, fill=1, stroke=0)
        c.setFillColorRGB(0, 0, 0)

        strategy_x = 150
        baseline_x = 260
        scenario_x = 340
        abs_red_x = 420
        pct_red_x = 500

        c.drawCentredString(strategy_x, y, "Strategy")
        c.drawCentredString(baseline_x, y, "Baseline")
        c.drawCentredString(scenario_x, y, "Scenario")
        c.drawCentredString(abs_red_x, y, "Abs. Red.")
        c.drawCentredString(pct_red_x, y, "% Red.")
        y -= 16
        c.setFont("Helvetica", 10)

        suggestions = get_emission_reduction_suggestions(labels, values, total)
        if suggestions:
            for s in suggestions:
                c.drawString(60, y, s[0])
                c.drawCentredString(baseline_x, y, f"{s[1]:.2f}")
                c.drawCentredString(scenario_x, y, f"{s[2]:.2f}")
                c.drawCentredString(abs_red_x, y, f"{s[3]:.2f}")
                c.drawCentredString(pct_red_x, y, f"{s[4]:.1f}%")
                y -= 14
                if y < 100:
                    c.showPage()
                    y = 770
        else:
            c.drawString(60, y, "Your emissions are well balanced. Keep up the good work!")
            y -= 14

        gap_between_sections = 32  
        y -= gap_between_sections

        # --- Calculate best material & country for suggestion ---
        selected_material = material_combo.get()
        selected_country = transport_country_combo.get()
        total_weight = get_total_weight()
        material_emission_lists = [steel_list, aluminium_list, cement_list, plastic_list, carbon_fiber_list]
        best = get_best_material_and_country(
            total_weight,
            country_coords_cache,
            country_list,
            materials,
            material_emission_lists,
            custom_emission_factors,
            selected_country,
            selected_material 
        )

        suggestions = get_emission_reduction_suggestions(labels, values, total) # Store the suggestions

        # --- Suggested Material & Source Country Section (Table Style) ---
        c.setFont("Helvetica-Bold", 13)  
        c.setFillColorRGB(0.2, 0.4, 0.6) 
        c.drawString(50, y, "Changed After Suggested Material & Source Country")
        c.setFillColorRGB(0, 0, 0) 
        y -= 22

        # Table header style - smaller font
        c.setFont("Helvetica-Bold", 9) 
        c.setFillColorRGB(0.93, 0.93, 0.93)
        c.rect(48, y-2, 520, 16, fill=1, stroke=0)
        c.setFillColorRGB(0, 0, 0)

        mat_x = 110      
        country_x = 220   
        total_x = 320     
        mat_em_x = 400   
        trans_em_x = 470  
        dist_x = 535      

        c.drawCentredString(mat_x, y, "Material")
        c.drawCentredString(country_x, y, "Country")
        c.drawCentredString(total_x, y, "Total (CO2eKg)")
        c.drawCentredString(mat_em_x, y, "Material")
        c.drawCentredString(trans_em_x, y, "Transport")
        c.drawCentredString(dist_x, y, "Dist (km)")
        y -= 16
        c.setFont("Helvetica", 9)  

        if best:
            matched_suggestion = None
            for s in suggestions:
                if "Source" in s[0] and best['material'] in s[0] and best['source_country'] in s[0]:
                    matched_suggestion = s
                    break
    
            highlight = (best["material"] != selected_material) or (best["source_country"] != selected_country)
            if highlight:
                c.setFillColorRGB(1, 0, 0)
            else:
                c.setFillColorRGB(0, 0, 0)
        
            c.drawCentredString(mat_x, y, best['material'])
            c.drawCentredString(country_x, y, best['source_country'])
            c.drawCentredString(total_x, y, f"{best['total_emission']:.2f}")
            c.drawCentredString(mat_em_x, y, f"{best['material_emission']:.2f}")
            c.drawCentredString(trans_em_x, y, f"{best['transport_emission']:.2f}")
            c.drawCentredString(dist_x, y, f"{best['distance_km']:.0f}")
            c.setFillColorRGB(0, 0, 0)
            y -= 14
        else:
            c.drawCentredString(mat_x, y, "N/A")
            y -= 14
        
        # --- Changed Machine Processes After Suggestions with highlight ---
        gap_between_sections = 32
        y -= gap_between_sections
        c.setFont("Helvetica-Bold", 13)
        c.setFillColorRGB(0.2, 0.4, 0.6)
        c.drawString(50, y, "Changed Machine Processes After Suggestions")
        c.setFillColorRGB(0, 0, 0)
        y -= 22
        c.setFont("Helvetica-Bold", 8)
        c.setFillColorRGB(0.93, 0.93, 0.93)
        c.rect(48, y-2, 520, 14, fill=1, stroke=0)
        c.setFillColorRGB(0, 0, 0)
        part_x = 80
        row_x = 120
        orig_machine_x = 200
        sugg_machine_x = 320
        orig_emission_x = 440
        sugg_emission_x = 520

        c.drawCentredString(part_x, y, "Part")
        c.drawCentredString(row_x, y, "Process")
        c.drawCentredString(orig_machine_x, y, "Original Machine")
        c.drawCentredString(sugg_machine_x, y, "Suggested Machine")
        c.drawCentredString(orig_emission_x, y, "Orig. Emission")
        c.drawCentredString(sugg_emission_x, y, "Sugg. Emission")
        y -= 12
        c.setFont("Helvetica", 8)

        changed_machine_rows = []
        for part_idx, part_data in enumerate(machine_parts_data, 1):
            if "rows" in part_data:
                for row_idx, row in enumerate(part_data["rows"], 1):
                    selected_machine = row.get("machine_type", "")
                    if not selected_machine or selected_machine not in machine_value_list:
                        continue
                    country_index = country_list.index(manufacturing_country_combo.get())
                    weight_after = get_weight_after_total()
                    weight_after_quantity = float(weight_after_quantity_entry.get()) if weight_after_quantity_entry.get() else 1.0
                    machining_time = float(row.get("machining_time", "10")) if row.get("machining_time") else 10
                    best_machine, best_emission = suggest_optimal_machine(
                        selected_machine, weight_after, weight_after_quantity, machining_time, country_index
                    )
                    changed_machine_rows.append({
                        "part": f"{part_idx}",
                        "row": row_idx,
                        "original_machine": selected_machine,
                        "suggested_machine": best_machine,
                        "original_emission": row.get("carbon_emission", "0.0"),
                        "suggested_emission": f"{best_emission:.2f}"
                    })

        for r in changed_machine_rows:
            y_start = y
            orig_lines = []
            sugg_lines = []
            for text, maxw, lines in [
                (r["original_machine"], 90, orig_lines),
                (r["suggested_machine"], 90, sugg_lines)
            ]:
                if c.stringWidth(text, "Helvetica", 8) > maxw:
                    words = text.split()
                    line = ""
                    for word in words:
                        test_line = f"{line} {word}".strip()
                        if c.stringWidth(test_line, "Helvetica", 8) > maxw:
                            lines.append(line)
                            line = word
                        else:
                            line = test_line
                    lines.append(line)
                else:
                    lines.append(text)
            max_lines = max(len(orig_lines), len(sugg_lines))
            c.setFillColorRGB(0, 0, 0)
            c.drawCentredString(part_x, y, r["part"])
            c.drawCentredString(row_x, y, str(r["row"]))
            # Original Machine
            c.setFillColorRGB(0, 0, 0)
            c.drawCentredString(orig_machine_x, y, orig_lines[0])
            # Suggested Machine
            if r["original_machine"] != r["suggested_machine"]:
                c.setFillColorRGB(1, 0, 0)  # Red
            else:
                c.setFillColorRGB(0, 0, 0)
            c.drawCentredString(sugg_machine_x, y, sugg_lines[0])
            # Original Emission
            c.setFillColorRGB(0, 0, 0)
            c.drawCentredString(orig_emission_x, y, str(r["original_emission"]))
            # Suggested Emission
            if r["original_emission"] != r["suggested_emission"]:
                c.setFillColorRGB(1, 0, 0)  # Red
            else:
                c.setFillColorRGB(0, 0, 0)
            c.drawCentredString(sugg_emission_x, y, str(r["suggested_emission"]))
            y -= 10
            c.setFillColorRGB(0, 0, 0)
            if y < 100:
                c.showPage()
                y = 770
                c.setFont("Helvetica", 8)

        # --- Changed Recycling After Suggestions with highlight ---
        gap_between_sections = 32
        y -= gap_between_sections
        c.setFont("Helvetica-Bold", 13)
        c.setFillColorRGB(0.2, 0.4, 0.6)
        c.drawString(50, y, "Changed Recycling After Suggestions")
        c.setFillColorRGB(0, 0, 0)
        y -= 22
        c.setFont("Helvetica-Bold", 8)
        c.setFillColorRGB(0.93, 0.93, 0.93)
        c.rect(48, y-2, 520, 14, fill=1, stroke=0)
        c.setFillColorRGB(0, 0, 0)
        orig_type_x = 120
        sugg_type_x = 260
        orig_emission_x = 380
        sugg_emission_x = 500

        c.drawCentredString(orig_type_x, y, "Original Metal Type")
        c.drawCentredString(sugg_type_x, y, "Suggested Metal Type")
        c.drawCentredString(orig_emission_x, y, "Original Emission")
        c.drawCentredString(sugg_emission_x, y, "Suggested Emission")
        y -= 12
        c.setFont("Helvetica", 8)

        original_recycling_type = profile_data.get("metal_recycling_dropdown", "")
        normalized_types = [t.strip().lower() for t in metal_recyling_types_list]
        try:
            orig_idx = normalized_types.index(original_recycling_type.strip().lower())
        except Exception:
            orig_idx = -1
        original_emission = metal_recyling_emission_list[orig_idx] if orig_idx != -1 else "N/A"
        min_emission = min([e for e in metal_recyling_emission_list if isinstance(e, (float, int))], default=original_emission)
        try:
            min_idx = metal_recyling_emission_list.index(min_emission)
            suggested_type = metal_recyling_types_list[min_idx]
        except Exception:
            suggested_type = original_recycling_type

        orig_type_lines = wrap_text(original_recycling_type, 90, c)
        sugg_type_lines = wrap_text(suggested_type, 90, c)
        max_lines = max(len(orig_type_lines), len(sugg_type_lines))
        for i in range(max_lines):
            # Original Metal Type
            c.setFillColorRGB(0, 0, 0)
            c.drawCentredString(orig_type_x, y, orig_type_lines[i] if i < len(orig_type_lines) else "")
            # Suggested Metal Type
            if (original_recycling_type != suggested_type) and i == 0:
                c.setFillColorRGB(1, 0, 0)
            else:
                c.setFillColorRGB(0, 0, 0)
            c.drawCentredString(sugg_type_x, y, sugg_type_lines[i] if i < len(sugg_type_lines) else "")
            # Original Emission
            c.setFillColorRGB(0, 0, 0)
            if i == 0:
                c.drawCentredString(orig_emission_x, y, str(original_emission))
                # Suggested Emission
                if str(original_emission) != str(min_emission):
                    c.setFillColorRGB(1, 0, 0)
                else:
                    c.setFillColorRGB(0, 0, 0)
                c.drawCentredString(sugg_emission_x, y, str(min_emission))
            y -= 10
            c.setFillColorRGB(0, 0, 0)
        if y < 100:
            c.showPage()
            y = 770

        # --- Changed Packaging After Suggestions with highlight ---
        y -= gap_between_sections
        c.setFont("Helvetica-Bold", 13)
        c.setFillColorRGB(0.2, 0.4, 0.6)
        c.drawString(50, y, "Changed Packaging After Suggestions")
        c.setFillColorRGB(0, 0, 0)
        y -= 22
        c.setFont("Helvetica-Bold", 8)
        c.setFillColorRGB(0.93, 0.93, 0.93)
        c.rect(48, y-2, 520, 14, fill=1, stroke=0)
        c.setFillColorRGB(0, 0, 0)
        orig_pack_x = 120
        sugg_pack_x = 260
        orig_pack_emission_x = 380
        sugg_pack_emission_x = 500

        c.drawCentredString(orig_pack_x, y, "Original Packaging Type")
        c.drawCentredString(sugg_pack_x, y, "Suggested Packaging Type")
        c.drawCentredString(orig_pack_emission_x, y, "Original Emission")
        c.drawCentredString(sugg_pack_emission_x, y, "Suggested Emission")
        y -= 12
        c.setFont("Helvetica", 8)

        original_packaging_type = profile_data.get("packaging_dropdown", "")
        normalized_pack_types = [t.strip().lower() for t in packaging_types_list]
        try:
            orig_pack_idx = normalized_pack_types.index(original_packaging_type.strip().lower())
        except Exception:
            orig_pack_idx = -1
        original_pack_emission = packaging_box_frame_list[orig_pack_idx] if orig_pack_idx != -1 else "N/A"
        min_pack_emission = min([e for e in packaging_box_frame_list if isinstance(e, (float, int))], default=original_pack_emission)
        try:
            min_pack_idx = packaging_box_frame_list.index(min_pack_emission)
            suggested_pack_type = packaging_types_list[min_pack_idx]
        except Exception:
            suggested_pack_type = original_packaging_type

        orig_pack_lines = wrap_text(original_packaging_type, 90, c)
        sugg_pack_lines = wrap_text(suggested_pack_type, 90, c)
        max_lines = max(len(orig_pack_lines), len(sugg_pack_lines))
        for i in range(max_lines):
            # Original Packaging Type
            c.setFillColorRGB(0, 0, 0)
            c.drawCentredString(orig_pack_x, y, orig_pack_lines[i] if i < len(orig_pack_lines) else "")
            # Suggested Packaging Type
            if (original_packaging_type != suggested_pack_type) and i == 0:
                c.setFillColorRGB(1, 0, 0)
            else:
                c.setFillColorRGB(0, 0, 0)
            c.drawCentredString(sugg_pack_x, y, sugg_pack_lines[i] if i < len(sugg_pack_lines) else "")
            # Original Emission
            c.setFillColorRGB(0, 0, 0)
            if i == 0:
                c.drawCentredString(orig_pack_emission_x, y, str(original_pack_emission))
                # Suggested Emission
                if str(original_pack_emission) != str(min_pack_emission):
                    c.setFillColorRGB(1, 0, 0)
                else:
                    c.setFillColorRGB(0, 0, 0)
                c.drawCentredString(sugg_pack_emission_x, y, str(min_pack_emission))
            y -= 10
            c.setFillColorRGB(0, 0, 0)
        if y < 100:
            c.showPage()
            y = 770

        c.save()
        messagebox.showinfo("Export Complete", f"Profile exported to {file}")

    def update_dependent_dropdowns_for_excel_template():
        """Creates and returns dictionaries mapping transport types to their respective variations and fuels"""
        transport_variations = {
            "Van": van_list,
            "HGV (Diesel)": HGV_diesel_list,
            "HGV Refrigerated (Diesel)": HGV_R_diesel_list,
            "Freight Flights": flight_list,
            "Rail": rail_list,
            "Sea Tanker": sea_tanker_list,
            "Cargo Ship": cargo_list
        }
    
        transport_fuels = {
            "Van": ["Diesel", "Petrol", "Battery"],
            "HGV (Diesel)": ["0% Laden", "50% Laden", "100% Laden"],
            "HGV Refrigerated (Diesel)": ["0% Laden", "50% Laden", "100% Laden"],
            "Freight Flights": ["International", "<1500km", ">1500km"],
            "Rail": ["Freight Train"],
            "Sea Tanker": [],  # For Sea Tanker, variation determines fuel
            "Cargo Ship": []   # For Cargo Ship, variation determines fuel
        }
    
        # Special case mappings for Sea Tanker and Cargo Ship
        sea_tanker_variation_fuels = {
            "Crude Tanker": extract_list(crude_tanker_cells),
            "Products Tanker": extract_list(products_tanker_cells),
            "Chemical Tanker": extract_list(chemical_tanker_cells),
            "LNG Tanker": extract_list(LNG_tanker_cells),
            "LPG Tanker": extract_list(LPG_tanker_cells)
        }
    
        cargo_ship_variation_fuels = {
            "Bulk Carrier": extract_list(Bulk_carrier_cells),
            "General Cargo": extract_list(General_cargo_cells),
            "Container Ship": extract_list(Container_ship_cells),
            "Vehicle Transport": extract_list(Vehicle_transport_cells),
            "RoRo Ferry": extract_list(Roro_ferry_cells)
        }
    
        return transport_variations, transport_fuels, sea_tanker_variation_fuels, cargo_ship_variation_fuels

    def import_from_excel():
        global is_importing
        is_importing = True
        """Imports data from the Excel template with proper handling of multiple processes per part"""
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx")],
            title="Import Excel Data"
        )

        if not file_path:
            return

        try:
            # Show loading cursor
            root.config(cursor="wait")
            root.update_idletasks()

            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            print("[DEBUG] Starting Excel import from:", file_path)

            # Create profile data dictionary
            profile_data = {}

            profile_data["product_name"] = str(ws['B5'].value or "")
            profile_data["description"] = str(ws['B6'].value or "")
            print(f"[DEBUG] Product: {profile_data['product_name']}, Description: {profile_data['description']}")

            profile_data["transport_country_combo"] = str(ws['B9'].value or "")
        
            # Get the transport country immediately for later use
            selected_transport_country = profile_data["transport_country_combo"]
        
            # Continue with remaining profile data
            profile_data["manufacturing_country_combo"] = str(ws['B10'].value or "")
            profile_data["material_combo"] = str(ws['B11'].value or "")
            profile_data["weight_entry"] = str(ws['B12'].value or "0")
            profile_data["quantity_entry_weight"] = str(ws['B13'].value or "1")
            profile_data["weight_after_entry"] = str(ws['B14'].value or "0")
            profile_data["weight_after_quantity_entry"] = str(ws['B15'].value or "1")

            # Packaging & Recycling (E9-E10)
            packaging_value = str(ws['E9'].value or "")
            recycling_value = str(ws['E10'].value or "")
            profile_data["packaging_dropdown"] = packaging_value
            profile_data["metal_recycling_dropdown"] = recycling_value
            print(f"[DEBUG] Packaging: {packaging_value}, Recycling: {recycling_value}")

            # Transport Rows (starts at row 21)
            transport_data_start = 21  # Transport data starts after headers
            transport_rows = []

            # First count the transport rows with data
            transport_count = 0
            for i in range(100):  # Template has 100 transport rows
                row_num = transport_data_start + i
                # Only count rows that have valid transport data (transport type AND at least one other field)
                if (ws.cell(row=row_num, column=1).value and  # Column A (transport type)
                    any(ws.cell(row=row_num, column=j).value for j in range(2, 7))):  
                    transport_count += 1

            print(f"[DEBUG] Found {transport_count} transport rows")
            profile_data["quantity_entry"] = str(transport_count)

            # Extract the actual transport data
            for i in range(transport_count):
                row_num = transport_data_start + i
                transport_type = ws.cell(row=row_num, column=1).value  # Column A
                variation = ws.cell(row=row_num, column=2).value      # Column B
                fuel = ws.cell(row=row_num, column=3).value          # Column C
                origin = ws.cell(row=row_num, column=4).value        # Column D
                destination = ws.cell(row=row_num, column=5).value    # Column E
                mode = ws.cell(row=row_num, column=6).value or "None" # Column F

                if transport_type:
                    transport_row = {
                        "transport": str(transport_type),
                        "variation": str(variation) if variation is not None else "",
                        "fuel": str(fuel) if fuel is not None else "",
                        "origin": str(origin) if origin is not None else "",
                        "destination": str(destination) if destination is not None else "",
                        "mode": str(mode),
                        "value_label": "" 
                    }
                    transport_rows.append(transport_row)
                    print(f"[DEBUG] Transport {i+1}: {transport_type}, Variation: {variation}, Fuel: {fuel}")

            profile_data["transport_rows"] = transport_rows
    
            # Machine Process Rows (starts at column I)
            machine_data_start = 21  # Machine data starts after headers
            parts_data = {}

            # Parse all machine rows, handling part headers and processes
            for i in range(100):  # Scan 100 more rows to catch all parts and processes
                row_num = machine_data_start + i
                part_cell = ws.cell(row=row_num, column=8).value    # Column H
                machine_type = ws.cell(row=row_num, column=9).value  # Column I
                machining_time = ws.cell(row=row_num, column=10).value # Column J
    
                # Skip rows that are empty
                if not part_cell:
                    continue
        
                # Skip rows with header text
                if isinstance(part_cell, str) and "Part" not in part_cell:
                    continue

                # Extract part number
                part_name = str(part_cell)
                try:
                    if "Part" in part_name:
                        part_num = int(''.join(filter(str.isdigit, part_name)))
                    else:
                        part_num = int(part_name) if part_name.isdigit() else len(parts_data) + 1
                except ValueError:
                    part_num = len(parts_data) + 1

                if part_num not in parts_data:
                    parts_data[part_num] = []

                # Add the process to this part
                parts_data[part_num].append({
                    "machine_type": str(machine_type) if machine_type else "",
                    "energy_use": "0.0",  
                    "machining_time": str(machining_time) if machining_time is not None else "0",
                    "carbon_emission": "0.0"  
                })
                print(f"[DEBUG] Machine Part {part_num}, Process: {machine_type}, Time: {machining_time}")

            # Convert to the format expected by the app
            machine_process_rows = []
            for part_num in sorted(parts_data.keys()):
                rows = parts_data[part_num]
                for row_idx, row in enumerate(rows, 1):
                    machine_process_rows.append({
                        "part": f"Part {part_num}",
                        "row": row_idx,
                        "machine_type": row["machine_type"],
                        "energy_use": row["energy_use"],
                        "machining_time": row["machining_time"],
                        "carbon_emission": row["carbon_emission"]
                    })

            profile_data["machine_process_rows"] = machine_process_rows
            profile_data["machine_process_quantity_entry"] = str(len(parts_data)) if parts_data else "1"
            print(f"[DEBUG] Found {len(parts_data)} machine parts with {len(machine_process_rows)} total processes")

            global machine_parts_data, current_part_index
            original_index = current_part_index

            try:
                machine_parts_data.clear()
                parts = {}
                for row in profile_data.get('machine_process_rows', []):
                    part_info = row.get('part', '')
                    if "Part" in str(part_info):
                        try:
                            part_num = int(str(part_info).split()[1])
                        except Exception:
                            continue
                        if part_num not in parts:
                            parts[part_num] = []
                        parts[part_num].append({
                            "machine_type": row.get("machine_type", ""),
                            "energy_use": row.get("energy_use", "0.0"),
                            "machining_time": row.get("machining_time", ""),
                            "carbon_emission": "0.0"
                        })

                for part_num in sorted(parts.keys()):
                    machine_parts_data.append({
                        "machine_row_qty": len(parts[part_num]),
                        "rows": parts[part_num],
                        "carbon_emission": "0.0"  
                    })
            except Exception as e:
                print(f"[DEBUG] Error rebuilding machine_parts_data: {e}")
                if not machine_parts_data:
                    machine_parts_data.append({
                        "machine_row_qty": 1,
                        "rows": [{"machine_type": "", "energy_use": "0.0", "machining_time": "", "carbon_emission": "0.0"}],
                        "carbon_emission": "0.0"
                    })

            # Make sure manufacturing country is set before calculations
            selected_country = profile_data.get("manufacturing_country_combo", "")
            if not selected_country and len(country_list) > 0:
                selected_country = country_list[0]
                profile_data["manufacturing_country_combo"] = selected_country

            # Get the country index for electricity emission factor
            country_idx = -1
            if selected_country in country_list:
                country_idx = country_list.index(selected_country)
            else:
                print(f"[DEBUG] Warning: Manufacturing country '{selected_country}' not found in country list")
                if country_list:
                    country_idx = 0  # Fallback to first country

            # Calculate emissions for each machine part
            for part_idx, part_data in enumerate(machine_parts_data):
                total_emission = 0.0
                if "rows" in part_data:
                    for row in part_data["rows"]:
                        try:
                            machine_type = row.get("machine_type", "")
                            machining_time = float(row.get("machining_time", "0"))
            
                            # Look up the energy use for this machine type
                            energy_use = 0.0
                            if machine_type in machine_value_list:
                                idx = machine_value_list.index(machine_type)
                                energy_use = float(specific_machine_energy_use_list[idx])
            
                            # Get manufacturing country electricity emission
                            electricity_emission = 0.0
                            if country_idx != -1:
                                electricity_emission = float(electricity_list[country_idx])
                                print(f"[DEBUG] Using electricity emission: {electricity_emission} for '{selected_country}'")
                            else:
                                print(f"[DEBUG] No valid electricity emission found for '{selected_country}'")
            
                            # Get weight after and quantity
                            weight_after = float(profile_data.get("weight_after_entry", "0"))
                            weight_after_quantity = float(profile_data.get("weight_after_quantity_entry", "1"))
            
                            # Calculate carbon emission
                            carbon_emission = energy_use * weight_after * weight_after_quantity * (machining_time / 60) * electricity_emission
                            row["energy_use"] = f"{energy_use:.2f}"
                            row["carbon_emission"] = f"{carbon_emission:.2f}"
                            total_emission += carbon_emission
                        except Exception as e:
                            print(f"[DEBUG] Error calculating machine emission: {e}")
            
                    # Set the total carbon emission for this part
                    part_data["carbon_emission"] = f"{total_emission:.2f}"

            # Add this to check if the UI widgets are correctly updated
            print("[DEBUG] Testing UI update functionality:")
            if len(machine_parts_data) > 0 and current_part_index < len(machine_parts_data):
                part_data = machine_parts_data[current_part_index]
                if "rows" in part_data:
                    for i, row in enumerate(part_data["rows"]):
                        print(f"Row {i+1} data before UI update:")
                        print(f"  Machine type: {row.get('machine_type', '')}")
                        print(f"  Energy use: {row.get('energy_use', '0.0')}")
                        print(f"  Machining time: {row.get('machining_time', '')}")
                        print(f"  Carbon emission: {row.get('carbon_emission', '0.0')}")    

            try:
                current_part_index = 0 
                update_machine_process_page() 
            
                # Force update of the UI with calculated values
                if len(machine_parts_data) > 0:
                    # Update profile_data with calculated emissions for machine processes
                    profile_data['machine_process_rows'] = []
                    for part_idx, part_data in enumerate(machine_parts_data, 1):
                        if "rows" in part_data:
                            for row_idx, row in enumerate(part_data["rows"], 1):
                                profile_data['machine_process_rows'].append({
                                    "part": f"Part {part_idx}",
                                    "row": row_idx,
                                    "machine_type": row.get("machine_type", ""),
                                    "energy_use": row.get("energy_use", "0.0"),
                                    "machining_time": row.get("machining_time", ""),
                                    "carbon_emission": row.get("carbon_emission", "0.0")
                                })
                
                    # Update widgets if they exist
                    if 'machine_process_row_widgets' in globals() and machine_process_row_widgets:
                        # Update each part in UI
                        for part_idx in range(len(machine_parts_data)):
                            current_part_index = part_idx
                            update_machine_process_page()
                    
                            part_data = machine_parts_data[part_idx]
                            if "rows" in part_data:
                                for row_idx, row in enumerate(part_data["rows"]):
                                    if row_idx < len(machine_process_row_widgets):
                                        row_widgets = machine_process_row_widgets[row_idx]
                                        try:
                                            if row_widgets["machine_type_combo"].winfo_exists():
                                                row_widgets["machine_type_combo"]["values"] = [""] + machine_value_list
                                                imported_value = row.get("machine_type", "")
                                                if not imported_value or imported_value.strip() == "":
                                                    row_widgets["machine_type_combo"].set("")
                                                    row_widgets["machine_type_combo"].update_idletasks()
                                                else:
                                                    row_widgets["machine_type_combo"].set(imported_value)
                                            if row_widgets["energy_use_value_label"].winfo_exists():
                                                row_widgets["energy_use_value_label"].config(text=row.get("energy_use", "0.0"))
                                            if row_widgets["machining_time_entry"].winfo_exists():
                                                row_widgets["machining_time_entry"].delete(0, tk.END)
                                                row_widgets["machining_time_entry"].insert(0, row.get("machining_time", ""))
                                            if row_widgets["carbon_emission_value_label"].winfo_exists():
                                                row_widgets["carbon_emission_value_label"].config(text=row.get("carbon_emission", "0.0"))
                                        except tk.TclError:
                                            print(f"[DEBUG] Widget access error for row {row_idx}")
                                        
                        current_part_index = 0
                        update_machine_process_page()
            except Exception as e:
                print(f"[DEBUG] Error updating machine process UI: {e}")
        
            try:
                if 'packaging_dropdown' in globals() and packaging_dropdown and hasattr(packaging_dropdown, 'winfo_exists') and packaging_dropdown.winfo_exists():
                    packaging_dropdown.set(packaging_value)
                    packaging_dropdown.event_generate("<<ComboboxSelected>>")
                    print(f"[DEBUG] Set packaging dropdown to: {packaging_value}")
            except (tk.TclError, NameError, AttributeError) as e:
                print(f"[DEBUG] Error setting packaging dropdown: {e}")
            
            try:
                if 'metal_recycling_dropdown' in globals() and metal_recycling_dropdown and hasattr(metal_recycling_dropdown, 'winfo_exists') and metal_recycling_dropdown.winfo_exists():
                    metal_recycling_dropdown.set(recycling_value)
                    metal_recycling_dropdown.event_generate("<<ComboboxSelected>>")
                    print(f"[DEBUG] Set metal recycling dropdown to: {recycling_value}")
            except (tk.TclError, NameError, AttributeError) as e:
                print(f"[DEBUG] Error setting recycling dropdown: {e}")
            
            # Calculate packaging emission
            try:
                if 'on_packaging_type_change' in globals():
                    on_packaging_type_change()
                    print(f"[DEBUG] Calculated packaging emission")
            except Exception as e:
                print(f"[DEBUG] Error calculating packaging emission: {e}")

            # Calculate recycling emission
            try:
                if 'update_recycled_emission_value' in globals():
                    update_recycled_emission_value()
                    print(f"[DEBUG] Calculated recycling emission")
            except Exception as e:
                print(f"[DEBUG] Error calculating recycling emission: {e}")

            try:
                # First update all the widgets with imported values
                if 'transport_country_combo' in globals() and transport_country_combo and transport_country_combo.winfo_exists():
                    transport_country_combo.set(selected_transport_country)
                    print(f"[DEBUG] Set transport_country_combo to: {selected_transport_country}")
                
                if 'material_combo' in globals() and material_combo and material_combo.winfo_exists():
                    material_combo.set(profile_data.get("material_combo", ""))
                
                if 'weight_entry' in globals() and weight_entry and weight_entry.winfo_exists():
                    weight_entry.delete(0, tk.END)
                    weight_entry.insert(0, profile_data.get("weight_entry", "0"))
                
                if 'quantity_entry_weight' in globals() and quantity_entry_weight and quantity_entry_weight.winfo_exists():
                    quantity_entry_weight.delete(0, tk.END)
                    quantity_entry_weight.insert(0, profile_data.get("quantity_entry_weight", "1"))
                
                if 'weight_after_entry' in globals() and weight_after_entry and weight_after_entry.winfo_exists():
                    weight_after_entry.delete(0, tk.END)
                    weight_after_entry.insert(0, profile_data.get("weight_after_entry", "0"))
                
                if 'weight_after_quantity_entry' in globals() and weight_after_quantity_entry and weight_after_quantity_entry.winfo_exists():
                    weight_after_quantity_entry.delete(0, tk.END)
                    weight_after_quantity_entry.insert(0, profile_data.get("weight_after_quantity_entry", "1"))
                
                if 'manufacturing_country_combo' in globals() and manufacturing_country_combo and manufacturing_country_combo.winfo_exists():
                    manufacturing_country_combo.set(profile_data.get("manufacturing_country_combo", ""))

                root.update_idletasks()
            except Exception as e:
                print(f"[DEBUG] Error setting widget values: {e}")
        
            try:
                if selected_transport_country in country_list:
                    print(f"[DEBUG] Applying emergency fix for Excel import: {selected_transport_country}")
        
                    # First set the transport_country_combo to ensure consistency
                    if 'transport_country_combo' in globals() and transport_country_combo and transport_country_combo.winfo_exists():
                        transport_country_combo.set(selected_transport_country)
                        print(f"[DEBUG] Set transport_country_combo to: {selected_transport_country}")
        
                    force_update_emission_labels_after_import(selected_transport_country)
        
                    # Then trigger the regular update functions for everything else
                    calculate_material_value()
                    recalculate_all()
                    update_sum_of_emissions()
                    update_pie_chart()
        
                    print(f"[DEBUG] Successfully updated Carbon Emission Overview for {selected_transport_country}")
                else:
                    print(f"[DEBUG] Warning: Imported transportation country '{selected_transport_country}' not found in country list")
            except Exception as e:
                print(f"[DEBUG] Error updating carbon emission overview: {e}")

            # Add the profile to the list
            if profile_data["product_name"]:
                if hasattr(cover_page, "profiles") and cover_page.profiles is not None:
                    if profile_data["product_name"] in cover_page.profiles:
                        if not messagebox.askyesno("Profile Exists", 
                                            f"A profile named '{profile_data['product_name']}' already exists.\nDo you want to overwrite it?"):
                            base_name = profile_data["product_name"]
                            counter = 1
                            while f"{base_name}_{counter}" in cover_page.profiles:
                                counter += 1
                            profile_data["product_name"] = f"{base_name}_{counter}"
    
                    cover_page.profiles[profile_data["product_name"]] = profile_data
                    save_profiles(cover_page.profiles)
    
                    # Update the profile dropdown
                    cover_page.profile_combo["values"] = sorted(cover_page.profiles.keys())
                    cover_page.selected_profile.set(profile_data["product_name"])
    
                    # Immediately load the imported profile
                    cover_page.profile_data = profile_data
                    cover_page.show_main_page_callback(profile_data)

        finally:
            root.config(cursor="") 
            is_importing = False

    def create_excel_template():
            """Creates an Excel template with transportation and machine process sections side by side"""
            from openpyxl import Workbook
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
    
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Save Excel Template"
            )
    
            if not file_path:
                return
    
            wb = Workbook()
            ws = wb.active
            ws.title = "SPHERE Data"
    
            # Get transport variation and fuel mappings
            transport_variations, transport_fuels, sea_tanker_variation_fuels, cargo_ship_variation_fuels = update_dependent_dropdowns_for_excel_template()
    
            # Create reference sheets
            lists_sheet = wb.create_sheet("Lists")
            transport_data_sheet = wb.create_sheet("TransportData")
    
            # --- Create all reference lists ---
            # Countries
            lists_sheet['A1'] = "Countries"
            for i, country in enumerate(country_list):
                lists_sheet[f'A{i+2}'] = country
            country_list_range = f'Lists!$A$2:$A${len(country_list)+1}'
    
            # Materials
            lists_sheet['B1'] = "Materials"
            for i, material in enumerate(materials):
                lists_sheet[f'B{i+2}'] = material
            material_list_range = f'Lists!$B$2:$B${len(materials)+1}'
    
            # Transport types
            lists_sheet['C1'] = "Transport Types"
            for i, transport in enumerate(transport_list):
                lists_sheet[f'C{i+2}'] = transport
            transport_list_range = f'Lists!$C$2:$C${len(transport_list)+1}'
    
            # Machine types
            lists_sheet['D1'] = "Machine Types"
            for i, machine in enumerate(machine_value_list):
                lists_sheet[f'D{i+2}'] = machine
            machine_list_range = f'Lists!$D$2:$D${len(machine_value_list)+1}'
    
            # Mode options
            lists_sheet['E1'] = "Modes"
            mode_options = ["None", "driving", "plane", "ship"]
            for i, mode in enumerate(mode_options):
                lists_sheet[f'E{i+2}'] = mode
            mode_list_range = f'Lists!$E$2:$E${len(mode_options)+1}'
    
            # Packaging types
            lists_sheet['F1'] = "Packaging"
            for i, packaging in enumerate(packaging_types_list):
                lists_sheet[f'F{i+2}'] = packaging
            packaging_list_range = f'Lists!$F$2:$F${len(packaging_types_list)+1}'
    
            # Metal recycling types
            lists_sheet['G1'] = "Recycling"
            for i, recycling in enumerate(metal_recyling_types_list):
                lists_sheet[f'G{i+2}'] = recycling
            recycling_list_range = f'Lists!$G$2:$G${len(metal_recyling_types_list)+1}'
    
            # --- Create transport data reference sheet ---
            transport_data_sheet['A1'] = "Transport Type"
            transport_data_sheet['B1'] = "Variations"
            transport_data_sheet['C1'] = "Fuels"
    
            # Write transport variations and fuels
            row_num = 1
            for transport_type, variations in transport_variations.items():
                # Write transport type and variations
                variation_start_row = row_num + 1
                for i, variation in enumerate(variations):
                    transport_data_sheet[f'A{variation_start_row+i}'] = transport_type
                    transport_data_sheet[f'B{variation_start_row+i}'] = variation
    
                # Define named range for variations
                variation_name = transport_type.replace(' ', '_').replace('(', '').replace(')', '')
                wb.create_named_range(
                    f"{variation_name}_Variations",
                    transport_data_sheet,
                    f'$B${variation_start_row}:$B${variation_start_row+len(variations)-1}'
                )
    
                # Create a named range for this transport's fuels
                fuels = transport_fuels.get(transport_type, [])
                fuel_start_row = row_num + 1
                for i, fuel in enumerate(fuels):
                    transport_data_sheet[f'C{fuel_start_row+i}'] = fuel
    
                if fuels:
                    wb.create_named_range(
                        f"{variation_name}_Fuels",
                        transport_data_sheet,
                        f'$C${fuel_start_row}:$C${fuel_start_row+len(fuels)-1}'
                    )
    
                # Update the row number for the next transport type
                row_num += max(len(variations), len(fuels))
    
            # Create special tables for Sea Tanker and Cargo Ship variations
            row_num += 2
            transport_data_sheet[f'A{row_num}'] = "Sea Tanker Variation"
            transport_data_sheet[f'B{row_num}'] = "Fuels"
            row_num += 1
    
            for variation, fuels in sea_tanker_variation_fuels.items():
                variation_row = row_num
                transport_data_sheet[f'A{variation_row}'] = variation
    
                for i, fuel in enumerate(fuels):
                    transport_data_sheet[f'B{variation_row+i}'] = fuel
    
                variation_name = variation.replace(' ', '_')
                wb.create_named_range(
                    f"SeaTanker_{variation_name}_Fuels",
                    transport_data_sheet,
                    f'$B${variation_row}:$B${variation_row+len(fuels)-1}'
                )
    
                row_num += max(1, len(fuels))
    
            row_num += 2
            transport_data_sheet[f'A{row_num}'] = "Cargo Ship Variation"
            transport_data_sheet[f'B{row_num}'] = "Fuels"
            row_num += 1
    
            for variation, fuels in cargo_ship_variation_fuels.items():
                variation_row = row_num
                transport_data_sheet[f'A{variation_row}'] = variation
    
                for i, fuel in enumerate(fuels):
                    transport_data_sheet[f'B{variation_row+i}'] = fuel
    
                variation_name = variation.replace(' ', '_')
                wb.create_named_range(
                    f"CargoShip_{variation_name}_Fuels",
                    transport_data_sheet,
                    f'$B${variation_row}:$B${variation_row+len(fuels)-1}'
                )
    
                row_num += max(1, len(fuels))
    
            # --- Set up styling ---
            header_font = Font(bold=True, size=12)
            subheader_font = Font(bold=True, size=11)
            header_fill = PatternFill("solid", fgColor="A8D5BA")
            instruction_fill = PatternFill("solid", fgColor="FFFFCC")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
    
            # --- Main sheet setup ---
            ws = wb.active
    
            # ---- Instructions ----
            ws.merge_cells('A1:K1')
            ws['A1'] = "SPHERE Carbon Emission Calculator - Data Entry Template"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws['A1'].fill = header_fill
    
            ws.merge_cells('A2:K2')
            ws['A2'] = "Instructions: Enter the number of rows needed in each quantity field, then fill in the data for each row."
            ws['A2'].font = Font(italic=True)
            ws['A2'].alignment = Alignment(horizontal='center')
            ws['A2'].fill = instruction_fill
    
            # ---- Basic Info Section ----
            row = 4
            ws['A'+str(row)] = "Basic Information"
            ws['A'+str(row)].font = header_font
            ws['A'+str(row)].fill = header_fill
            ws.merge_cells(f'A{row}:B{row}')
    
            row += 1
            ws['A'+str(row)] = "Product Name"
            ws['B'+str(row)] = ""
    
            row += 1
            ws['A'+str(row)] = "Description"
            ws['B'+str(row)] = ""
    
            # ---- Material Information and Packaging & Recycling Side-by-Side ----
            row += 2
    
            # Left side: Material Information
            ws['A'+str(row)] = "Material Information"
            ws['A'+str(row)].font = header_font
            ws['A'+str(row)].fill = header_fill
            ws.merge_cells(f'A{row}:B{row}')
    
            # Right side: Packaging & Recycling
            ws['D'+str(row)] = "Packaging & Recycling"
            ws['D'+str(row)].font = header_font
            ws['D'+str(row)].fill = header_fill
            ws.merge_cells(f'D{row}:E{row}')
    
            # Material Information fields
            material_start_row = row + 1
    
            ws['A'+str(material_start_row)] = "Transport Country"
            ws['B'+str(material_start_row)] = ""
            country_dv = DataValidation(type="list", formula1=country_list_range, allow_blank=True, showErrorMessage=True)
            ws.add_data_validation(country_dv)
            country_dv.add(f'B{material_start_row}')
    
            material_start_row += 1
            ws['A'+str(material_start_row)] = "Manufacturing Country"
            ws['B'+str(material_start_row)] = ""
            country_dv.add(f'B{material_start_row}')
    
            material_start_row += 1
            ws['A'+str(material_start_row)] = "Material"
            ws['B'+str(material_start_row)] = ""
            material_dv = DataValidation(type="list", formula1=material_list_range, allow_blank=True, showErrorMessage=True)
            ws.add_data_validation(material_dv)
            material_dv.add(f'B{material_start_row}')
    
            material_start_row += 1
            ws['A'+str(material_start_row)] = "Raw Weight (kg)"
            ws['B'+str(material_start_row)] = ""
    
            material_start_row += 1
            ws['A'+str(material_start_row)] = "Raw Weight Quantity"
            ws['B'+str(material_start_row)] = "1"
    
            material_start_row += 1
            ws['A'+str(material_start_row)] = "Weight After Manufacturing (kg)"
            ws['B'+str(material_start_row)] = ""
    
            material_start_row += 1
            ws['A'+str(material_start_row)] = "Weight After Manufacturing Quantity"
            ws['B'+str(material_start_row)] = "1"

            from openpyxl.worksheet.datavalidation import DataValidation

            # Calculate the correct row numbers for each field
            raw_weight_row = material_start_row - 3
            raw_weight_qty_row = material_start_row - 2
            weight_after_row = material_start_row - 1
            weight_after_qty_row = material_start_row

            # Decimal validation for weights
            decimal_dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True, showErrorMessage=True)
            decimal_dv.error = "Please enter a valid number (no letters or symbols)."
            decimal_dv.errorTitle = "Invalid Input"
            ws.add_data_validation(decimal_dv)
            decimal_dv.add(f'B{raw_weight_row}')
            decimal_dv.add(f'B{weight_after_row}')   
            decimal_dv.add(f'B{raw_weight_qty_row}')
            decimal_dv.add(f'B{weight_after_qty_row}')

            # Packaging & Recycling fields (placed to the right of Material Information)
            packaging_row = row + 1
            ws['D'+str(packaging_row)] = "Packaging Type"
            ws['E'+str(packaging_row)] = ""
            packaging_dv = DataValidation(type="list", formula1=packaging_list_range, allow_blank=True, showErrorMessage=True)
            ws.add_data_validation(packaging_dv)
            packaging_dv.add(f'E{packaging_row}')
    
            packaging_row += 1
            ws['D'+str(packaging_row)] = "Recycled Metal Type"
            ws['E'+str(packaging_row)] = ""
            recycling_dv = DataValidation(type="list", formula1=recycling_list_range, allow_blank=True, showErrorMessage=True)
            ws.add_data_validation(recycling_dv)
            recycling_dv.add(f'E{packaging_row}')
    
            # ---- TRANSPORTATION AND MACHINE SECTIONS ----
            # Reset row to the maximum of Material Information and Packaging & Recycling sections + gap
            row = max(material_start_row, packaging_row) + 2

            note_row = row  # This is the row above "Transportation Options"
            ws.merge_cells(f'A{note_row}:F{note_row}')
            ws[f'A{note_row}'] = (
                "Note: When adding more rows, copy an existing row to maintain the validations."
            )
            ws[f'A{note_row}'].font = Font(color="FF0000", italic=True)
            ws[f'A{note_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            ws.merge_cells(f'H{note_row}:K{note_row}')
            ws[f'H{note_row}'] = "Note: 1 = machine part. Multiple rows of Part 1 = Process for that part. Same logic applies to other parts."
            ws[f'H{note_row}'].font = Font(color="FF0000", italic=True)
            ws[f'H{note_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws[f'H{note_row}'].fill = PatternFill("solid", fgColor="FFFFCC")    
            
            row += 1  # Move to the next row for "Transportation Options"
            # ---- Transportation Section (Left Side) ----
            ws['A'+str(row)] = "Transportation Options"
            ws['A'+str(row)].font = header_font
            ws['A'+str(row)].fill = header_fill
            ws.merge_cells(f'A{row}:F{row}')
    
            # ---- Machine Process Section (Right Side) ----
            # Place the machine process section exactly at the same row as transportation
            right_col = 8
    
            ws.cell(row=row, column=right_col).value = "Machine Processes"
            ws.cell(row=row, column=right_col).font = header_font
            ws.cell(row=row, column=right_col).fill = header_fill
            ws.merge_cells(f'{get_column_letter(right_col)}{row}:{get_column_letter(right_col+3)}{row}')
    
            # Transport quantity input
            transport_qty_row = row + 1
            ws['A'+str(transport_qty_row)] = "Number of Transport Rows"
            ws['B'+str(transport_qty_row)] = "1"
            ws['B'+str(transport_qty_row)].font = Font(bold=True)
            ws['C'+str(transport_qty_row)] = "← Enter quantity here"
            ws['C'+str(transport_qty_row)].font = Font(italic=True, color="FF0000")
    
            # Machine quantity input - same row as transport quantity
            machine_qty_row = transport_qty_row
            ws.cell(row=machine_qty_row, column=right_col).value = "Number of Machine Parts"
            ws.cell(row=machine_qty_row, column=right_col+1).value = "1"
            ws.cell(row=machine_qty_row, column=right_col+1).font = Font(bold=True)
            ws.cell(row=machine_qty_row, column=right_col+2).value = "← Enter quantity here"
            ws.cell(row=machine_qty_row, column=right_col+2).font = Font(italic=True, color="FF0000")
    
            # Transport headers
            transport_header_row = transport_qty_row + 1
            for col, header in enumerate(['Transport Type', 'Variation', 'Fuel', 'Origin', 'Destination', 'Mode']):
                cell = ws.cell(row=transport_header_row, column=col+1)
                cell.value = header
                cell.font = subheader_font
                cell.fill = header_fill
                cell.border = thin_border
    
            # Machine headers - same row as transport headers
            machine_header_row = transport_header_row
            for col, header in enumerate(['Part Number', 'Machine Type', 'Machining Time (min)', 'Notes']):
                cell = ws.cell(row=machine_header_row, column=right_col+col)
                cell.value = header
                cell.font = subheader_font
                cell.fill = header_fill
                cell.border = thin_border
    
            # Transport data rows
            transport_data_start = transport_header_row + 1
            transport_dv = DataValidation(type="list", formula1=transport_list_range, allow_blank=True, showErrorMessage=True)
            mode_dv = DataValidation(type="list", formula1=mode_list_range, allow_blank=True, showErrorMessage=True)
            ws.add_data_validation(transport_dv)
            ws.add_data_validation(mode_dv)
    
            for i in range(5):
                row = transport_data_start + i
                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = thin_border
        
                # Apply transport validations
                transport_dv.add(f'A{row}') 
                mode_dv.add(f'F{row}') 
        
                # Create INDIRECT formulas for variations based on transport type
                variation_formula = f'=INDIRECT(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(A{row}," ","_"),"(",""),")","")'+'&"_Variations")'
                variation_dv = DataValidation(type="list", formula1=variation_formula, allow_blank=True, showErrorMessage=True)
                ws.add_data_validation(variation_dv)
                variation_dv.add(f'B{row}')
        
                # Create complex INDIRECT formulas for fuels based on transport type and variation
                fuel_formula = (
                    f'=IF(A{row}="Sea Tanker",INDIRECT("SeaTanker_"&SUBSTITUTE(B{row}," ","_")&"_Fuels"),'
                    f'IF(A{row}="Cargo Ship",INDIRECT("CargoShip_"&SUBSTITUTE(B{row}," ","_")&"_Fuels"),'
                    f'INDIRECT(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(A{row}," ","_"),"(",""),")","")&"_Fuels")))'
                )
                fuel_dv = DataValidation(type="list", formula1=fuel_formula, allow_blank=True, showErrorMessage=True)
                ws.add_data_validation(fuel_dv)
                fuel_dv.add(f'C{row}')
    
            # Machine data rows - aligned with transport rows
            machine_data_start = transport_data_start
            machine_dv = DataValidation(type="list", formula1=machine_list_range, allow_blank=True, showErrorMessage=True)
            ws.add_data_validation(machine_dv)
    
            for i in range(5):
                row = machine_data_start + i
                # Add borders
                for col in range(right_col, right_col+4):
                    ws.cell(row=row, column=col).border = thin_border
        
                # Set initial values for machine row
                ws.cell(row=row, column=right_col).value = f"Part {i//2 + 1}"  # Group rows by part
                machine_dv.add(f'{get_column_letter(right_col+1)}{row}')  # Apply machine type validation
    
            # Adjust column widths
            column_widths = {
                1: 18,  # A
                2: 18,  # B
                3: 15,  # C
                4: 18,  # D
                5: 18,  # E
                6: 12,  # F
                7: 2,   # G (separator)
                8: 15,  # H
                9: 22,  # I
                10: 18, # J
                11: 15  # K
            }
    
            for col, width in column_widths.items():
                ws.column_dimensions[get_column_letter(col)].width = width
  
            # Hide the reference sheets
            lists_sheet.sheet_state = 'hidden'
            transport_data_sheet.sheet_state = 'hidden'
    
            try:
                from openpyxl.styles import Font, Alignment, PatternFill

                ws2 = wb.create_sheet(title="Instructions")

                # Define styles
                title_font = Font(size=14, bold=True, color="FFFFFF")
                title_fill = PatternFill("solid", fgColor="4F81BD")  # Blue
                section_font = Font(size=12, bold=True, color="FFFFFF")
                section_fill = PatternFill("solid", fgColor="70AD47")  # Green
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill("solid", fgColor="A9D08E")  # Light green
                note_font = Font(bold=True, color="FF0000")
                tip_fill = PatternFill("solid", fgColor="FFF2CC")  # Light yellow
                wrap_alignment = Alignment(wrap_text=True, vertical="top")
                center_alignment = Alignment(horizontal='center', vertical='center')

                # Title
                ws2.merge_cells('A1:H1')
                ws2['A1'] = "SPHERE Carbon Emission Calculator - Data Entry Instructions"
                ws2['A1'].font = title_font
                ws2['A1'].fill = title_fill
                ws2['A1'].alignment = center_alignment

                row = 3
                ws2.merge_cells(f'A{row}:H{row}')
                ws2[f'A{row}'] = "1. Basic Information"
                ws2[f'A{row}'].font = section_font
                ws2[f'A{row}'].fill = section_fill
                ws2[f'A{row}'].alignment = center_alignment
                row += 1

                # Basic Info Table Header
                ws2.append(["Field", "Instruction", "Example Label", "Example Value"])
                for col in "ABCD":
                    ws2[f"{col}{row}"].font = header_font
                    ws2[f"{col}{row}"].fill = header_fill
                    ws2[f"{col}{row}"].alignment = center_alignment
                row += 1

                ws2.append(["Product Name", "Enter the name of your product.", "Example:", "Aluminum Gear Housing"])
                ws2.append(["Description", "Provide a brief description or notes.", "Example:", "Prototype for lightweight automotive gearbox."])
                row = ws2.max_row + 2

                ws2.merge_cells(f'A{row}:H{row}')
                ws2[f'A{row}'] = "2. Material Information"
                ws2[f'A{row}'].font = section_font
                ws2[f'A{row}'].fill = section_fill
                ws2[f'A{row}'].alignment = center_alignment
                row += 1

                ws2.append(["Field", "Instruction", "Example Label", "Example Value"])
                for col in "ABCD":
                    ws2[f"{col}{row}"].font = header_font
                    ws2[f"{col}{row}"].fill = header_fill
                    ws2[f"{col}{row}"].alignment = center_alignment
                row += 1

                ws2.append(["Transport Country", "Select the country from which the product will be transported.", "Example:", "Singapore"])
                ws2.append(["Manufacturing Country", "Select the country where the product is manufactured.", "Example:", "China"])
                ws2.append(["Material", "Select the main material used.", "Example:", "Aluminium"])
                ws2.append(["Raw Weight (kg)", "Enter the total raw material weight in kilograms.", "Example:", "2.5"])
                ws2.append(["Raw Weight Quantity", "Enter the number of raw material units.", "Example:", "100"])
                ws2.append(["Weight After Manufacturing (kg)", "Enter the weight after manufacturing.", "Example:", "2.0"])
                ws2.append(["Weight After Manufacturing Quantity", "Enter the number of finished products.", "Example:", "100"])
                row = ws2.max_row + 2

                ws2.merge_cells(f'A{row}:H{row}')
                ws2[f'A{row}'] = "3. Packaging & Recycling"
                ws2[f'A{row}'].font = section_font
                ws2[f'A{row}'].fill = section_fill
                ws2[f'A{row}'].alignment = center_alignment
                row += 1

                ws2.append(["Field", "Instruction", "Example Label", "Example Value"])
                for col in "ABCD":
                    ws2[f"{col}{row}"].font = header_font
                    ws2[f"{col}{row}"].fill = header_fill
                    ws2[f"{col}{row}"].alignment = center_alignment
                row += 1

                ws2.append(["Packaging Type", "Select the type of packaging used.", "Example:", "Cardboard Box"])
                ws2.append(["Recycled Metal Type", "Select the type of recycled metal, if any.", "Example:", "Recycled Aluminium"])
                row = ws2.max_row + 2

                ws2.merge_cells(f'A{row}:H{row}')
                ws2[f'A{row}'] = "4. Transport Rows"
                ws2[f'A{row}'].font = section_font
                ws2[f'A{row}'].fill = section_fill
                ws2[f'A{row}'].alignment = center_alignment
                row += 1

                ws2.merge_cells(f'A{row}:H{row}')
                ws2[f'A{row}'] = "6. Example Data Entry (Summary Table)"
                ws2[f'A{row}'].font = section_font
                ws2[f'A{row}'].fill = section_fill
                ws2[f'A{row}'].alignment = center_alignment
                row += 1

                ws2.append(["Product Name", "Gear Housing"])
                ws2.append(["Description", "Prototype for gearbox"])
                ws2.append(["Transport Country", "China"])
                ws2.append(["Manufacturing Country", "Singapore"])
                ws2.append(["Material", "Aluminium"])
                ws2.append(["Raw Weight (kg)", 2.5])
                ws2.append(["Raw Weight Quantity", 100])
                ws2.append(["Weight After Manufacturing (kg)", 2.0])
                ws2.append(["Weight After Manufacturing Quantity", 100])
                ws2.append(["Packaging Type", "Cardboard Box"])
                ws2.append(["Recycled Metal Type", "Recycled Aluminium"])
                row = ws2.max_row + 2

                ws2.merge_cells(f'A{row}:H{row}')
                ws2[f'A{row}'] = "Transport Example"
                ws2[f'A{row}'].font = section_font
                ws2[f'A{row}'].fill = section_fill
                ws2[f'A{row}'].alignment = center_alignment
                row += 1

                ws2.append(["#", "Transport", "Variation", "Fuel", "Origin", "Destination", "Mode"])
                for col in "ABCDEFG":
                    ws2[f"{col}{row}"].font = header_font
                    ws2[f"{col}{row}"].fill = header_fill
                    ws2[f"{col}{row}"].alignment = center_alignment
                row += 1

                ws2.append([1, "Van", "Class Two", "Diesel", "211805", "211113", "driving"])
                ws2.append([2, "Freight Flights", "International", "International", "211113", "819663", "plane"])
                ws2.append([3, "Van", "Class Two", "Diesel", "819663", "486154", "driving"])  
                row = ws2.max_row + 2

                ws2.merge_cells(f'A{row}:H{row}')
                ws2[f'A{row}'] = "Machine Process Example"
                ws2[f'A{row}'].font = section_font
                ws2[f'A{row}'].fill = section_fill
                ws2[f'A{row}'].alignment = center_alignment
                row += 1

                ws2.append(["Part", "Machine Type", "Machining Time (min)"])
                for col in "ABC":
                    ws2[f"{col}{row}"].font = header_font
                    ws2[f"{col}{row}"].fill = header_fill
                    ws2[f"{col}{row}"].alignment = center_alignment
                row += 1

                ws2.append(["Process 1", "CNC (YCM4)", 15])
                ws2.append(["Process 1", "Surface Grinder (S.G)", 10])
                ws2.append(["Process 2", "CNC (RB8)", 20])
                row += 5

                ws2['A{}'.format(row)] = "Note:"
                ws2['A{}'.format(row)].font = note_font
                ws2['A{}'.format(row)].fill = tip_fill
                ws2['B{}'.format(row)] = "For multiple machine operations per part, use the same part number for related rows."
                ws2['B{}'.format(row)].fill = tip_fill
                row += 1

                ws2['A{}'.format(row)] = "Tip:"
                ws2['A{}'.format(row)].font = note_font
                ws2['A{}'.format(row)].fill = tip_fill
                ws2['B{}'.format(row)] = "If you need more rows for transport or machine processes, copy an existing row to maintain data validation."
                ws2['B{}'.format(row)].fill = tip_fill
                row += 1

                ws2['A{}'.format(row)] = "Tip:"
                ws2['A{}'.format(row)].font = note_font
                ws2['A{}'.format(row)].fill = tip_fill
                ws2['B{}'.format(row)] = "Hover over info icons (ℹ️) in the app for field-specific guidance."
                ws2['B{}'.format(row)].fill = tip_fill
                row += 1

                ws2['A{}'.format(row)] = "Tip:"
                ws2['A{}'.format(row)].font = note_font
                ws2['A{}'.format(row)].fill = tip_fill
                ws2['B{}'.format(row)] = "Use the Undo button to revert to your last saved state."
                ws2['B{}'.format(row)].fill = tip_fill
                row = ws2.max_row + 2

                for col in "ABCDEFGH":
                    ws2.column_dimensions[col].width = 20

                wb.save(file_path)
                messagebox.showinfo("Template Created", 
                    "Excel template saved successfully.\n\n"
                    "The template shows Material Information and Packaging & Recycling side by side, "
                    "and Transportation and Machine Process sections side by side.\n\n"
                    "1. Set the desired number of rows in each quantity field\n"
                    "2. Fill in data for each row (copy an existing row to add more)\n"
                    "3. For multiple machine operations on the same part, use the same part number")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save template: {str(e)}")

    def import_profile_from_file():
        """Import a profile from a .sphere file sent by someone without access to the app"""
        file_path = filedialog.askopenfilename(
            filetypes=[("SPHERE Files", "*.sphere"), ("JSON Files", "*.json"), ("All Files", "*.*")],
            title="Import External Profile Data"
        )
    
        if not file_path:
            return  # User cancelled
    
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                wizard_data = json.load(f)
        
            # Convert to the format expected by the app
            new_profile_data = {}
        
            # Basic info
            new_profile_data["product_name"] = wizard_data.get("product_name", "")
            new_profile_data["description"] = wizard_data.get("description", "")
        
            # Material and weight
            new_profile_data["transport_country_combo"] = wizard_data.get("transport_country", "")
            new_profile_data["manufacturing_country_combo"] = wizard_data.get("manufacturing_country", "")
            new_profile_data["material_combo"] = wizard_data.get("material", "")
            new_profile_data["weight_entry"] = wizard_data.get("weight", "")
            new_profile_data["quantity_entry_weight"] = wizard_data.get("weight_quantity", "1")
            new_profile_data["weight_after_entry"] = wizard_data.get("weight_after", "")
            new_profile_data["weight_after_quantity_entry"] = wizard_data.get("weight_after_quantity", "1")
        
            # Transport options
            new_profile_data["quantity_entry"] = wizard_data.get("transport_quantity", "1")
            new_profile_data["transport_rows"] = []
            for row_data in wizard_data.get("transport_rows", []):
                new_profile_data["transport_rows"].append({
                    "transport": row_data.get("transport", ""),
                    "variation": row_data.get("variation", ""),
                    "fuel": row_data.get("fuel", ""),
                    "origin": row_data.get("origin", ""),
                    "destination": row_data.get("destination", ""),
                    "mode": row_data.get("mode", "None"),
                    "value_label": ""  # This will be calculated later
                })
        
            # Machine process
            parts = wizard_data.get("parts", [])
            new_profile_data["machine_process_quantity_entry"] = wizard_data.get("parts_quantity", str(len(parts)))
            new_profile_data["machine_process_rows"] = []
        
            for i, part in enumerate(parts):
                for j, process in enumerate(part.get("processes", [])):
                    new_profile_data["machine_process_rows"].append({
                        "part": f"Part {i+1}",
                        "row": j+1,
                        "machine_type": process.get("machine_type", ""),
                        "energy_use": process.get("energy_use", "0.0"),
                        "machining_time": process.get("machining_time", ""),
                        "carbon_emission": ""  # This will be calculated later
                    })
        
            # Packaging and recycling
            new_profile_data["packaging_dropdown"] = wizard_data.get("packaging", "")
            new_profile_data["metal_recycling_dropdown"] = wizard_data.get("recycling", "")
        
            # Ask for a profile name if it's empty
            if not new_profile_data["product_name"]:
                new_name = simpledialog.askstring("Profile Name", "Enter a name for this profile:")
                if not new_name:
                    return  # User cancelled
                new_profile_data["product_name"] = new_name
        
            # Add the profile to the list
            if hasattr(cover_page, "profiles"):
                cover_page.profiles[new_profile_data["product_name"]] = new_profile_data
                save_profiles(cover_page.profiles)
            
                # Reload the profiles dropdown
                cover_page.profile_combo["values"] = sorted(cover_page.profiles.keys())
                cover_page.selected_profile.set(new_profile_data["product_name"])
            
                # Show success message
                messagebox.showinfo("Import Successful", 
                    f"Profile '{new_profile_data['product_name']}' has been imported and is ready to use.")
        
        except Exception as e:
            messagebox.showerror("Import Error", f"Error importing profile: {str(e)}")

    def confirm_save_before_back():
        result = messagebox.askyesnocancel(
            "Save Changes",
            "Do you want to save changes before going back?\n\nYes: Save and go back\nNo: Discard changes\nCancel: Stay on this page"
        )
        return result  # True (Yes), False (No), or None (Cancel)

    def on_back_button():
        global is_navigating_back, restoring_profile
    
        store_full_state(widgets)
    
        response = confirm_save_before_back()
    
        if response is None:
            return

        is_navigating_back = True
        restoring_profile = True
    
        if response is True:  # User clicked Yes - save changes
            print("[DEBUG][on_back_button] User chose YES: saving changes before going back.")
            save_profile_inputs(profile_data, widgets)
            if active_profile_name and hasattr(cover_page, "profiles"):
                cover_page.profiles[active_profile_name] = dict(profile_data)
                save_profiles(cover_page.profiles)
            back_to_cover_callback()
        else:  # User clicked No - discard changes
            print("[DEBUG][on_back_button] User chose NO: discarding changes and going back.")
            # Immediately navigate back without saving
            back_to_cover_callback()

    def set_navigating_back_false_and_enable_widgets():
        global is_navigating_back
        is_navigating_back = False
        for widget in widgets.values():
            try:
                widget.config(state="normal")
            except Exception:
                pass

    def set_restoring_profile_false():
        global restoring_profile
        restoring_profile = False

    back_button = tk.Button(
        header_frame,
        text="Back",
        font=("Arial", 12, "bold"),
        bg="#CC0000",
        fg="white",
        command=on_back_button
    )
    back_button.pack(side="right", padx=10, pady=10)

    def get_coordinates(location_name):
        geolocator = Nominatim(user_agent="sea_route_app")
        try:
            location = geolocator.geocode(location_name, timeout=10)
            if location:
                return [location.longitude, location.latitude]
            else:
                print(f"Could not find coordinates for '{location_name}'. Please try again.")
                return None
        except GeocoderTimedOut:
            print("Geocoding service timed out. Please try again.")
            return None

    geocode_cache = {}
    def get_cached_coordinates(location_name):
        if location_name in geocode_cache:
            return geocode_cache[location_name]
        coordinates = get_coordinates(location_name)
        if coordinates:
            geocode_cache[location_name] = coordinates
        return coordinates

    from concurrent.futures import ThreadPoolExecutor
    thread_pool = ThreadPoolExecutor(max_workers=2)
    def calculate_sea_route_async(origin, destination, result_label):
        future = thread_pool.submit(get_cached_sea_route, origin, destination)
        def update_result(future):
            result = future.result()
            def update_label():
                if not result_label.winfo_exists():
                    return  # Widget was destroyed
                if result and len(result) >= 3:
                    figure, distance_nautical, distance_km = result[:3]
                    if distance_km is not None and distance_nautical is not None:
                        result_label.config(text=f"Distance: {distance_km:.2f} km (by ship)\n({distance_nautical:.2f} nautical miles)")
                    else:
                        result_label.config(text="Error calculating sea route distance.")
                else:
                    result_label.config(text="Error calculating sea route distance.")
            try:
                result_label.after(0, update_label)
            except Exception:
                pass  
        future.add_done_callback(lambda f: update_result(f))

    def calculate_distance_dynamic(origin_entry, destination_entry, mode_var, result_label):
        origin = origin_entry.get().strip()
        destination = destination_entry.get().strip()
        mode = mode_var.get()

        if not origin or not destination:
            result_label.config(text="Enter both origin and destination.")
            return

        if mode == "None":
            result_label.config(text="Select a valid mode.")
            return
        try:
            map_client = googlemaps.Client(API_KEY)
            origin_geocode = map_client.geocode(origin)
            destination_geocode = map_client.geocode(destination)

            if not origin_geocode or not destination_geocode:
                messagebox.showerror("Geocoding Error", "Invalid origin or destination address.")
                result_label.config(text="Invalid origin or destination.")
                return

            origin_coords = (
                origin_geocode[0]['geometry']['location']['lat'],
                origin_geocode[0]['geometry']['location']['lng']
            )
            destination_coords = (
                destination_geocode[0]['geometry']['location']['lat'],
                destination_geocode[0]['geometry']['location']['lng']
            )

            if mode == "plane":
                distance = geodesic(origin_coords, destination_coords).kilometers
                result_label.config(text=f"Distance: {distance:.2f} km (by plane)")
            elif mode == "ship":
                # Update the result_label immediately
                result_label.config(text="Calculating sea route distance...")
                calculate_sea_route_async(origin, destination, result_label)

                def calculate_sea_route():
                    result = visualize_sea_route(origin, destination)
                    if result and len(result) >= 3:
                        figure, distance_nautical, distance_km = result[:3]
                        if distance_km is not None and distance_nautical is not None:
                            if result_label.winfo_exists():
                                result_label.config(
                                    text=f"Distance: {distance_km:.2f} km (by ship)\n({distance_nautical:.2f} nautical miles)"
                                )
                        else:
                            if result_label.winfo_exists():
                                result_label.config(text="Error calculating sea route distance.")
                    else:
                        if result_label.winfo_exists():
                            result_label.config(text="Error calculating sea route distance.")
                threading.Thread(target=calculate_sea_route).start()
            else:
                distance_result = map_client.distance_matrix(origins=[origin], destinations=[destination], mode=mode)
                distance = distance_result['rows'][0]['elements'][0]['distance']['text']
                duration = distance_result['rows'][0]['elements'][0]['duration']['text']
                result_label.config(text=f"Distance: {distance}\nDuration: {duration}")

                print(f"[DEBUG] Origin Coordinates: {origin_coords}")
                print(f"[DEBUG] Destination Coordinates: {destination_coords}")
        except Exception as e:
            messagebox.showerror("API Error", f"An error occurred while contacting the map service:\n{e}")
            result_label.config(text=f"Error: {e}")

    def calculate_distance_dynamic_with_callback(origin_entry, destination_entry, mode_var, result_label, callback=None):
        origin = origin_entry.get().strip()
        destination = destination_entry.get().strip()
        mode = mode_var.get()

        if not origin or not destination:
            result_label.config(text="Enter both origin and destination.")
            if callback:
                callback()
            return

        if mode == "None":
            result_label.config(text="Select a valid mode.")
            if callback:
                callback()
            return
        try:
            map_client = googlemaps.Client(API_KEY)
            origin_geocode = map_client.geocode(origin)
            destination_geocode = map_client.geocode(destination)

            if not origin_geocode or not destination_geocode:
                messagebox.showerror("Geocoding Error", "Invalid origin or destination address.")
                result_label.config(text="Invalid origin or destination.")
                if callback:
                    callback()
                return

            origin_coords = (
                origin_geocode[0]['geometry']['location']['lat'],
                origin_geocode[0]['geometry']['location']['lng']
            )
            destination_coords = (
                destination_geocode[0]['geometry']['location']['lat'],
                destination_geocode[0]['geometry']['location']['lng']
            )

            if mode == "plane":
                distance = geodesic(origin_coords, destination_coords).kilometers
                result_label.config(text=f"Distance: {distance:.2f} km (by plane)")
                if callback:
                    callback()
            elif mode == "ship":
                result_label.config(text="Calculating sea route distance...")
                def after_sea_route():
                    if callback:
                        callback()
                calculate_sea_route_async(origin, destination, result_label)
                result_label.after(500, after_sea_route)
            else:
                distance_result = map_client.distance_matrix(origins=[origin], destinations=[destination], mode=mode)
                distance = distance_result['rows'][0]['elements'][0]['distance']['text']
                duration = distance_result['rows'][0]['elements'][0]['duration']['text']
                result_label.config(text=f"Distance: {distance}\nDuration: {duration}")
                if callback:
                    callback()
        except Exception as e:
            messagebox.showerror("API Error", f"An error occurred while contacting the map service:\n{e}")
            result_label.config(text=f"Error: {e}")
            if callback:
                callback()

    def update_transport_types(event):
        selected_transport = event.widget.get()
        print(f"[DEBUG] update_transport_types called for: {selected_transport}")
        transport_combo = event.widget
        clear_variations_and_fuel_types(transport_combo)

        transport_combo.value_label.config(text="0.0")

        if selected_transport.lower() not in ["van", "hgv (diesel)", "hgv refrigerated (diesel)"]:
            transport_combo.value_frame.config(text="Carbon Emission Amount\nper 1000kgkm")
        else:
            transport_combo.value_frame.config(text="Carbon Emission Amount\nper km")

        if selected_transport:
            transport_combo.variation_combo.config(state="readonly")
        else:
            transport_combo.variation_combo.set('')
            transport_combo.variation_combo.config(state="disabled")

        # Set up transport-specific options and enable/disable fuel combo accordingly
        if selected_transport.lower() == "van":
            transport_combo.variation_combo["values"] = van_list
            transport_combo.fuel_combo["values"] = ["Diesel", "Petrol", "Battery"]
            transport_combo.fuel_combo.config(state="readonly")
        elif selected_transport.lower() == "hgv (diesel)":
            transport_combo.variation_combo["values"] = HGV_diesel_list
            transport_combo.fuel_combo["values"] = ["0% Laden", "50% Laden", "100% Laden"]
            transport_combo.fuel_combo.config(state="readonly")
        elif selected_transport.lower() == "hgv refrigerated (diesel)":
            transport_combo.variation_combo["values"] = HGV_R_diesel_list
            transport_combo.fuel_combo["values"] = ["0% Laden", "50% Laden", "100% Laden"]
            transport_combo.fuel_combo.config(state="readonly")
        elif selected_transport.lower() == "freight flights":
            transport_combo.variation_combo["values"] = flight_list
            transport_combo.fuel_combo["values"] = ["International", "<1500km", ">1500km"]
            transport_combo.fuel_combo.config(state="readonly")
        elif selected_transport.lower() == "rail":
            transport_combo.variation_combo["values"] = rail_list
            transport_combo.fuel_combo["values"] = ["Freight Train"]
            transport_combo.fuel_combo.config(state="readonly")
        elif selected_transport.lower() == "sea tanker":
            transport_combo.variation_combo["values"] = sea_tanker_list
            transport_combo.fuel_combo["values"] = []
            transport_combo.fuel_combo.set('')
            transport_combo.fuel_combo.config(state="readonly")
        elif selected_transport.lower() == "cargo ship":
            transport_combo.variation_combo["values"] = cargo_list
            transport_combo.fuel_combo["values"] = []
            transport_combo.fuel_combo.set('')
            transport_combo.fuel_combo.config(state="readonly")
        else:
            transport_combo.variation_combo["values"] = []
            transport_combo.variation_combo.set('')
            transport_combo.variation_combo.config(state="readonly")
            transport_combo.fuel_combo["values"] = []
            transport_combo.fuel_combo.set('')
            transport_combo.fuel_combo.config(state="readonly")
            transport_combo.value_label.config(text="")  # Clear the output

        # Dynamically update mode options
        mode_menu = transport_combo.mode_menu  # Access the mode menu
        if selected_transport.lower() in ["van", "hgv (diesel)", "hgv refrigerated (diesel)"]:
            mode_options = ["driving"]  # Only driving mode
        elif selected_transport.lower() == "freight flights":
            mode_options = ["plane"]  # Only plane mode
        elif selected_transport.lower() in ["sea tanker", "cargo ship"]:
            mode_options = ["ship"]  # Only ship mode
        else:
            mode_options = ["None", "driving", "plane", "ship"]  # Full list of modes

        # Update the mode menu
        menu = mode_menu["menu"]
        menu.delete(0, "end")  
        for mode in mode_options:
            menu.add_command(label=mode, command=tk._setit(transport_combo.mode_var, mode))
        transport_combo.mode_var.set("None")  # Reset to default

    def clear_variations_and_fuel_types(widget):
        if hasattr(widget, 'variation_combo'):
            widget.variation_combo.set('')  
            widget.variation_combo["values"] = []  
        if hasattr(widget, 'fuel_combo'):
            widget.fuel_combo.set('') 
            widget.fuel_combo["values"] = []  

    def update_value(event, transport_combo):
        selected_transport = transport_combo.get()  
        selected_variation = transport_combo.variation_combo.get()
        selected_fuel = transport_combo.fuel_combo.get()
        index = -1

        print(f"[DEBUG] Selected Transport: {selected_transport}, Variation: {selected_variation}, Fuel: {selected_fuel}")

        if selected_transport.lower() == "van":
            index = van_list.index(selected_variation) if selected_variation in van_list else -1
            if selected_fuel == "Diesel" and index != -1:
                transport_combo.value_label.config(text=van_diesel_list[index])
            elif selected_fuel == "Petrol" and index != -1:
                transport_combo.value_label.config(text=van_petrol_list[index])
            elif selected_fuel == "Battery" and index != -1:
                transport_combo.value_label.config(text=van_battery_list[index])
            else:
                transport_combo.value_label.config(text="")

        elif selected_transport.lower() == "hgv (diesel)":
            index = HGV_diesel_list.index(selected_variation) if selected_variation in HGV_diesel_list else -1
            if selected_fuel == "0% Laden" and index != -1:
                transport_combo.value_label.config(text=HGV_0_diesel_list[index])
            elif selected_fuel == "50% Laden" and index != -1:
                transport_combo.value_label.config(text=HGV_50_diesel_list[index])
            elif selected_fuel == "100% Laden" and index != -1:
                transport_combo.value_label.config(text=HGV_100_diesel_list[index])
            else:
                transport_combo.value_label.config(text="")

        elif selected_transport.lower() == "hgv refrigerated (diesel)":
            index = HGV_R_diesel_list.index(selected_variation) if selected_variation in HGV_R_diesel_list else -1
            if selected_fuel == "0% Laden" and index != -1:
                transport_combo.value_label.config(text=HGV_R_0_diesel_list[index])
            elif selected_fuel == "50% Laden" and index != -1:
                transport_combo.value_label.config(text=HGV_R_50_diesel_list[index])
            elif selected_fuel == "100% Laden" and index != -1:
                transport_combo.value_label.config(text=HGV_R_100_diesel_list[index])
            else:
                transport_combo.value_label.config(text="")

        elif selected_transport.lower() == "freight flights" or selected_transport.lower() == "belly freight flights":
            index = flight_list.index(selected_variation) if selected_variation in flight_list else -1
            if selected_fuel == "International" and index != -1:
                if index < len(Int_flight_list):
                    transport_combo.value_label.config(text=Int_flight_list[index])
                else:
                    transport_combo.value_label.config(text="NIL")
            elif selected_fuel == "<1500km" and index != -1:
                if index < len(L_1500_flight_list):
                    transport_combo.value_label.config(text=L_1500_flight_list[index])
                else:
                    transport_combo.value_label.config(text="NIL")
            elif selected_fuel == ">1500km" and index != -1:
                if index < len(M_1500_flight_list):
                    transport_combo.value_label.config(text=M_1500_flight_list[index])
                else:
                    transport_combo.value_label.config(text="NIL")
            else:
                transport_combo.value_label.config(text="")

        elif selected_transport.lower() == "rail":
            index = rail_list.index(selected_variation) if selected_variation in rail_list else -1
            if selected_fuel == "Freight Train" and index != -1:
                if index < len(freight_train_rail_list):
                    transport_combo.value_label.config(text=freight_train_rail_list[index])
                else:
                    transport_combo.value_label.config(text="NIL")
            else:
                transport_combo.value_label.config(text="")

        elif selected_transport.lower() == "sea tanker":
            index = sea_tanker_list.index(selected_variation) if selected_variation in sea_tanker_list else -1
            if selected_variation == "Crude Tanker" and index != -1:
                transport_combo.fuel_combo["values"] = extract_list(crude_tanker_cells)
                update_fuel_value(crude_value_cells, transport_combo)
            elif selected_variation == "Products Tanker" and index != -1:
                transport_combo.fuel_combo["values"] = extract_list(products_tanker_cells)
                update_fuel_value(products_value_cells, transport_combo)
            elif selected_variation == "Chemical Tanker" and index != -1:
                transport_combo.fuel_combo["values"] = extract_list(chemical_tanker_cells)
                update_fuel_value(chemical_value_cells, transport_combo)
            elif selected_variation == "LNG Tanker" and index != -1:
                transport_combo.fuel_combo["values"] = extract_list(LNG_tanker_cells)
                update_fuel_value(LNG_value_cells, transport_combo)
            elif selected_variation == "LPG Tanker" and index != -1:
                transport_combo.fuel_combo["values"] = extract_list(LPG_tanker_cells)
                update_fuel_value(LPG_value_cells, transport_combo)
            else:
                transport_combo.fuel_combo["values"] = []

        elif selected_transport.lower() == "cargo ship":
            index = cargo_list.index(selected_variation) if selected_variation in cargo_list else -1
            if selected_variation == "Bulk Carrier" and index != -1:
                transport_combo.fuel_combo["values"] = extract_list(Bulk_carrier_cells)
                update_fuel_value(bulk_value_cells, transport_combo)
            elif selected_variation == "General Cargo" and index != -1:
                transport_combo.fuel_combo["values"] = extract_list(General_cargo_cells)
                update_fuel_value(general_value_cells, transport_combo)
            elif selected_variation == "Container Ship" and index != -1:
                transport_combo.fuel_combo["values"] = extract_list(Container_ship_cells)
                update_fuel_value(container_value_cells, transport_combo)
            elif selected_variation == "Vehicle Transport" and index != -1:
                transport_combo.fuel_combo["values"] = extract_list(Vehicle_transport_cells)
                update_fuel_value(vehicle_value_cells, transport_combo)
            elif selected_variation == "RoRo Ferry" and index != -1:
                transport_combo.fuel_combo["values"] = extract_list(Roro_ferry_cells)
                update_fuel_value(roro_value_cells, transport_combo)
            else:
                transport_combo.fuel_combo["values"] = []
        else:
            transport_combo.value_label.config(text="")
        print(f"[DEBUG] Updated Value Label: {transport_combo.value_label.cget('text')}")

    def update_fuel_value(value_cells, widget):
        selected_fuel = widget.fuel_combo.get()
        if selected_fuel:
            if selected_fuel in widget.fuel_combo["values"]:
                index = widget.fuel_combo["values"].index(selected_fuel)
                widget.value_label.config(text=value_cells[index][0].value)
            else:
                widget.value_label.config(text="Variation Unavailable")
        else:
            widget.value_label.config(text="")

    def update_transportation_country_values(selected_transport_country):
        """
        Updates the carbon emission overview based on the selected transportation country.
        """
        try:
            print(f"[DEBUG] Selected Transportation Country: {selected_transport_country}")
            if selected_transport_country in country_list:
                transport_index = country_list.index(selected_transport_country)
                print(f"[DEBUG] Transportation Country Index: {transport_index}")

                # Update material emission labels in materials_group_frame
                steel_label.config(text=steel_list[transport_index])
                print(f"[DEBUG] Steel Label set to: {steel_list[transport_index]}")
                aluminium_label.config(text=aluminium_list[transport_index])
                print(f"[DEBUG] Aluminium Label set to: {aluminium_list[transport_index]}")
                cement_label.config(text=cement_list[transport_index])
                print(f"[DEBUG] Cement Label set to: {cement_list[transport_index]}")
                plastic_label.config(text=plastic_list[transport_index])
                print(f"[DEBUG] Plastic Label set to: {plastic_list[transport_index]}")
                carbon_fiber_label.config(text=carbon_fiber_list[transport_index])
                print(f"[DEBUG] Carbon Fiber Label set to: {carbon_fiber_list[transport_index]}")

                custom_elec = custom_emission_factors.get("electricity")
                if custom_elec is not None and custom_elec != "":
                    try:
                        electricity_emission = float(custom_elec)
                        electricity_label.config(text=f"{electricity_emission:.2f}")
                        print(f"[DEBUG] [Custom] Electricity Label set to: {electricity_emission:.2f}")
                    except Exception:
                        electricity_label.config(text="0.0")
                        print(f"[DEBUG] [Custom] Invalid electricity emission value, set to 0.0")
                else:
                    try:
                        electricity_emission = float(electricity_list[transport_index])
                        electricity_label.config(text=f"{electricity_emission:.2f}")
                        print(f"[DEBUG] Electricity Label set to: {electricity_emission:.2f}")
                    except (ValueError, IndexError):
                        electricity_label.config(text="0.0")
                        print(f"[DEBUG] Missing or invalid electricity emission value, set to 0.0")

                # Debugging: Log updated values
                print(f"[DEBUG] Updated Steel Label: {steel_label.cget('text')}")
                print(f"[DEBUG] Updated Aluminium Label: {aluminium_label.cget('text')}")
                print(f"[DEBUG] Updated Cement Label: {cement_label.cget('text')}")
                print(f"[DEBUG] Updated Plastic Label: {plastic_label.cget('text')}")
                print(f"[DEBUG] Updated Carbon Fiber Label: {carbon_fiber_label.cget('text')}")
                print(f"[DEBUG] Updated Electricity Label: {electricity_label.cget('text')}")
            else:
                print(f"[DEBUG] Invalid Transportation Country: {selected_transport_country}")
        except Exception as e:
            print(f"[DEBUG] Error updating transportation country values: {e}")

    def update_transportation_country_values_direct(selected_transport_country):
        """
        Force direct update of carbon emission overview labels regardless of widget state
        """
        try:
            print("\n[DEBUG] ===== FORCE UPDATING EMISSION LABELS =====")
            print(f"[DEBUG] Force updating emission labels for country: {selected_transport_country}")
            if selected_transport_country in country_list:
                transport_index = country_list.index(selected_transport_country)
                print(f"[DEBUG] Transportation Country Index: {transport_index}")
    
                # Get the emission values for this country
                steel_value = custom_emission_factors.get("materials", {}).get("Steel", steel_list[transport_index])
                aluminium_value = custom_emission_factors.get("materials", {}).get("Aluminium", aluminium_list[transport_index])
                cement_value = custom_emission_factors.get("materials", {}).get("Cement", cement_list[transport_index])
                plastic_value = custom_emission_factors.get("materials", {}).get("Plastic", plastic_list[transport_index])
                carbon_fiber_value = custom_emission_factors.get("materials", {}).get("Carbon Fiber", carbon_fiber_list[transport_index])
                electricity_value = electricity_list[transport_index]

                # Print debug for each value with distinctive formatting
                print(f"[DEBUG] >>> Steel Label set to: {steel_value}")
                print(f"[DEBUG] >>> Aluminium Label set to: {aluminium_value}")
                print(f"[DEBUG] >>> Cement Label set to: {cement_value}")
                print(f"[DEBUG] >>> Plastic Label set to: {plastic_value}")
                print(f"[DEBUG] >>> Carbon Fiber Label set to: {carbon_fiber_value}")
                print(f"[DEBUG] >>> Electricity Label set to: {electricity_value}")

                # Update global emission labels directly
                global steel_label, aluminium_label, cement_label, plastic_label, carbon_fiber_label, electricity_label
            
                # Use direct configuration to ensure UI updates
                if 'steel_label' in globals() and steel_label.winfo_exists():
                    steel_label.config(text=str(steel_value))
                    print(f"[DEBUG] UI Steel Label updated to: {steel_label.cget('text')}")
            
                if 'aluminium_label' in globals() and aluminium_label.winfo_exists():
                    aluminium_label.config(text=str(aluminium_value))
                    print(f"[DEBUG] UI Aluminium Label updated to: {aluminium_label.cget('text')}")
            
                if 'cement_label' in globals() and cement_label.winfo_exists():
                    cement_label.config(text=str(cement_value))
                    print(f"[DEBUG] UI Cement Label updated to: {cement_label.cget('text')}")
            
                if 'plastic_label' in globals() and plastic_label.winfo_exists():
                    plastic_label.config(text=str(plastic_value))
                    print(f"[DEBUG] UI Plastic Label updated to: {plastic_label.cget('text')}")
            
                if 'carbon_fiber_label' in globals() and carbon_fiber_label.winfo_exists():
                    carbon_fiber_label.config(text=str(carbon_fiber_value))
                    print(f"[DEBUG] UI Carbon Fiber Label updated to: {carbon_fiber_label.cget('text')}")
            
                if 'electricity_label' in globals() and electricity_label.winfo_exists():
                    electricity_label.config(text=str(electricity_value))
                    print(f"[DEBUG] UI Electricity Label updated to: {electricity_label.cget('text')}")
            
                # Force UI update immediately
                if 'root' in globals() and root.winfo_exists():
                    root.update_idletasks()
                    print("[DEBUG] Forced UI update with root.update_idletasks()")
            
                print("[DEBUG] calculate_material_value() triggered")
                calculate_material_value()
            
                # Additional recalculation to ensure all dependent values are updated
                recalculate_all()
    
                print(f"[DEBUG] Successfully updated all emission labels for {selected_transport_country}")
                print(f"[DEBUG] Updated Carbon Emission Overview for country: {selected_transport_country}")
                print("[DEBUG] ===== EMISSION LABELS UPDATE COMPLETE =====\n")
                return True
            else:
                print(f"[DEBUG] Country not found in country list: {selected_transport_country}")
                return False
        except Exception as e:
            print(f"[DEBUG] ERROR in force update emission labels: {e}")
            import traceback
            traceback.print_exc() 
            return False

    def validate_float_input(P):
        """
        Validates if the input string can be converted to a float.
        Allows empty input for flexibility.
        """
        print(f"[DEBUG] Validating Input: '{P}'")  # Debugging
        if not P:  # Allow empty input
            return True
        try:
            float(P)  
            return True
        except ValueError:
            return False  # Reject invalid input

    def on_transport_country_change(event):
        if restoring_profile:
            return
        selected_transport_country = transport_country_combo.get()
        print(f"[DEBUG] Transport Country Changed: {selected_transport_country}")

        # Update Carbon Emission Overview
        update_transportation_country_values(selected_transport_country)

        # Recalculate Material Value
        calculate_material_value()

    def update_manufacturing_country_values(selected_manufacturing_country):
        """
        Updates other values based on the selected manufacturing country.
        Does not affect carbon emission values.
        """
        try:
            if selected_manufacturing_country in country_list:
                manufacturing_index = country_list.index(selected_manufacturing_country)
                print(f"[DEBUG] Manufacturing Country Index: {manufacturing_index}")

                print(f"[DEBUG] Manufacturing Country Selected: {selected_manufacturing_country}")
            else:
                print(f"[DEBUG] Invalid Manufacturing Country: {selected_manufacturing_country}")
        except Exception as e:
            print(f"[DEBUG] Error updating manufacturing country values: {e}")

    def rebuild_carbon_emission_overview(selected_country):
        """
        Complete rebuild of carbon emission overview by reusing the existing materials_group_frame
        """
        try:
            print(f"\n[DEBUG] REBUILDING carbon emission overview for {selected_country}")
            if selected_country in country_list:
                transport_index = country_list.index(selected_country)
            
                # Get emission values (use custom factors if available)
                steel_value = custom_emission_factors.get("materials", {}).get("Steel", steel_list[transport_index])
                aluminium_value = custom_emission_factors.get("materials", {}).get("Aluminium", aluminium_list[transport_index])
                cement_value = custom_emission_factors.get("materials", {}).get("Cement", cement_list[transport_index])
                plastic_value = custom_emission_factors.get("materials", {}).get("Plastic", plastic_list[transport_index])
                carbon_fiber_value = custom_emission_factors.get("materials", {}).get("Carbon Fiber", carbon_fiber_list[transport_index])
                electricity_value = electricity_list[transport_index]
            
                # Clear materials_group_frame (keep the frame but remove children)
                for widget in materials_group_frame.winfo_children():
                    widget.destroy()
                     
                # Configure the grid
                materials_group_frame.grid_rowconfigure(0, weight=1)
                materials_group_frame.grid_rowconfigure(1, weight=1)
                materials_group_frame.grid_rowconfigure(2, weight=1)
                materials_group_frame.grid_columnconfigure(0, weight=1)
                materials_group_frame.grid_columnconfigure(1, weight=1)
            
                # Steel
                global steel_frame, steel_label
                steel_frame = ttk.Labelframe(materials_group_frame, text="Steel (CO2eKg)", style="Eco.TLabelframe")
                steel_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
                steel_frame.grid_rowconfigure(0, weight=1)
                steel_frame.grid_columnconfigure(0, weight=1)
                steel_label = ttk.Label(steel_frame, text=str(steel_value), style="Eco.TLabel")
                steel_label.grid(row=0, column=0, sticky="")
            
                # Aluminium
                global aluminium_frame, aluminium_label
                aluminium_frame = ttk.Labelframe(materials_group_frame, text="Aluminium (CO2eKg)", style="Eco.TLabelframe")
                aluminium_frame.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)
                aluminium_frame.grid_rowconfigure(0, weight=1)
                aluminium_frame.grid_columnconfigure(0, weight=1)
                aluminium_label = ttk.Label(aluminium_frame, text=str(aluminium_value), style="Eco.TLabel")
                aluminium_label.grid(row=0, column=0, sticky="")
            
                # Cement
                global cement_frame, cement_label
                cement_frame = ttk.Labelframe(materials_group_frame, text="Cement (CO2eKg)", style="Eco.TLabelframe")
                cement_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
                cement_frame.grid_rowconfigure(0, weight=1)
                cement_frame.grid_columnconfigure(0, weight=1)
                cement_label = ttk.Label(cement_frame, text=str(cement_value), style="Eco.TLabel")
                cement_label.grid(row=0, column=0, sticky="")
            
                # Carbon Fiber
                global carbon_fiber_frame, carbon_fiber_label
                carbon_fiber_frame = ttk.Labelframe(materials_group_frame, text="Global Carbon Fiber (CO2eKg)", style="Eco.TLabelframe")
                carbon_fiber_frame.grid(row=1, column=1, sticky="nsew", padx=10, pady=10)
                carbon_fiber_frame.grid_rowconfigure(0, weight=1)
                carbon_fiber_frame.grid_columnconfigure(0, weight=1)
                carbon_fiber_label = ttk.Label(carbon_fiber_frame, text=str(carbon_fiber_value), style="Eco.TLabel")
                carbon_fiber_label.grid(row=0, column=0, sticky="")
            
                # Plastic
                global plastic_frame, plastic_label
                plastic_frame = ttk.Labelframe(materials_group_frame, text="Global Plastic (CO2eKg)", style="Eco.TLabelframe")
                plastic_frame.grid(row=2, column=0, sticky="nsew", padx=10, pady=10)
                plastic_frame.grid_rowconfigure(0, weight=1)
                plastic_frame.grid_columnconfigure(0, weight=1)
                plastic_label = ttk.Label(plastic_frame, text=str(plastic_value), style="Eco.TLabel")
                plastic_label.grid(row=0, column=0, sticky="")
            
                # Electricity
                global electricity_frame, electricity_label
                electricity_frame = ttk.Labelframe(materials_group_frame, text="Electricity (CO2eKg/kWh)", style="Eco.TLabelframe")
                electricity_frame.grid(row=2, column=1, sticky="nsew", padx=10, pady=10)
                electricity_frame.grid_rowconfigure(0, weight=1)
                electricity_frame.grid_columnconfigure(0, weight=1)
                electricity_label = ttk.Label(electricity_frame, text=str(electricity_value), style="Eco.TLabel")
                electricity_label.grid(row=0, column=0, sticky="")
            
                materials_group_frame.update_idletasks()
            
                print(f"[DEBUG] Successfully rebuilt carbon emission overview for {selected_country}")
                print(f"[DEBUG] Steel Label: {steel_label.cget('text')}")
                print(f"[DEBUG] Aluminium Label: {aluminium_label.cget('text')}")
                print(f"[DEBUG] Cement Label: {cement_label.cget('text')}")
                print(f"[DEBUG] Plastic Label: {plastic_label.cget('text')}")
                print(f"[DEBUG] Carbon Fiber Label: {carbon_fiber_label.cget('text')}")
                print(f"[DEBUG] Electricity Label: {electricity_label.cget('text')}")
            
                return True
            else:
                print(f"[DEBUG] Country not found in country list: {selected_country}")
                return False
        except Exception as e:
            print(f"[DEBUG] ERROR in rebuild_carbon_emission_overview: {e}")
            import traceback
            traceback.print_exc()
            return False

    def force_update_emission_labels_after_import(selected_transport_country):
        """
        Emergency fix for Excel imports - directly accesses and updates UI elements
        regardless of application state
        """
        try:
            print("\n[DEBUG] === EMERGENCY FIX FOR EXCEL IMPORT ===")
            print(f"[DEBUG] Forcing emission label updates for: {selected_transport_country}")
        
            # Get country index
            if selected_transport_country not in country_list:
                print(f"[DEBUG] ERROR: Country '{selected_transport_country}' not in country list!")
                return False
            
            transport_index = country_list.index(selected_transport_country)
        
            # Get emission values
            emission_values = {
                "Steel": custom_emission_factors.get("materials", {}).get("Steel", steel_list[transport_index]),
                "Aluminium": custom_emission_factors.get("materials", {}).get("Aluminium", aluminium_list[transport_index]),
                "Cement": custom_emission_factors.get("materials", {}).get("Cement", cement_list[transport_index]),
                "Plastic": custom_emission_factors.get("materials", {}).get("Plastic", plastic_list[transport_index]),
                "Carbon Fiber": custom_emission_factors.get("materials", {}).get("Carbon Fiber", carbon_fiber_list[transport_index]),
                "Electricity": electricity_list[transport_index]
            }
        
            for key, value in emission_values.items():
                print(f"[DEBUG] {key} emission value: {value}")
        
            global steel_label, aluminium_label, cement_label, plastic_label, carbon_fiber_label, electricity_label
            global steel_frame, aluminium_frame, cement_frame, plastic_frame, carbon_fiber_frame, electricity_frame
        
            for label_name in ['steel_label', 'aluminium_label', 'cement_label', 'plastic_label', 'carbon_fiber_label', 'electricity_label']:
                if label_name in globals() and globals()[label_name] and globals()[label_name].winfo_exists():
                    globals()[label_name].destroy()
        
            # Create new labels with the correct values
            steel_label = ttk.Label(steel_frame, text=str(emission_values["Steel"]), style="Eco.TLabel")
            steel_label.grid(row=0, column=0, sticky="")
        
            aluminium_label = ttk.Label(aluminium_frame, text=str(emission_values["Aluminium"]), style="Eco.TLabel")
            aluminium_label.grid(row=0, column=0, sticky="")
        
            cement_label = ttk.Label(cement_frame, text=str(emission_values["Cement"]), style="Eco.TLabel")
            cement_label.grid(row=0, column=0, sticky="")
        
            plastic_label = ttk.Label(plastic_frame, text=str(emission_values["Plastic"]), style="Eco.TLabel")
            plastic_label.grid(row=0, column=0, sticky="")
        
            carbon_fiber_label = ttk.Label(carbon_fiber_frame, text=str(emission_values["Carbon Fiber"]), style="Eco.TLabel")
            carbon_fiber_label.grid(row=0, column=0, sticky="")
        
            electricity_label = ttk.Label(electricity_frame, text=str(emission_values["Electricity"]), style="Eco.TLabel")
            electricity_label.grid(row=0, column=0, sticky="")
        
            # Verify the new labels show the correct values
            print(f"[DEBUG] New Steel Label text: {steel_label.cget('text')}")
            print(f"[DEBUG] New Aluminium Label text: {aluminium_label.cget('text')}")
            print(f"[DEBUG] New Cement Label text: {cement_label.cget('text')}")
            print(f"[DEBUG] New Plastic Label text: {plastic_label.cget('text')}")
            print(f"[DEBUG] New Carbon Fiber Label text: {carbon_fiber_label.cget('text')}")
            print(f"[DEBUG] New Electricity Label text: {electricity_label.cget('text')}")
        
            # Force UI refresh
            root.update_idletasks()
        
            print("[DEBUG] === EMERGENCY FIX COMPLETE ===\n")
            return True
        except Exception as e:
            import traceback
            print(f"[DEBUG] ERROR in emergency fix: {e}")
            traceback.print_exc()
            return False

    def calculate_material_value():
        global restoring_profile
        if restoring_profile:
            return
        print("[DEBUG] calculate_material_value() triggered")

        selected_transport_country = transport_country_combo.get()
        selected_material = material_combo.get()

        total_weight = get_total_weight()
        print(f"[DEBUG] Total Weight (user unit): '{total_weight}'")
        print(f"[DEBUG] Current unit: '{current_unit}'")

        weight_in_kg = total_weight
        print(f"[DEBUG] Weight in kg for emission calculation: '{weight_in_kg}'")

        try:
            if weight_in_kg == 0.0:
                raise ValueError("Total weight is zero")
        except ValueError as e:
            print(f"[DEBUG] Invalid total weight: {e}")
            calculated_value_label.config(text="Value: 0.0")
            return

        material_emission = None
        if selected_material in custom_emission_factors.get("materials", {}):
            try:
                material_emission = float(custom_emission_factors["materials"][selected_material])
                print(f"[DEBUG] Using custom emission factor for {selected_material}: {material_emission}")
            except Exception:
                material_emission = None

        if material_emission is None and selected_transport_country in country_list and selected_material in materials:
            transport_index = country_list.index(selected_transport_country)
            material_index = materials.index(selected_material)
            material_emission_lists = [steel_list, aluminium_list, cement_list, plastic_list, carbon_fiber_list]
            material_emission = float(material_emission_lists[material_index][transport_index])
            print(f"[DEBUG] Using default emission factor for {selected_material}: {material_emission}")

        if material_emission is not None:
            total_emission = weight_in_kg * material_emission
            calculated_value_label.config(text=f"Value: {total_emission:.2f}")
            print(f"[DEBUG] Calculated Material Emission: {total_emission:.2f}")
        else:
            print("[DEBUG] Invalid transportation country or material selection")
            calculated_value_label.config(text="Value: 0.0")

        recalculate_all_carbon_emissions()
        update_sum_of_emissions()

    def calculate_total_emission(value_label, result_label, total_emission_label, weight_entry, transport_combo):
        row_widgets = None
        for row in transport_row_widgets:
            if (
                row.get("value_label") == value_label and
                row.get("result_label") == result_label and
                row.get("total_emission_label") == total_emission_label and
                row.get("weight_entry") == weight_entry and
                row.get("transport_combo") == transport_combo
            ):
                row_widgets = row
                break

        try:
            # Debugging: Log the inputs
            print(f"[DEBUG] Value Label: {value_label.cget('text')}")
            print(f"[DEBUG] Result Label: {result_label.cget('text')}")
            print(f"[DEBUG] Weight Entry: {weight_entry.get()}")
            print(f"[DEBUG] Transport Combo: {transport_combo.get()}")

            # Get the carbon emission amount
            transport_type_name = transport_combo.get()
            custom_tr = custom_emission_factors.get("transport", {}).get(transport_type_name)
            if custom_tr is not None and custom_tr != "":
                try:
                    carbon_emission = float(custom_tr)
                    print(f"[DEBUG] Using custom emission factor for transport {transport_type_name}: {carbon_emission}")
                except Exception:
                    try:
                        carbon_emission = float(value_label.cget("text"))
                    except ValueError:
                        total_emission_label.config(text="Invalid carbon emission")
                        return
            else:
                try:
                    carbon_emission = float(value_label.cget("text"))
                except ValueError:
                    total_emission_label.config(text="Invalid carbon emission")
                    return

            # Get the distance (strip "km" and convert to float)
            distance_text = result_label.cget("text")
            if "Distance:" in distance_text:
                try:
                    distance = float(distance_text.split(":")[1].split("km")[0].strip().replace(",", ""))
                except ValueError:
                    total_emission_label.config(text="Invalid distance")
                    return
            else:
                total_emission_label.config(text="Invalid distance")
                return

            # Use the new total weight (weight * quantity)
            total_weight = get_total_weight()

            # Check the transport type and apply the appropriate formula
            transport_type = transport_combo.get().lower()
            if transport_type not in ["van", "hgv (diesel)", "hgv refrigerated (diesel)"]:
                # Formula: total_weight * (distance / 1000) * carbon emission
                total_emission = total_weight * (distance / 1000) * carbon_emission
            else:
                # Default formula: distance * carbon emission
                total_emission = carbon_emission * distance

            # Update the total emission label
            # --- Ensure rounding here ---
            total_emission = round(total_emission, 2)
            total_emission_label.config(text=f"Total Emission: {total_emission:.2f} CO2ekg")
            print(f"[DEBUG] Calculated Total Emission: {total_emission_label.cget('text')}")
        except Exception as e:
            total_emission_label.config(text="Error calculating total emission")
            print(f"[DEBUG] Error in calculate_total_emission: {e}")

    total_emission_labels = [] # Global list to track total emission labels

    def debounce_weight_calculation():
        global debounce_timer
        if debounce_timer:
            debounce_timer.cancel()  
        print("[DEBUG] Starting debounce timer for weight_entry")
        debounce_timer = Timer(0.5, lambda: [calculate_material_value(), recalculate_all_emissions()])
        debounce_timer.start()

    def recalculate_all_emissions(event=None):
        print("[DEBUG] Recalculating emissions for all rows...")
        for row_widgets in transport_row_widgets:
            print(f"[DEBUG] Processing row: {row_widgets}")
            if "value_label" not in row_widgets or "transport_combo" not in row_widgets:
                print("[DEBUG] Missing required widgets in row, skipping...")
                continue
            calculate_total_emission(
                value_label=row_widgets["value_label"],
                result_label=row_widgets["result_label"],
                total_emission_label=row_widgets["total_emission_label"],
                weight_entry=weight_entry,
                transport_combo=row_widgets["transport_combo"]
            )
        update_sum_of_emissions()  # Update the total sum of emissions

    def recalculate_all():
        global restoring_profile
        if restoring_profile:
            return
        print("[DEBUG] Recalculating all values...")
        update_transportation_country_values(transport_country_combo.get())  # Update Carbon Emission Overview
        calculate_material_value()  # Recalculate material value
        recalculate_all_carbon_emissions()  # Recalculate carbon emissions
        update_recycled_weight(weight_entry, weight_after_entry, recycled_weight_value_label) 
        update_sum_of_emissions()  
       
    def update_sum_of_emissions():
        try:
            try:
                material_emission = float(calculated_value_label.cget("text").split(":")[1].strip())
            except (IndexError, ValueError):
                material_emission = 0.0  # Default to 0 if parsing fails
                print(f"[DEBUG] Material Emission: {calculated_value_label.cget('text')}")

            print(f"[DEBUG] Material Emission: {material_emission}")  
            # Sum up all total_emission_label values (each row rounded before summing)
            total_emission = material_emission
            for label in total_emission_labels[:]:
                if label and label.winfo_exists():
                    emission_text = label.cget("text")
                    print(f"[DEBUG] Processing Label: {label.cget('text')}") 
                    if "Total Emission:" in emission_text:
                        import re
                        emission_value = float(re.sub(r'[^0-9.\-]+', '', emission_text.split(":")[1]))
                        emission_value = round(emission_value, 2)  # Round each row before summing
                        total_emission += emission_value
                else:
                    total_emission_labels.remove(label)
                    print(f"Error parsing label: {label.cget('text')}")
                    continue  
            total_emission = round(total_emission, 2)  # Final rounding for display

            # Add carbon_emission from all machine parts (sum all parts, not just current page)
            for part_data in machine_parts_data:
                try:
                    carbon_emission_value = float(part_data.get("carbon_emission", "0.0"))
                    total_emission += carbon_emission_value
                except (ValueError, TypeError):
                    print(f"Error parsing part carbon emission: {part_data.get('carbon_emission', 'N/A')}")
                    continue  

            # Add packaging emission (parse and remove CO2eKg)
            try:
                packaging_text = packaging_total_emission_label.cget("text")
                packaging_emission = float(packaging_text.split(":")[1].replace("CO2eKg", "").strip())
                print(f"[DEBUG] Packaging Emission: {packaging_emission}")
            except (ValueError, AttributeError, IndexError):
                packaging_emission = 0.0
            total_emission += packaging_emission

            # Deduct recycled emission value (parse and remove CO2eKg)
            try:
                recycled_emission = float(recycled_emission_value_label.cget("text").replace("CO2eKg", "").strip())
            except (ValueError, AttributeError):
                recycled_emission = 0.0
            total_emission -= recycled_emission

            # Round the total emission to 2 decimal places for consistency
            total_emission = round(total_emission, 2)
            profile_data["total_emission"] = total_emission  # Comparison Value
            # Update the sum_emission_label
            update_pie_chart()
            sum_emission_label.config(text=f"Total Sum of Emission: {total_emission:.2f} CO2eKg")
            print(f"[DEBUG] Updated Sum of Emissions (after recycling deduction): {sum_emission_label.cget('text')}")

            # --- Threshold highlight logic ---
            highlight_color = "#006400"  # Default dark green
            if threshold_enabled_var.get():
                try:
                    threshold = float(threshold_value_var.get())
                    if total_emission > threshold:
                        highlight_color = "#FF0000"  # Red if over threshold
                    else:
                        highlight_color = "#32CD32"  # Lime green if under threshold
                except ValueError:
                    highlight_color = "#006400"  # Invalid input, keep default
            sum_emission_inner_frame.configure(bg=highlight_color)
            sum_emission_label.configure(background=highlight_color)
            threshold_label.configure(bg=highlight_color)
            threshold_checkbox.configure(bg=highlight_color, activebackground=highlight_color)
        except Exception as e:
            sum_emission_label.config(text="Total Sum of Emission: 0.0 CO2eKg")
            print(f"[DEBUG] Error in update_sum_of_emissions: {e}")

    def update_pie_chart():
        global pie_chart_debounce_id
        if pie_chart_debounce_id:
            root.after_cancel(pie_chart_debounce_id)
        pie_chart_debounce_id = root.after(500, update_pie_chart_mainthread)

    def update_pie_chart_mainthread():
        print("[DEBUG] update_pie_chart_mainthread called")
        try:
            # Check calculated_value_label exists before using
            if not (calculated_value_label and calculated_value_label.winfo_exists()):
                return
            try:
                material_emission = float(calculated_value_label.cget("text").split(":")[1].strip())
            except (IndexError, ValueError):
                material_emission = 0.0

            transport_emission = 0.0
            for label in total_emission_labels[:]:
                if label and label.winfo_exists():
                    emission_text = label.cget("text")
                    if "Total Emission:" in emission_text:
                        import re
                        emission_value = float(re.sub(r'[^0-9.\-]+', '', emission_text.split(":")[1]))
                        transport_emission += emission_value
                else:
                    total_emission_labels.remove(label)
                    continue

            machine_emission = 0.0
            for part_data in machine_parts_data:
                try:
                    carbon_emission_value = float(part_data.get("carbon_emission", "0.0"))
                    machine_emission += carbon_emission_value
                except (ValueError, TypeError):
                    continue

            # Check packaging_total_emission_label exists before using
            if packaging_total_emission_label and packaging_total_emission_label.winfo_exists():
                try:
                    packaging_text = packaging_total_emission_label.cget("text")
                    packaging_emission = float(packaging_text.split(":")[1].replace("CO2eKg", "").strip())
                except (ValueError, AttributeError, IndexError):
                    packaging_emission = 0.0
            else:
                packaging_emission = 0.0

            # Check recycled_emission_value_label exists before using
            if recycled_emission_value_label and recycled_emission_value_label.winfo_exists():
                try:
                    recycled_emission = float(recycled_emission_value_label.cget("text").replace("CO2eKg", "").strip())
                except (ValueError, AttributeError):
                    recycled_emission = 0.0
            else:
                recycled_emission = 0.0

            labels = [
                'Raw Material Emission',
                'Transport Emission',
                'Machine Emission',
                'Packaging Emission',
                'Recycled Emission'
            ]
            values = [
                material_emission,
                transport_emission,
                machine_emission,
                packaging_emission,
                recycled_emission if recycled_emission > 0 else 0
            ]
            colors = ['#FF9999', '#66B3FF', '#99FF99', '#FFB347', '#FFD700']
            
            filtered = [(label, value, color) for label, value, color in zip(labels, values, colors) if value > 0]
            print(f"[DEBUG] Pie chart data: {filtered}")
            if not filtered:
                print("[DEBUG] No data to plot in pie chart.")
                return

            if chart_type_var.get() == "Pie Chart":
                _draw_pie_chart(filtered)
                print("[DEBUG] _draw_pie_chart called")
            else:
                _draw_pareto_chart(filtered)
                print("[DEBUG] _draw_pareto_chart called")

            if root and root.winfo_exists():
                global pie_chart_debounce_id
                pie_chart_debounce_id = root.after(500, update_pie_chart_mainthread)
        except Exception as e:
            print(f"[DEBUG] Error in update_pie_chart_mainthread: {e}")

    def _draw_pie_chart(filtered):
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

        global pie_chart_canvas

        filtered_labels, filtered_values, filtered_colors = zip(*filtered)
        total = sum(filtered_values)
        explode = [0.12 if v / total < 0.08 else 0.0 for v in filtered_values]

        if pie_chart_canvas is not None and pie_chart_canvas.get_tk_widget().winfo_exists():
            pie_chart_canvas.figure.clf()
            fig = pie_chart_canvas.figure
            ax = fig.add_subplot(111)
        else:
            fig, ax = plt.subplots(figsize=(6, 3))
            pie_chart_canvas = FigureCanvasTkAgg(fig, master=pie_chart_frame)
            pie_chart_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        # Draw the pie chart
        ax.set_position([0.32, 0.08, 0.6, 0.84])
        wedges, _ = ax.pie(
            filtered_values,
            startangle=90,
            colors=filtered_colors,
            labels=None,
            explode=explode,
            autopct=None,
            pctdistance=0.8,
            textprops={'fontsize': 8, 'color': 'black', 'weight': 'bold'}
        )
        ax.axis('equal')
        ax.legend(
            wedges,
            filtered_labels,
            title="Emission Types",
            loc="lower left",
            bbox_to_anchor=(-0.35, 0),
            fontsize=7,
            title_fontsize=8,
            frameon=True
        )

        # Tooltip logic
        tooltip = ax.annotate(
            "",
            xy=(0, 0),
            xytext=(10, 10),
            textcoords="offset points",
            bbox=dict(boxstyle="round", fc="w", ec="black", lw=0.5),
            arrowprops=dict(arrowstyle="->", lw=0.5),
            ha='left',
            va='bottom',
            fontsize=9,
            weight='normal',
            color='black'
        )
        tooltip.set_visible(False)

        def on_motion(event):
            if event.inaxes != ax or event.xdata is None or event.ydata is None:
                tooltip.set_visible(False)
                pie_chart_canvas.draw_idle()
                return

            found = False
            for i, wedge in enumerate(wedges):
                if wedge.contains_point([event.x, event.y], radius=1.5):
                    tooltip.xy = (event.xdata, event.ydata)
                    tooltip.set_text(f"{filtered_labels[i]}:\n{filtered_values[i]:.2f} CO2eKg")
                    tooltip.set_visible(True)
                    found = True
                    break
            if not found:
                tooltip.set_visible(False)
            pie_chart_canvas.draw_idle()

        pie_chart_canvas.mpl_connect("motion_notify_event", on_motion)

        pie_chart_canvas.draw()
        plt.close(fig)

    def _draw_pareto_chart(filtered):
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

        global pie_chart_canvas
        if pie_chart_canvas is not None and pie_chart_canvas.get_tk_widget().winfo_exists():
            pie_chart_canvas.figure.clf()
            fig = pie_chart_canvas.figure
            ax1 = fig.add_subplot(111)
        else:
            fig, ax1 = plt.subplots(figsize=(6, 3))
            pie_chart_canvas = FigureCanvasTkAgg(fig, master=pie_chart_frame)
            pie_chart_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        labels, values, colors = zip(*filtered)
        sorted_data = sorted(zip(values, labels, colors), reverse=True)
        sorted_values, sorted_labels, sorted_colors = zip(*sorted_data)
        cum_values = [sum(sorted_values[:i+1]) for i in range(len(sorted_values))]
        total = sum(sorted_values)
        cum_perc = [v / total * 100 for v in cum_values]

        # Use integer positions for bars and ticks
        x = range(len(sorted_labels))
        bars = ax1.bar(x, sorted_values, color=sorted_colors)
        ax1.set_ylabel('Emission (CO2eKg)', fontsize=9)
        ax1.set_xticks(x)
        ax1.set_xticklabels(sorted_labels, rotation=30, ha='right', fontsize=7)
        fig.subplots_adjust(bottom=0.28)  # Increase bottom margin for label visibility

        # Secondary axis for cumulative percentage
        ax2 = ax1.twinx()
        ax2.plot(x, cum_perc, color='red', marker='o', linestyle='-', linewidth=2)
        ax2.set_ylabel('Cumulative %', fontsize=9)
        ax2.set_ylim(0, 110)
        ax2.grid(False)

        # Annotate bar values
        for i, v in enumerate(sorted_values):
            ax1.text(i, v, f"{v:.1f}", ha='center', va='bottom', fontsize=8)

        pie_chart_canvas.draw()
        plt.close(fig)

    debounce_weight_id = None
    def debounce_weight_calculation():
        global debounce_weight_id
        if debounce_weight_id:
            root.after_cancel(debounce_weight_id)
        print("[DEBUG] Starting debounce timer for weight_entry")
        debounce_weight_id = root.after(500, lambda: [
            calculate_material_value(),
            recalculate_all_emissions(),
            recalculate_all_carbon_emissions(),
            update_sum_of_emissions()
        ])

    debounce_calc_ids = {}

    def generate_transport_rows(quantity, total_emission_label=None):
        global transport_row_widgets
        print(f"[DEBUG] Entered generate_transport_rows with quantity={quantity}")
        print(f"Generating {quantity} transport rows")
        transport_row_widgets.clear()

        try:
            inner_box_frame.winfo_exists()  # This will raise TclError if the widget doesn't exist
        except (tk.TclError, NameError):
            print("[DEBUG] inner_box_frame no longer exists, cannot generate transport rows")
            return

        if quantity == 0:
            for child in inner_box_frame.winfo_children():
                child.destroy()
            return

        # Get all existing transport rows inside the inner_box_frame
        existing_rows = [
            child
            for child in inner_box_frame.winfo_children()
            if isinstance(child, tk.Frame) and "boundary_frame" in str(child)
        ]

        # Remove excess rows if quantity is decreased
        if len(existing_rows) > quantity:
            for i in range(quantity, len(existing_rows)):
                boundary_frame = existing_rows[i]
                boundary_frame.destroy()  
            del transport_row_widgets[quantity:]  

        # Create new rows if quantity exceeds current rows
        for i in range(len(existing_rows), quantity):
            row = i  

            boundary_frame = tk.Frame(inner_box_frame, bg="#A8D5BA", padx=2, pady=2, name=f"boundary_frame_{i}")
            boundary_frame.grid(row=row, column=0, columnspan=2, sticky="ew", padx=5, pady=5)

            # Bind mouse wheel events for each transport row
            boundary_frame.bind("<MouseWheel>", forward_mouse_wheel_to_canvas)  # For Windows/macOS
            boundary_frame.bind("<Button-4>", lambda event: canvas.yview_scroll(-1, "units"))  # For Linux (scroll up)
            boundary_frame.bind("<Button-5>", lambda event: canvas.yview_scroll(1, "units"))   # For Linux (scroll down)

            # Create the row frame inside the boundary frame
            row_frame = ttk.Frame(boundary_frame, style="Eco.TFrame")
            row_frame.grid(row=0, column=0, sticky="nsew")

            # Configure the grid weights for the row frame
            row_frame.grid_columnconfigure(0, weight=2)  
            row_frame.grid_columnconfigure(1, weight=2) 

            # Left column: Transport, Variation, and Fuel frames
            left_column_frame = ttk.Frame(row_frame, style="Eco.TFrame")
            left_column_frame.grid(row=0, column=0, sticky="nsew", padx=5, pady=2)

            # Transport List
            transport_frame = ttk.Labelframe(left_column_frame, text=f"Choice of Transport {i+1}", style="Eco.TLabelframe")
            transport_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=2)
            transport_combo = ttk.Combobox(transport_frame, width=25, state="readonly")
            transport_combo.grid(row=0, column=0, sticky="ew")
            transport_combo["values"] = transport_list

            # Bind mouse wheel events for transport_frame
            transport_frame.bind("<MouseWheel>", forward_mouse_wheel_to_canvas)  
            transport_frame.bind("<Button-4>", lambda event: canvas.yview_scroll(-1, "units"))  
            transport_frame.bind("<Button-5>", lambda event: canvas.yview_scroll(1, "units"))   

            # Variations
            variation_frame = ttk.Labelframe(left_column_frame, text=f"Different Types {i+1}", style="Eco.TLabelframe")
            variation_frame.grid(row=1, column=0, sticky="ew", padx=5, pady=2)
            variation_combo = ttk.Combobox(variation_frame, width=25, state="readonly")
            variation_combo.grid(row=0, column=0, sticky="ew")

            # Bind mouse wheel events for variation_frame
            variation_frame.bind("<MouseWheel>", forward_mouse_wheel_to_canvas)  
            variation_frame.bind("<Button-4>", lambda event: canvas.yview_scroll(-1, "units"))  
            variation_frame.bind("<Button-5>", lambda event: canvas.yview_scroll(1, "units"))   

            # Fuel Types
            fuel_frame = ttk.Labelframe(left_column_frame, text=f"Fuel Types {i+1}", style="Eco.TLabelframe")
            fuel_frame.grid(row=2, column=0, sticky="ew", padx=5, pady=2)
            fuel_combo = ttk.Combobox(fuel_frame, width=25, state="readonly")
            fuel_combo.grid(row=0, column=0, sticky="ew")

            # Bind mouse wheel events for fuel_frame
            fuel_frame.bind("<MouseWheel>", forward_mouse_wheel_to_canvas)  
            fuel_frame.bind("<Button-4>", lambda event: canvas.yview_scroll(-1, "units"))  
            fuel_frame.bind("<Button-5>", lambda event: canvas.yview_scroll(1, "units"))   

            # Carbon Emission Output
            value_frame = ttk.Labelframe(left_column_frame, text=f"Carbon Emission Amount {i+1}", style="Eco.TLabelframe")
            value_frame.grid(row=3, column=0, sticky="ew", padx=5, pady=2)
            value_label = ttk.Label(value_frame, text="0.0", style="Eco.TLabel")  
            value_label.grid(row=0, column=0, sticky="ew")

            # Bind mouse wheel events for value_frame
            value_frame.bind("<MouseWheel>", forward_mouse_wheel_to_canvas) 
            value_frame.bind("<Button-4>", lambda event: canvas.yview_scroll(-1, "units"))  
            value_frame.bind("<Button-5>", lambda event: canvas.yview_scroll(1, "units"))   

            # Right column: Distance Calculator
            distance_frame = ttk.Labelframe(row_frame, text=f"Distance Calculator {i+1}", style="Eco.TLabelframe")
            distance_frame.grid(row=0, column=1, sticky="nsew", padx=5, pady=2)

            # Bind mouse wheel events for distance_frame
            distance_frame.bind("<MouseWheel>", forward_mouse_wheel_to_canvas)  
            distance_frame.bind("<Button-4>", lambda event: canvas.yview_scroll(-1, "units"))  
            distance_frame.bind("<Button-5>", lambda event: canvas.yview_scroll(1, "units"))   

            tk.Label(distance_frame, text="Origin:").grid(row=0, column=0, padx=2, pady=1, sticky="w")
            origin_entry = tk.Entry(distance_frame, width=20)
            origin_entry.grid(row=0, column=1, padx=2, pady=1)

            tk.Label(distance_frame, text="Destination:").grid(row=1, column=0, padx=2, pady=1, sticky="w")
            destination_entry = tk.Entry(distance_frame, width=20)
            destination_entry.grid(row=1, column=1, padx=2, pady=1)

            tk.Label(distance_frame, text="Mode:").grid(row=2, column=0, padx=2, pady=1, sticky="w")
            mode_var = tk.StringVar(value="None")
            mode_options = ["None", "driving", "plane", "ship"]
            mode_menu = ttk.OptionMenu(distance_frame, mode_var, *mode_options)
            mode_menu.grid(row=2, column=1, padx=2, pady=1, sticky="w")

            result_label = ttk.Label(distance_frame, text="Distance: 0.0 km", style="Eco.TLabel")
            result_label.grid(row=3, column=0, columnspan=2, pady=2)

            # Attach mode_menu to transport_combo for dynamic updates
            transport_combo.mode_menu = mode_menu
            transport_combo.mode_var = mode_var

            # Total Emission Output
            total_emission_frame = ttk.Labelframe(distance_frame, text=f"Total Emission {i+1}", style="Eco.TLabelframe")
            total_emission_frame.grid(row=4, column=0, columnspan=2, sticky="nsew", padx=5, pady=2)
            total_emission_label = ttk.Label(total_emission_frame, text="Total Emission: 0.0 CO2eKg", style="Eco.TLabel")
            total_emission_label.grid(row=0, column=0, sticky="ew")

            # Bind mouse wheel events for total_emission_frame
            total_emission_frame.bind("<MouseWheel>", forward_mouse_wheel_to_canvas)  
            total_emission_frame.bind("<Button-4>", lambda event: canvas.yview_scroll(-1, "units"))  
            total_emission_frame.bind("<Button-5>", lambda event: canvas.yview_scroll(1, "units"))  

            total_emission_labels.append(total_emission_label) 
            total_emission_label.bind("<Configure>", lambda event: update_sum_of_emissions())  # Bind the event

            transport_combo.variation_combo = variation_combo
            transport_combo.fuel_combo = fuel_combo
            transport_combo.value_label = value_label
            transport_combo.value_frame = value_frame

            # Bind events to dynamically calculate distance and total emission
            origin_entry.bind("<FocusOut>", lambda event: [
                calculate_distance_dynamic(origin_entry, destination_entry, mode_var, result_label),
                auto_save_inputs()
            ])
            origin_entry.bind("<Return>", lambda event: [
                calculate_distance_dynamic(origin_entry, destination_entry, mode_var, result_label),
                auto_save_inputs()
            ])
            destination_entry.bind("<FocusOut>", lambda event: [
                calculate_distance_dynamic(origin_entry, destination_entry, mode_var, result_label),
                auto_save_inputs()
            ])
            destination_entry.bind("<Return>", lambda event: [
                calculate_distance_dynamic(origin_entry, destination_entry, mode_var, result_label),
                auto_save_inputs()
            ])

            destination_entry.bind("<FocusOut>", lambda event: auto_save_inputs())
            mode_var.trace("w", lambda *args, o=origin_entry, d=destination_entry, m=mode_var, r=result_label: calculate_distance_dynamic(o, d, m, r))
            result_label.bind("<Configure>", lambda event, v=value_label, r=result_label, t=total_emission_label, w=weight_entry, tc=transport_combo: [calculate_total_emission(v, r, t, w, tc), update_sum_of_emissions()])
            total_emission_label.bind("<Configure>", lambda event: update_sum_of_emissions())

            # Bind events to dynamically update and recalculate total emission
            # Transport Combo
            def on_transport_selected(event, tc=transport_combo, r=result_label, t=total_emission_label):
                if restoring_profile:
                    return
                try:
                    update_transport_types(event)
                    auto_save_inputs()
                    calculate_total_emission(tc.value_label, r, t, weight_entry, tc)
                    update_sum_of_emissions()
                except Exception as e:
                    print(f"[ERROR] Exception in on_transport_selected: {e}")
            transport_combo.bind("<<ComboboxSelected>>", lambda event, tc=transport_combo, r=result_label, t=total_emission_label: on_transport_selected(event, tc, r, t))

            # Variation Combo
            def on_variation_selected(event, tc=transport_combo, r=result_label, t=total_emission_label):
                if restoring_profile:
                    return
                try:
                    print(f"[DEBUG] Variation Combo Changed: {tc.variation_combo.get()}")
                    update_value(event, tc)
                    calculate_total_emission(tc.value_label, r, t, weight_entry, tc)
                    update_sum_of_emissions()
                except Exception as e:
                    print(f"[ERROR] Exception in on_variation_selected: {e}")
            variation_combo.bind("<<ComboboxSelected>>", lambda event, tc=transport_combo, r=result_label, t=total_emission_label: on_variation_selected(event, tc, r, t))

            # Fuel Combo
            def on_fuel_selected(event, tc=transport_combo, r=result_label, t=total_emission_label):
                if restoring_profile:
                    return
                try:
                    print(f"[DEBUG] Fuel Combo Changed: {tc.fuel_combo.get()}")
                    update_value(event, tc)
                    calculate_total_emission(tc.value_label, r, t, weight_entry, tc)
                    update_sum_of_emissions()
                except Exception as e:
                    print(f"[ERROR] Exception in on_fuel_selected: {e}")
            fuel_combo.bind("<<ComboboxSelected>>", lambda event, tc=transport_combo, r=result_label, t=total_emission_label: on_fuel_selected(event, tc, r, t))

            # Append the row widgets to the global list
            transport_row_widgets.append({
                "transport_combo": transport_combo,
                "variation_combo": variation_combo,
                "fuel_combo": fuel_combo,
                "origin_entry": origin_entry,
                "destination_entry": destination_entry,
                "result_label": result_label,
                "total_emission_label": total_emission_label,
                "weight_entry": weight_entry,
                "value_label": value_label,
            })
            
        print(f"[DEBUG] Number of transport rows after generation: {len(transport_row_widgets)}")

        canvas.yview_moveto(0) 
        outlined_canvas.update_idletasks()
        outlined_canvas.configure(scrollregion=outlined_canvas.bbox("all"))

    transport_rows_state = {}

    def save_transport_rows_state():
        global transport_rows_state
        transport_rows_state.clear()  
        for i, child in enumerate(root.winfo_children()):
            if isinstance(child, ttk.Labelframe) and f"Choice of Transport {i+1}" in child.cget("text"):
                transport_combo = child.winfo_children()[0]
                variation_combo = transport_combo.variation_combo
                fuel_combo = transport_combo.fuel_combo
                value_label = transport_combo.value_label

                # Save the current state of the row
                transport_rows_state[i] = {
                    "transport": transport_combo.get(),
                    "variation": variation_combo.get(),
                    "fuel": fuel_combo.get(),
                    "value": value_label.cget("text")
                }

    def restore_transport_rows_state():
           for i, child in enumerate(root.winfo_children()):
               if isinstance(child, ttk.Labelframe) and f"Choice of Transport {i+1}" in child.cget("text"):
                   transport_combo = child.winfo_children()[0]
                   variation_combo = transport_combo.variation_combo
                   fuel_combo = transport_combo.fuel_combo
                   value_label = transport_combo.value_label

                   # Restore the saved state of the row
                   if i in transport_rows_state:
                       state = transport_rows_state[i]
                       transport_combo.set(state["transport"])
                       variation_combo.set(state["variation"])
                       fuel_combo.set(state["fuel"])
                       value_label.config(text=state["value"])

           # Rebind <Configure> for all total_emission_labels
           for label in total_emission_labels[:]:
                if label and label.winfo_exists():
                    label.bind("<Configure>", lambda event: update_sum_of_emissions())
                else:
                    total_emission_labels.remove(label)
           print("Valid total_emission_labels:", [label for label in total_emission_labels if label.winfo_exists()])
   
    def clean_up_transport_rows_state(quantity):
           keys_to_remove = [key for key in transport_rows_state if key >= quantity]
           for key in keys_to_remove:
               del transport_rows_state[key]

           # Remove invalid labels
           global total_emission_labels
           total_emission_labels = [label for label in total_emission_labels if label.winfo_exists()]  
   
    def on_quantity_submit():
        try:
            quantity = int(quantity_entry.get())
            print("Quantity Submitted:", quantity)  
            generate_transport_rows(quantity)  # Regenerate transport rows
        except ValueError:
            messagebox.showerror("Invalid Input", "Please enter a valid integer for quantity.")

    def on_quantity_submit():
        quantity = int(quantity_entry.get())
        generate_transport_rows(quantity)

    def on_quantity_change():
        global restoring_profile
        if restoring_profile:
            return
        print("[DEBUG] on_quantity_change triggered")
        try:
            quantity_str = quantity_entry.get().strip()
            if not quantity_str:
                quantity = 0
            else:
                quantity = int(quantity_str)
            if quantity < 0:
                quantity = 0

            print("Quantity Changed:", quantity)

            # Only regenerate if quantity actually changed
            if quantity != len(transport_row_widgets):
                if quantity > 0:
                    outlined_box_frame.grid()
                    inner_box_frame.pack(fill="both", expand=True)
                    generate_transport_rows(quantity, None)
                    outlined_canvas.update_idletasks()
                    outlined_canvas.configure(scrollregion=outlined_canvas.bbox("all"))

                    # --- Dynamically enable/disable scrolling ---
                    bbox = outlined_canvas.bbox("all")
                    height = bbox[3] - bbox[1] if bbox else 0
                    visible_height = outlined_canvas.winfo_height()
                    if height > visible_height:
                        scrollbar.configure(command=canvas.yview)
                        canvas.bind("<MouseWheel>", on_mouse_wheel)
                        canvas.bind("<Button-4>", on_mouse_wheel)
                        canvas.bind("<Button-5>", on_mouse_wheel)
                    else:
                        scrollbar.configure(command=None)
                        canvas.unbind("<MouseWheel>")
                        canvas.unbind("<Button-4>")
                        canvas.unbind("<Button-5>")
                else:
                    outlined_box_frame.grid_remove()
                    inner_box_frame.pack_forget()
                    outlined_canvas.configure(scrollregion=(0, 0, 0, 0))
                    scrollbar.configure(command=None)
                    canvas.unbind("<MouseWheel>")
                    canvas.unbind("<Button-4>")
                    canvas.unbind("<Button-5>")

                clean_up_transport_rows_state(quantity)
        except ValueError:
            print("Invalid quantity input")
            outlined_box_frame.grid_remove()
            inner_box_frame.pack_forget()
            outlined_canvas.configure(scrollregion=(0, 0, 0, 0))
            scrollbar.configure(command=None)
            canvas.unbind("<MouseWheel>")
            canvas.unbind("<Button-4>")
            canvas.unbind("<Button-5>")

    sea_route_cache = {}
    def get_cached_sea_route(origin, destination):
        key = (origin, destination)
        if key in sea_route_cache:
            return sea_route_cache[key]
        result = visualize_sea_route(origin, destination)
        if result:
            sea_route_cache[key] = result
        return result

    # Function to visualize the sea route
    def visualize_sea_route(origin_name, destination_name):
        if not origin_name or not destination_name:
            print("Origin or destination is empty.")
            return None, None, None

        # Get coordinates for origin and destination
        origin = get_coordinates(origin_name)
        destination = get_coordinates(destination_name)

        if not origin:
            print(f"Could not resolve coordinates for origin: {origin_name}")
            return None, None, None
        if not destination:
            print(f"Could not resolve coordinates for destination: {destination_name}")
            return None, None, None

        try:
            # Get the route using searoute
            route = sr.searoute(origin, destination, units="naut")
            if not route or 'properties' not in route or 'geometry' not in route:
                print("Invalid route data received from searoute.")
                return None, None, None

            distance_nautical = route['properties']['length']
            distance_km = distance_nautical * 1.852  # Convert to kilometers

            # Extract route coordinates
            route_coordinates = route['geometry']['coordinates']

            # Plot the map
            fig = plt.figure(figsize=(10, 6))
            ax = plt.axes(projection=ccrs.PlateCarree())
            ax.set_global()

            # Add map features
            ax.add_feature(cfeature.LAND, edgecolor='black')
            ax.add_feature(cfeature.COASTLINE)
            ax.add_feature(cfeature.BORDERS, linestyle=':')

            # Plot the route
            lons, lats = zip(*route_coordinates)
            ax.plot(lons, lats, color='blue', linewidth=2, marker='o', transform=ccrs.Geodetic(), label="Sea Route")

            # Plot origin and destination
            ax.plot(origin[0], origin[1], 'ro', transform=ccrs.Geodetic(), label="Origin")
            ax.plot(destination[0], destination[1], 'go', transform=ccrs.Geodetic(), label="Destination")

            # Add legend and title
            plt.legend()
            plt.title(f"Sea Route: {distance_nautical:.1f} Nautical Miles / {distance_km:.1f} Kilometers")

            # Return the figure and distances
            return fig, distance_nautical, distance_km

        except Exception as e:
            print(f"Error calculating sea route: {e}")
            return None, None, None

    # Configure the grid layout for the root window
    root.grid_columnconfigure(0, weight=3)  
    root.grid_columnconfigure(1, weight=1)  
    root.grid_rowconfigure(0, weight=1)     

    # Create a frame for the "SP" text and additional text at the top
    header_frame = tk.Frame(root, bg="#A8D5BA", height=100)  # Set a fixed height for the header
    header_frame.pack(side="top", fill="x")  

    export_button = tk.Button(
    header_frame,
    text="Export Excel Data",
    command=export_profile_to_excel,
    font=("Arial", 10),
    bg="#4CAF50",
    fg="white"
    )
    export_button.pack(side="right", padx=5, pady=10)

    excel_template_button = tk.Button(
        header_frame,
        text="Create Excel Template",
        command=create_excel_template,
        font=("Arial", 10),
        bg="#4CAF50",
        fg="white"
    )
    excel_template_button.pack(side="right", padx=5, pady=10)

    import_excel_button = tk.Button(
        header_frame,
        text="Import Excel Data",
        command=import_from_excel,
        font=("Arial", 10),
        bg="#2196F3",
        fg="white"
    )
    import_excel_button.pack(side="right", padx=5, pady=10)

    pdf_button = tk.Button(
        header_frame,
        text="Export PDF Report",
        command=export_profile_to_pdf,
        font=("Arial", 10),
        bg="#795548",
        fg="white"
    )
    pdf_button.pack(side="right", padx=5, pady=10)

    def show_emission_breakdown():
        breakdown_win = tk.Toplevel(root)
        breakdown_win.title("Emission Breakdown")
        tree = ttk.Treeview(breakdown_win, columns=("Category", "Value"), show="headings")
        tree.heading("Category", text="Category")
        tree.heading("Value", text="Emission (CO2eKg)")
        tree.pack(fill="both", expand=True)

        # Get the values as you do in update_pie_chart_mainthread
        try:
            material_emission = float(calculated_value_label.cget("text").split(":")[1].strip())
        except (IndexError, ValueError):
            material_emission = 0.0

        # Sum all transport emissions as floats, then round the total
        import re
        transport_emission = round(
            sum(
                float(re.sub(r'[^0-9.\-]+', '', label.cget("text").split(":")[1]))
                for label in total_emission_labels[:]
                if label and label.winfo_exists() and "Total Emission:" in label.cget("text")
            ),
            2
        )

        machine_emission = 0.0
        for part_data in machine_parts_data:
            try:
                carbon_emission_value = float(part_data.get("carbon_emission", "0.0"))
                machine_emission += carbon_emission_value
            except (ValueError, TypeError):
                continue

        if packaging_total_emission_label and packaging_total_emission_label.winfo_exists():
            try:
                packaging_text = packaging_total_emission_label.cget("text")
                packaging_emission = float(packaging_text.split(":")[1].replace("CO2eKg", "").strip())
            except (ValueError, AttributeError, IndexError):
                packaging_emission = 0.0
        else:
            packaging_emission = 0.0

        if recycled_emission_value_label and recycled_emission_value_label.winfo_exists():
            try:
                recycled_emission = float(recycled_emission_value_label.cget("text").replace("CO2eKg", "").strip())
            except (ValueError, AttributeError):
                recycled_emission = 0.0
        else:
            recycled_emission = 0.0

        # Insert the actual values
        tree.insert("", "end", values=("Raw Material", f"{material_emission:.2f}"))
        tree.insert("", "end", values=("Transport", f"{transport_emission:.2f}"))
        tree.insert("", "end", values=("Machine", f"{machine_emission:.2f}"))
        tree.insert("", "end", values=("Packaging", f"{packaging_emission:.2f}"))
        tree.insert("", "end", values=("Recycled", f"-{recycled_emission:.2f}"))

    breakdown_button = tk.Button(
    header_frame,
    text="Breakdown",
    command=show_emission_breakdown,
    font=("Arial", 10),
    bg="#FFB300",
    fg="white"
    )
    breakdown_button.pack(side="right", padx=5, pady=10)

    edit_emission_button = tk.Button(
        header_frame,
        text="Advanced Customization",
        command=lambda: show_custom_emission_factors_dialog(
            root, materials, on_update=recalculate_all
        ),
        font=("Arial", 10),
        bg="#9C27B0",
        fg="white"
    )
    edit_emission_button.pack(side="right", padx=(5, 0), pady=10)

    info_icon_edit_emission = tk.Label(
        header_frame,
        text="ℹ️",
        bg="#A8D5BA", 
        fg="black",
        font=("Arial", 12)
    )
    info_icon_edit_emission.pack(side="right", padx=(0, 5), pady=10)

    ToolTip(
        info_icon_edit_emission,
        "Advanced Customization: For advanced emission factors customization."
        "\nBreakdown: View a detailed breakdown of emissions by category."
        "\nExport PDF Report: Generate a comprehensive PDF report with all emission data, charts, and optimization suggestions."
        "\nImport Excel Data: Import emission data from an Excel file using the SPHERE template format."
        "\nCreate Excel Template: Creates a structured Excel workbook to fill in or sharing to others. "
        "\nExport Excel Data: Export all current emission data to an Excel file for external analysis or sharing."
    )

    # Create a sub-frame for the "SP" text
    sp_frame = tk.Frame(header_frame, bg="#A8D5BA")
    sp_frame.pack(side="left", padx=5, pady=10) 

    # Add the "SP" text to the sp_frame
    sp_label = tk.Label(
        sp_frame,
        text="SP",
        font=("Ariel", 75, "bold"),
        fg="#CC0000",
        bg="#A8D5BA"
    )
    sp_label.pack()

    # Add a shadow effect for the "SP" text
    shadow_label = tk.Label(
        sp_frame,
        text="SP",
        font=("Ariel", 75, "bold"),
        fg="#FF2424",
        bg="#A8D5BA"
    )
    shadow_label.place(x=2, y=2) 

    # Create a sub-frame for "Singapore" and "Polytechnic" text
    text_frame = tk.Frame(header_frame, bg="#A8D5BA")
    text_frame.pack(side="left", padx=0, pady=10) 

    # Add the "Singapore" text to the text_frame
    singapore_label = tk.Label(
        text_frame,
        text="Singapore",
        font=("Ariel", 23), 
        fg="black",
        bg="#A8D5BA"
    )
    singapore_label.pack(anchor="w")  # Align to the left

    # Add the "Polytechnic" text below "Singapore"
    polytechnic_label = tk.Label(
        text_frame,
        text="Polytechnic",
        font=("Ariel", 23),  
        fg="black",
        bg="#A8D5BA"
    )
    polytechnic_label.pack(anchor="w", pady=(0,0)) 

    # Create a canvas and a scrollbar
    canvas = tk.Canvas(root, bg="#A8D5BA", highlightthickness=0)  # Set eco-design background for the canvas
    scrollbar = ttk.Scrollbar(root, orient="vertical", command=canvas.yview)
    scrollable_frame = ttk.Frame(canvas, style="Eco.TFrame")  # Use a custom style for the frame

    # Configure the canvas and scrollbar
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    # Function to forward mouse wheel events from scrollable_frame and its children to the canvas
    def forward_mouse_wheel_to_canvas(event):
        canvas.yview_scroll(-1 * (event.delta // 120), "units")  
        return "break"  

    # Bind mouse wheel events for scrollable_frame
    scrollable_frame.bind("<MouseWheel>", forward_mouse_wheel_to_canvas)  # For Windows/macOS
    scrollable_frame.bind("<Button-4>", lambda event: canvas.yview_scroll(-1, "units"))  # For Linux (scroll up)
    scrollable_frame.bind("<Button-5>", lambda event: canvas.yview_scroll(1, "units"))   # For Linux (scroll down)

    # Bind mouse wheel events for all child widgets of scrollable_frame
    for child in scrollable_frame.winfo_children():
        child.bind("<MouseWheel>", forward_mouse_wheel_to_canvas) 
        child.bind("<Button-4>", lambda event: canvas.yview_scroll(-1, "units"))  
        child.bind("<Button-5>", lambda event: canvas.yview_scroll(1, "units"))   

    # Function to update the scroll region dynamically
    def update_scroll_region(event):
        canvas.configure(scrollregion=canvas.bbox("all"))

    # Bind the <Configure> event to update the scroll region
    scrollable_frame.bind("<Configure>", update_scroll_region)

    # Place the canvas and scrollbar in the left column
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    # Add a red box frame beside the left side of right_frame
    recycling_box_frame = tk.Frame(
        root,
        bg="#A8D5BA",
        width=450,
        height=300,  # Reduced height
        highlightbackground="black",
        highlightthickness=2
    )
    recycling_box_frame.pack(side="left", fill="y", padx=(10, 0), pady=10)
    recycling_box_frame.pack_propagate(False)

    # Add a header inside recycling_box_frame
    recycling_header = tk.Label(
        recycling_box_frame,
        text="Recycling",
        bg="#A8D5BA",
        fg="black",
        font=("Arial", 16, "bold")
    )
    recycling_header.pack(side="top", pady=5)

    # Create a recycling frame with a background that blends with "#A8D5BA"
    recycling_inner_frame = tk.Frame(recycling_box_frame, bg="#A8D5BA")
    recycling_inner_frame.pack(fill="x", padx=10, pady=10)

    # --- Add a separator line before the packaging section ---
    style = ttk.Style()
    style.configure("Black.TSeparator", background="black")
    separator = ttk.Separator(recycling_box_frame, orient="horizontal", style="Black.TSeparator")
    separator.pack(fill="x", padx=0, pady=(0, 5))
    
    # --- Packaging Section: Add this below the separator ---
    packaging_box_frame = tk.Frame(recycling_box_frame, bg="#A8D5BA")
    packaging_box_frame.pack(fill="x", padx=10, pady=(0, 10))

    # Optional: Add a header or placeholder for future packaging inputs
    packaging_box_frame.grid_columnconfigure(0, weight=1)
    packaging_box_frame.grid_columnconfigure(1, weight=1)
    packaging_box_frame.grid_columnconfigure(2, weight=1)

    packaging_header = tk.Label(
        packaging_box_frame,
        text="Packaging",
        bg="#A8D5BA",
        fg="black",
        font=("Arial", 12, "bold")
    )
    packaging_header.grid(row=0, column=0, columnspan=3, sticky="ew", pady=(5, 2))

    select_packaging_label = tk.Label(
        packaging_box_frame,
        text="Packaging type:",
        bg="#A8D5BA",
        fg="black",
        font=("Arial", 10, "bold")
    )
    select_packaging_label.grid(row=2, column=0, sticky="w", padx=5, pady=5)

    packaging_dropdown = ttk.Combobox(packaging_box_frame, width=23, state="readonly")
    packaging_dropdown["values"] = packaging_types_list
    packaging_dropdown.set("Select Packaging type")
    packaging_dropdown.grid(row=2, column=1, sticky="e", padx=5, pady=5)
    
    packaging_emission_value_label = tk.Label(
        packaging_box_frame,
        text="Emission: N/A",
        bg="#A8D5BA",
        fg="black",
        font=("Arial", 10)
    )
    packaging_emission_value_label.grid(row=2, column=2, sticky="w", padx=5, pady=5)

    packaging_total_emission_label = tk.Label(
        packaging_box_frame,
        text="Packaging Emission: 0.00CO2eKg",
        bg="#A8D5BA",
        fg="black",
        font=("Arial", 10, "bold")
    )
    packaging_total_emission_label.grid(row=3, column=0, columnspan=3, sticky="w", padx=5, pady=5)

    # Callback to update the emission value based on the selected packaging type
    def update_packaging_emission_value(event=None):
        selected = packaging_dropdown.get().strip().lower()
        normalized_types = [t.strip().lower() for t in packaging_types_list]
        print(f"[DEBUG] Selected: '{selected}'")
        print(f"[DEBUG] Normalized types: {normalized_types}")
        try:
            index = normalized_types.index(selected)
            emission_value = packaging_box_frame_list[index]
            print(f"[DEBUG] Index: {index}, Emission Value: {emission_value}")
            if emission_value != "N/A":
                emission_str = f"{float(emission_value):.2f}"
            else:
                emission_str = "N/A"
        except (ValueError, TypeError) as e:
            print(f"[DEBUG] Exception: {e}")
            emission_str = "N/A"
        packaging_emission_value_label.config(
            text=f"Emission: {emission_str}CO2eKg",
            font=("Arial", 8)
        )

    def update_packaging_emission_total(*args):
        selected = packaging_dropdown.get().strip().lower()
        normalized_types = [t.strip().lower() for t in packaging_types_list]
        print(f"[DEBUG] [update_packaging_emission_total] Selected: '{selected}'")
        print(f"[DEBUG] [update_packaging_emission_total] Normalized types: {normalized_types}")
        try:
            index = normalized_types.index(selected)
            emission_value = packaging_box_frame_list[index]
            print(f"[DEBUG] [update_packaging_emission_total] Index: {index}, Emission Value: {emission_value}")
            emission_value_float = float(emission_value) if emission_value != "N/A" else 0.0
        except (ValueError, TypeError) as e:
            print(f"[DEBUG] [update_packaging_emission_total] Exception: {e}")
            emission_value_float = 0.0
        try:
            weight_after = get_weight_after_total()
            print(f"[DEBUG] [update_packaging_emission_total] Weight after: {weight_after}")
            total_grams = emission_value_float * weight_after
            total_kg = total_grams / 1000.0
            print(f"[DEBUG] [update_packaging_emission_total] total_grams: {total_grams}, total_kg: {total_kg}")
            packaging_total_emission_label.config(
                text=f"Packaging Emission: {total_kg:.2f}CO2eKg"
            )
        except Exception as e:
            print(f"[DEBUG] [update_packaging_emission_total] Exception in calculation: {e}")
            packaging_total_emission_label.config(
                text="Packaging Emission: 0.00CO2eKg"
            )
        update_sum_of_emissions()

    def on_packaging_type_change(event=None):
        print("[DEBUG] on_packaging_type_change called")
        update_packaging_emission_value(event)
        update_packaging_emission_total()

    # Bind the callback to the dropdown
    def on_packaging_selected(event):
        if restoring_profile:
            return
        try:
            on_packaging_type_change(event)
            auto_save_inputs()
        except Exception as e:
            print(f"[ERROR] Exception in on_packaging_selected: {e}")
    packaging_dropdown.bind("<<ComboboxSelected>>", on_packaging_selected)

    # Column 0: "Select Metal type:" label in black
    select_metal_label = tk.Label(
        recycling_inner_frame,
        text="Recycle Metal type:",
        bg="#A8D5BA",   # Background to blend in
        fg="black",     # Black text color
        font=("Arial", 10, "bold")
    )
    select_metal_label.grid(row=0, column=0, sticky="w", padx=5, pady=5)

    # Callback to update the emission value based on the selected metal type
    def update_metal_recycling_value(event):
        print("[DEBUG] update_metal_recycling_value called, value:", metal_recycling_dropdown.get())
        selected = metal_recycling_dropdown.get().strip().lower()
        normalized_types = [t.strip().lower() for t in metal_recyling_types_list]
        try:
            index = normalized_types.index(selected)
            emission_value = metal_recyling_emission_list[index]
            if emission_value != "N/A":
                emission_str = f"{float(emission_value):.2f}"
            else:
                emission_str = "N/A"
        except ValueError:
            emission_str = "N/A"
        metal_recycling_value_label.config(
            text=f"Emission: {emission_str}CO2eKg",
            font=("Arial", 8)
        )
        # Always update the calculation after the label is set
        update_recycled_emission_value()

    # Column 1: Metal recycling dropdown (using ttk)
    metal_recycling_dropdown = ttk.Combobox(recycling_inner_frame, width=18, state="readonly")
    metal_recycling_dropdown["values"] = metal_recyling_types_list
    metal_recycling_dropdown.set("Select Metal type")
    metal_recycling_dropdown.grid(row=0, column=1, sticky="e", padx=0, pady=5)

    def on_metal_recycling_selected(event):
        if restoring_profile:
            return
        try:
            update_metal_recycling_value(event)
            auto_save_inputs()
        except Exception as e:
            print(f"[ERROR] Exception in on_metal_recycling_selected: {e}")
    metal_recycling_dropdown.bind("<<ComboboxSelected>>", on_metal_recycling_selected)

    # Column 2: Emission output label in black
    metal_recycling_value_label = tk.Label(
        recycling_inner_frame,
        text="Emission: N/A",
        bg="#A8D5BA",  
        fg="black",
        font=("Arial", 10)
    )
    metal_recycling_value_label.grid(row=0, column=2, sticky="w", padx=5, pady=5)

    def update_recycled_emission_value(*args):
        print("[DEBUG] update_recycled_emission_value called")
        # Always update the emission label based on the dropdown
        selected = metal_recycling_dropdown.get().strip().lower()
        normalized_types = [t.strip().lower() for t in metal_recyling_types_list]
        try:
            index = normalized_types.index(selected)
            emission_value = metal_recyling_emission_list[index]
            if emission_value != "N/A":
                emission_str = f"{float(emission_value):.2f}"
            else:
                emission_str = "N/A"
        except ValueError:
            emission_str = "N/A"
        metal_recycling_value_label.config(
            text=f"Emission: {emission_str}CO2eKg",
            font=("Arial", 8)
        )
        try:
            selected_country = manufacturing_country_combo.get()
            if selected_country in country_list:
                country_index = country_list.index(selected_country)
                manufacturing_electricity = float(electricity_list[country_index])
            else:
                manufacturing_electricity = 0.0

            # Use the just-updated emission_str for calculation
            try:
                metal_emission = float(emission_str) if emission_str != "N/A" else 0.0
            except Exception:
                metal_emission = 0.0

            recycled_weight = float(recycled_weight_value_label.cget("text"))

            recycled_emission = (manufacturing_electricity / 0.3) * metal_emission * recycled_weight

            print(f"[DEBUG] update_recycled_emission_value: manufacturing_electricity = {manufacturing_electricity}")
            print(f"[DEBUG] update_recycled_emission_value: metal_emission = {metal_emission}")
            print(f"[DEBUG] update_recycled_emission_value: recycled_weight = {recycled_weight}")
            print(f"[DEBUG] update_recycled_emission_value: recycled_emission = {recycled_emission}")

            recycled_emission_value_label.config(
                text=f"{recycled_emission:.2f}CO2eKg",
                font=("Arial", 10, "bold")
            )
        except Exception as e:
            print(f"[DEBUG] update_recycled_emission_value: Error - {e}")
            recycled_emission_value_label.config(
                text="0.00CO2eKg",
                font=("Arial", 10, "bold")
            )
        update_sum_of_emissions()

    # Create a new frame for the pie chart above the right_frame
    pie_chart_container = ttk.Labelframe(
        root,
        text="Emission Contributions",
        style="Custom.TLabelframe"
    )
    pie_chart_container.pack(side="top", fill="x", padx=10, pady=10)  # Place it above the right_frame

    # Chart type selector for emission chart
    chart_type_var = tk.StringVar(value="Pie Chart")
    chart_type_selector = ttk.Combobox(
        pie_chart_container,
        textvariable=chart_type_var,
        values=["Pie Chart", "Pareto Chart"],
        state="readonly",
        width=15
    )
    chart_type_selector.pack(side="top", anchor="ne", padx=10, pady=5)
    chart_type_selector.bind("<<ComboboxSelected>>", lambda e: update_pie_chart_mainthread())

    # Update the pie_chart_frame to use the new container
    pie_chart_frame = ttk.Frame(pie_chart_container)
    pie_chart_frame.pack(fill="both", expand=True)

    # --- Scrollable right_frame setup ---
    right_frame = tk.Frame(
        root,
        bg="#073B3A",
        width=625,
        highlightbackground="white",
        highlightcolor="white",
        highlightthickness=2
    )
    right_frame.pack(side="right", fill="y", padx=10, pady=10)
    right_frame.pack_propagate(False)

    # Create a canvas and a vertical scrollbar in right_frame
    right_canvas = tk.Canvas(right_frame, bg="#073B3A", highlightthickness=0, width=625)
    right_scrollbar = ttk.Scrollbar(right_frame, orient="vertical", command=right_canvas.yview)
    right_canvas.configure(yscrollcommand=right_scrollbar.set)

    right_canvas.pack(side="left", fill="both", expand=True)
    right_scrollbar.pack(side="right", fill="y")

    # Create a frame inside the canvas
    right_scrollable_frame = tk.Frame(right_canvas, bg="#073B3A")
    right_canvas.create_window((0, 0), window=right_scrollable_frame, anchor="nw")

    # Update scrollregion when the frame changes size
    def update_right_scrollregion(event):
        right_canvas.configure(scrollregion=right_canvas.bbox("all"))

    right_scrollable_frame.bind("<Configure>", update_right_scrollregion)

    # Enable mouse wheel scrolling for right_frame
    def right_mouse_wheel(event):
        right_canvas.yview_scroll(-1 * (event.delta // 120), "units")
        return "break"

    right_scrollable_frame.bind("<MouseWheel>", right_mouse_wheel)
    right_scrollable_frame.bind("<Button-4>", lambda event: right_canvas.yview_scroll(-1, "units"))  
    right_scrollable_frame.bind("<Button-5>", lambda event: right_canvas.yview_scroll(1, "units"))   

    # Create a custom style for frames and labels
    style = ttk.Style()
    style.configure("Eco.TFrame", background="#A8D5BA")  
    style.configure("Eco.TLabel", background="#A8D5BA")  
    style.configure("Eco.TLabelframe", background="#A8D5BA", borderwidth=0) 
    style.configure("Eco.TLabelframe.Label", background="#A8D5BA", font=("TkDefaultFont", 10, "bold")) 

    # Create custom styles for frames and labels in the right_frame
    style.configure(
        "Custom.TLabelframe",
        background="#073B3A",  # Match right_frame background
        foreground="white",    # White font color
        borderwidth=0          # Optional: Remove border
    )
    style.configure(
        "Custom.TLabelframe.Label",
        background="#073B3A",  # Match right_frame background
        foreground="white",    # White font color
        font=("TkDefaultFont", 10, "bold")  # Optional: Bold font
    )
    style.configure(
        "Custom.TLabel",
        background="#073B3A",  # Match right_frame background
        foreground="white"     # White font color
    )

    def validate_float_input(P):
        """
        Validates if the input string can be converted to a float.
        Allows empty input for flexibility.
        """
        print(f"[DEBUG] Validating Input: '{P}'")  # Debugging
        if not P:  # Allow empty input
            return True
        try:
            float(P)  # Attempt to convert to float
            return True
        except ValueError:
            return False  # Reject invalid input

    # Register validate_float with the Tkinter root
    validate_float = root.register(validate_float_input)

    # Add Manufacturing Country Selection to the right_frame
    manufacturing_country_frame = ttk.Labelframe(
        right_scrollable_frame,
        text="Manufacturing Country of Choice",
        style="Custom.TLabelframe"  # Use a custom style
    )
    manufacturing_country_frame.pack(fill="x", padx=10, pady=10)  

    manufacturing_country_label = ttk.Label(
        manufacturing_country_frame,
        text="Select Country:",
        style="Custom.TLabel"  # Use a custom style
    )
    manufacturing_country_label.grid(row=0, column=0, sticky="w", padx=5, pady=5)

    manufacturing_country_combo = ttk.Combobox(manufacturing_country_frame, width=25, state="readonly")
    manufacturing_country_combo.grid(row=0, column=1, sticky="ew", padx=5, pady=5)
    manufacturing_country_combo["values"] = country_list

    # Info icon with tooltip for manufacturing country
    info_icon_manufacturing_country = tk.Label(
        manufacturing_country_frame,
        text="ℹ️",
        bg="#073B3A",
        fg="white",
        font=("Arial", 12)
    )
    info_icon_manufacturing_country.grid(row=0, column=2, sticky="w", padx=(5, 0))
    ToolTip(
        info_icon_manufacturing_country,
        "Select the country where the product is manufactured."
    )

    def on_manufacturing_country_selected(event):
        if restoring_profile:
            return
        try:
            recalculate_all_carbon_emissions()
            update_recycled_emission_value()
            auto_save_inputs()
        except Exception as e:
            print(f"[ERROR] Exception in on_manufacturing_country_selected: {e}")
    manufacturing_country_combo.bind("<<ComboboxSelected>>", on_manufacturing_country_selected)

    # Weight Input
    weight_frame = ttk.Labelframe(scrollable_frame, text="Raw Weight Input", style="Eco.TLabelframe")
    weight_frame.grid(row=2, column=0, sticky="ew", padx=15, pady=0)  

    weight_label = ttk.Label(weight_frame, text="Enter Weight (kg):", style="Eco.TLabel")
    weight_label.grid(row=0, column=0, sticky="w")

    # Add validation for weight_entry to accept decimal values
    validate_float = root.register(validate_float_input)
    weight_entry = ttk.Entry(weight_frame, width=10, validate="key", validatecommand=(validate_float, '%P'))
    weight_entry.grid(row=0, column=1, sticky="w")

    # Info icon for weight_entry
    info_icon_weight_entry = tk.Label(
        weight_frame,
        text="ℹ️",
        bg="#A8D5BA",
        font=("Arial", 12)
    )
    info_icon_weight_entry.grid(row=0, column=2, sticky="w", padx=(5, 0))
    ToolTip(
        info_icon_weight_entry,
        "Enter the weight of the raw material in kilograms. Decimals are allowed.\nAlso, enter the number of raw material quantity."
    )

    # Add Quantity input under Weight Input
    quantity_label_weight = ttk.Label(weight_frame, text="Quantity:", style="Eco.TLabel")
    quantity_label_weight.grid(row=1, column=0, sticky="w")

    quantity_entry_weight = ttk.Entry(weight_frame, width=10, validate="key", validatecommand=(validate_float, '%P'))
    quantity_entry_weight.grid(row=1, column=1, sticky="w")
    quantity_entry_weight.insert(0, "1")  # Default to 1

    # Add Total Weight label
    total_weight_label = ttk.Label(weight_frame, text="Total Weight (kg): 0.0", style="Eco.TLabel")
    total_weight_label.grid(row=2, column=0, columnspan=2, sticky="w")

    def get_total_weight():
        try:
            weight = float(weight_entry.get()) if weight_entry.get() else 0.0
        except ValueError:
            weight = 0.0
        try:
            quantity = float(quantity_entry_weight.get()) if quantity_entry_weight.get() else 1.0
        except ValueError:
            quantity = 1.0
        # Convert to selected unit
        return weight * quantity

    def update_total_weight_and_recalculate(*args):
        global restoring_profile
        print("[DEBUG] update_total_weight_and_recalculate called")
        print("  weight_entry:", weight_entry.get())
        print("  quantity_entry_weight:", quantity_entry_weight.get())
        if restoring_profile:
            print("  [DEBUG] Skipping because restoring_profile is True")
            return
        total_weight = get_total_weight()
        print("  [DEBUG] Calculated total_weight:", total_weight)
        total_weight_label.config(text=f"Total Weight (kg): {total_weight:.2f}")
        update_recycled_weight(weight_entry, weight_after_entry, recycled_weight_value_label)
        calculate_material_value()
        recalculate_all_emissions()

    weight_entry.bind("<KeyRelease>", update_total_weight_and_recalculate)
    weight_entry.bind("<FocusOut>", update_total_weight_and_recalculate)
    quantity_entry_weight.bind("<KeyRelease>", update_total_weight_and_recalculate)
    quantity_entry_weight.bind("<FocusOut>", update_total_weight_and_recalculate)

    # Global debounce timer
    debounce_timer = None

    def debounce_weight_calculation():
        global debounce_timer
        if debounce_timer:
            debounce_timer.cancel() 
        print("[DEBUG] Starting debounce timer for weight_entry")
        debounce_timer = Timer(0.5, lambda: [
            calculate_material_value(), 
            recalculate_all_emissions(),  
            recalculate_all_carbon_emissions(),  
            update_sum_of_emissions()  
        ])
        debounce_timer.start()

    # --- Weight after Manufacturing Frame (with Quantity) ---
    weight_after_frame = ttk.Labelframe(
        right_scrollable_frame,
        text="Weight after Manufacturing",
        style="Custom.TLabelframe"
    )
    weight_after_frame.pack(fill="x", padx=10, pady=10)

    # Label for weight after manufacturing
    weight_after_label = ttk.Label(
        weight_after_frame,
        text="Enter Weight (kg):",
        style="Custom.TLabel"
    )
    weight_after_label.grid(row=0, column=0, sticky="w", padx=5, pady=5)

    weight_after_entry = ttk.Entry(weight_after_frame, width=10, validate="key", validatecommand=(validate_float, '%P'))
    weight_after_entry.grid(row=0, column=1, sticky="w", padx=5, pady=5)

    # Quantity input for weight after manufacturing
    weight_after_quantity_label = ttk.Label(weight_after_frame, text="Quantity:", style="Custom.TLabel")
    weight_after_quantity_label.grid(row=1, column=0, sticky="w", padx=5, pady=5)

    # Info icon for weight_after_quantity_entry
    info_icon_weight_after_qty = tk.Label(
        weight_after_frame,
        text="ℹ️",
        bg="#073B3A",
        fg="white",
        font=("Arial", 12)
    )
    info_icon_weight_after_qty.grid(row=1, column=2, sticky="w", padx=(5, 0), pady=5)
    ToolTip(
        info_icon_weight_after_qty,
        "Enter the weight of finished product after manufacturing in kilograms. \nPlease key in finish product quantity at the bottom entry."
    )

    weight_after_quantity_entry = ttk.Entry(weight_after_frame, width=10, validate="key", validatecommand=(validate_float, '%P'))
    weight_after_quantity_entry.grid(row=1, column=1, sticky="w", padx=5, pady=5)
    weight_after_quantity_entry.insert(0, "1")  # Default to 1

    weight_after_entry.bind("<KeyRelease>", lambda event: [update_packaging_emission_value(), update_packaging_emission_total()])
    weight_after_entry.bind("<FocusOut>", lambda event: [update_packaging_emission_value(), update_packaging_emission_total()])
    weight_after_quantity_entry.bind("<KeyRelease>", lambda event: [update_packaging_emission_value(), update_packaging_emission_total()])
    weight_after_quantity_entry.bind("<FocusOut>", lambda event: [update_packaging_emission_value(), update_packaging_emission_total()])

    # Total weight after manufacturing label
    weight_after_total_weight_label = ttk.Label(weight_after_frame, text="Total Weight (kg): 0.0", style="Custom.TLabel")
    weight_after_total_weight_label.grid(row=2, column=0, columnspan=2, sticky="w", padx=5, pady=5)

    def update_unit_labels():
        unit = "kg"
        weight_label.config(text=f"Enter Weight ({unit}):")
        total_weight_label.config(text=f"Total Weight ({unit}): {get_total_weight():.2f}")
        weight_after_label.config(text=f"Enter Weight ({unit}):")
        weight_after_total_weight_label.config(text=f"Total Weight ({unit}): {get_weight_after_total():.2f}")
        recycled_weight_label.config(text=f"Recycled Weight ({unit}):")

    # --- Parts for Assembly Input ---
    parts_quantity_frame = ttk.Labelframe(
        right_scrollable_frame,
        text="How many parts for assembly",
        style="Custom.TLabelframe"
    )
    parts_quantity_frame.pack(fill="x", padx=10, pady=10)

    # Info icon with tooltip for parts_quantity_frame
    info_icon_parts_quantity = tk.Label(
        parts_quantity_frame,
        text="ℹ️",
        bg="#073B3A",
        fg="white",
        font=("Arial", 12)
    )
    info_icon_parts_quantity.grid(row=0, column=2, sticky="w", padx=(5, 0), pady=5)
    ToolTip(
        info_icon_parts_quantity,
        "Specify the number of parts that will be assembled for the final product."
    )

    parts_quantity_label = ttk.Label(parts_quantity_frame, text="Number of Parts:", style="Custom.TLabel")
    parts_quantity_label.grid(row=0, column=0, sticky="w", padx=5, pady=5)

    parts_quantity_entry = ttk.Entry(parts_quantity_frame, width=5)
    parts_quantity_entry.grid(row=0, column=1, sticky="w", padx=5, pady=5)
    parts_quantity_entry.insert(0, "1")

    # "Recycled Weight" output
    recycled_weight_label = ttk.Label(
        weight_after_frame,
        text="Recycled Weight (kg):",
        style="Custom.TLabel"
    )
    recycled_weight_label.grid(row=0, column=2, sticky="w", padx=5, pady=5)

    recycled_weight_value_label = ttk.Label(
        weight_after_frame,
        text="0.0",
        style="Custom.TLabel"
    )
    recycled_weight_value_label.grid(row=0, column=3, sticky="w", padx=5, pady=5)

    def get_weight_after_total():
        try:
            w = float(weight_after_entry.get()) if weight_after_entry.get() else 0.0
        except ValueError:
            w = 0.0
        try:
            q = float(weight_after_quantity_entry.get()) if weight_after_quantity_entry.get() else 1.0
        except ValueError:
            q = 1.0
        # Convert to selected unit
        return w * q

    global get_weight_after_total_global
    get_weight_after_total_global = get_weight_after_total

    def update_weight_after_total_and_recycled(*args):
        global restoring_profile
        if restoring_profile:
            return
        total_weight_after = get_weight_after_total()
        weight_after_total_weight_label.config(text=f"Total Weight (kg): {total_weight_after:.2f}")
        try:
            total_weight = get_total_weight()
        except Exception:
            total_weight = 0.0
        recycled_weight = total_weight - total_weight_after
        recycled_weight_value_label.config(text=f"{recycled_weight:.2f}")
        update_packaging_emission_total()
        update_recycled_emission_value()
        recalculate_all_carbon_emissions()
        calculate_material_value()
        recalculate_all_emissions()
        update_sum_of_emissions()

    # Bindings for instant update
    weight_after_entry.bind("<KeyRelease>", update_weight_after_total_and_recycled)
    weight_after_entry.bind("<FocusOut>", update_weight_after_total_and_recycled)
    weight_after_quantity_entry.bind("<KeyRelease>", update_weight_after_total_and_recycled)
    weight_after_quantity_entry.bind("<FocusOut>", update_weight_after_total_and_recycled)

    # Bind weight_entry to update recycled weight and emission instantly
    def on_weight_entry_change(event):
        if restoring_profile:
            return
        update_recycled_weight(weight_entry, weight_after_entry, recycled_weight_value_label)
        calculate_material_value()
        recalculate_all_emissions()

    recycled_emission_label = tk.Label(
        recycling_inner_frame,
        text="Recycled Emission:",
        bg="#A8D5BA",
        fg="black",
        font=("Arial", 10, "bold")
    )
    recycled_emission_label.grid(row=1, column=0, sticky="w", padx=5, pady=5)

    recycled_emission_value_label = tk.Label(
        recycling_inner_frame,
        text="0.0",
        bg="#A8D5BA",
        fg="black",
        font=("Arial", 10)
    )
    recycled_emission_value_label.grid(row=1, column=1, sticky="w", padx=5, pady=5)
    recycled_weight_value_label.bind("<Configure>", lambda event: update_recycled_emission_value())

    def update_carbon_emission(machining_time_entry, energy_use_value_label, carbon_emission_value_label, weight_after_entry):
        try:
            machining_time = float(machining_time_entry.get()) if machining_time_entry.get() else 0.0
            energy_use = float(energy_use_value_label.cget("text")) if energy_use_value_label.cget("text") else 0.0
            weight_after = float(weight_after_entry.get()) if weight_after_entry.get() else 0.0

            try:
                weight_after_quantity = float(weight_after_quantity_entry.get()) if weight_after_quantity_entry.get() else 1.0
            except Exception:
                weight_after_quantity = 1.0

            selected_manufacturing_country = manufacturing_country_combo.get()
            if selected_manufacturing_country in country_list:
                manufacturing_index = country_list.index(selected_manufacturing_country)
                try:
                    electricity_emission = float(electricity_list[manufacturing_index])
                    print(f"[DEBUG] update_carbon_emission: Electricity Emission for '{selected_manufacturing_country}' = {electricity_emission}")
                except (ValueError, IndexError):
                    electricity_emission = 0.0
                    print(f"[DEBUG] update_carbon_emission: Invalid electricity emission for '{selected_manufacturing_country}'")
            else:
                electricity_emission = 0.0
                print(f"[DEBUG] update_carbon_emission: Manufacturing country '{selected_manufacturing_country}' not found")

            carbon_emission = energy_use * weight_after * weight_after_quantity * (machining_time / 60) * electricity_emission
            carbon_emission_value_label.config(text=f"{carbon_emission:.2f}")

            print(f"[DEBUG] update_carbon_emission: energy_use = {energy_use}, weight_after = {weight_after}, "
                  f"weight_after_quantity = {weight_after_quantity}, machining_time = {machining_time}, "
                  f"electricity_emission = {electricity_emission}")
            print(f"[DEBUG] update_carbon_emission: Calculated carbon emission = {carbon_emission:.2f}")

            update_sum_of_emissions()

        except ValueError as e:
            print(f"[DEBUG] update_carbon_emission: Error - {e}")
            carbon_emission_value_label.config(text="Invalid")

    def recalculate_all_carbon_emissions():
        """
        Recalculates carbon emissions for all rows and updates the corresponding labels.
        """
        print("[DEBUG] Recalculating carbon emissions for all machine rows...")
    
        selected_country = manufacturing_country_combo.get()
        country_idx = -1
        electricity_emission = 0.0
    
        if selected_country in country_list:
            country_idx = country_list.index(selected_country)
            try:
                electricity_emission = float(electricity_list[country_idx])
                print(f"[DEBUG] Using electricity emission: {electricity_emission} for '{selected_country}'")
            except (ValueError, IndexError) as e:
                print(f"[DEBUG] Error getting electricity emission for '{selected_country}': {e}")
        else:
            print(f"[DEBUG] Warning: Manufacturing country '{selected_country}' not found in country list")
    
        # Get weight after values
        try:
            weight_after = float(weight_after_entry.get()) if weight_after_entry.get() else 0.0
            weight_after_quantity = float(weight_after_quantity_entry.get()) if weight_after_quantity_entry.get() else 1.0
            print(f"[DEBUG] Weight after: {weight_after}, Quantity: {weight_after_quantity}")
        except ValueError as e:
            print(f"[DEBUG] Error parsing weight after values: {e}")
            weight_after = 0.0
            weight_after_quantity = 1.0
    
        for row_idx, row_widgets in enumerate(machine_process_row_widgets):
            try:
                machine_type = row_widgets["machine_type_combo"].get()
                print(f"[DEBUG] Row {row_idx+1} machine type: {machine_type}")
            
                energy_use = 0.0
                if machine_type in machine_value_list:
                    idx = machine_value_list.index(machine_type)
                    energy_use = float(specific_machine_energy_use_list[idx])
                    print(f"[DEBUG] Energy use for {machine_type}: {energy_use}")
                else:
                    print(f"[DEBUG] Machine type '{machine_type}' not found in list")
            
                machining_time = 0.0
                try:
                    machining_time = float(row_widgets["machining_time_entry"].get()) if row_widgets["machining_time_entry"].get() else 0.0
                    print(f"[DEBUG] Machining time: {machining_time}")
                except ValueError:
                    print(f"[DEBUG] Invalid machining time: {row_widgets['machining_time_entry'].get()}")
            
                carbon_emission = energy_use * weight_after * weight_after_quantity * (machining_time / 60) * electricity_emission
                print(f"[DEBUG] Calculated carbon emission: {carbon_emission:.2f}")
            
                row_widgets["energy_use_value_label"].config(text=f"{energy_use:.2f}")
                row_widgets["carbon_emission_value_label"].config(text=f"{carbon_emission:.2f}")

                if current_part_index < len(machine_parts_data):
                    part_data = machine_parts_data[current_part_index]
                    if "rows" in part_data and row_idx < len(part_data["rows"]):
                        part_data["rows"][row_idx]["energy_use"] = f"{energy_use:.2f}"
                        part_data["rows"][row_idx]["carbon_emission"] = f"{carbon_emission:.2f}"
        
            except Exception as e:
                print(f"[DEBUG] Error recalculating carbon emissions for row {row_idx+1}: {e}")
    
        recalculate_all_machine_part_emissions()

    def recalculate_all_machine_part_emissions():
        """Calculates total carbon emissions for each machine part by summing up row emissions"""
        print("[DEBUG] recalculate_all_machine_part_emissions called")
        for part_idx, part_data in enumerate(machine_parts_data):
            total = 0.0
            if "rows" in part_data:
                for row in part_data["rows"]:
                    try:
                        emission_value = row.get("carbon_emission", "0.0")
                        # Strip any text and convert to float
                        if isinstance(emission_value, str):
                            emission_value = emission_value.replace("CO2eKg", "").strip()
                        total += float(emission_value)
                    except (ValueError, TypeError) as e:
                        print(f"[DEBUG] Error parsing emission value '{row.get('carbon_emission')}': {e}")
                        # Continue with other rows even if one fails
            
                # Update the part's carbon_emission value
                part_data["carbon_emission"] = f"{total:.2f}"
                print(f"[DEBUG] Part {part_idx+1} total emission set to: {total:.2f}")
            
                # Update the StringVar for UI if it exists
                if "part_total_emission_var" in part_data and part_data["part_total_emission_var"] is not None:
                    try:
                        part_data["part_total_emission_var"].set(f"{total:.2f}")
                    except Exception as e:
                        print(f"[DEBUG] Error updating part_total_emission_var: {e}")
    
        # Log the updated data structure
        for i, part in enumerate(machine_parts_data):
            print(f"[DEBUG] After recalc - Part {i+1}: {part.get('carbon_emission', '0.0')} CO2eKg")
            if "rows" in part:
                for j, row in enumerate(part["rows"]):
                    print(f"  Process {j+1}: {row.get('machine_type', '')}, {row.get('carbon_emission', '0.0')} CO2eKg") 

    machine_process_page_frame = tk.Frame(right_scrollable_frame, bg="#073B3A")
    machine_process_page_frame.pack(fill="x", padx=10, pady=10)

    # Navigation buttons
    nav_frame = tk.Frame(right_scrollable_frame, bg="#073B3A")
    nav_frame.pack(fill="x", padx=10, pady=5)
    prev_btn = ttk.Button(nav_frame, text="← Previous")
    prev_btn.pack(side="left", padx=5)
    next_btn = ttk.Button(nav_frame, text="Next →")
    next_btn.pack(side="right", padx=5)

    def validate_machine_process_rows():
        global is_importing
        if is_importing:
            next_btn.config(state="normal")
            return
        valid_row_exists = any(
            row["machine_type_combo"].get().strip() and row["machining_time_entry"].get().strip()
            for row in machine_process_row_widgets
        )
        if valid_row_exists:
            next_btn.config(state="normal")
        else:
               next_btn.config(state="disabled")

    def update_machine_process_page():
        global current_part_index
        for widget in machine_process_page_frame.winfo_children():
            widget.destroy()
        if not machine_parts_data:
            return

        if current_part_index < 0 or current_part_index >= len(machine_parts_data):
            # Reset to first part if out of range
            current_part_index = 0
            if not machine_parts_data:
                return

        part_data = machine_parts_data[current_part_index]
        if "machine_row_qty" not in part_data:
            part_data["machine_row_qty"] = 1
        if "rows" not in part_data or len(part_data["rows"]) != part_data["machine_row_qty"]:
            part_data["rows"] = [
                {"machine_type": "", "energy_use": "", "machining_time": "", "carbon_emission": ""}
                for _ in range(part_data["machine_row_qty"])
            ]

        # --- Use persistent StringVar for part total emission ---
        if "part_total_emission_var" not in part_data:
            part_data["part_total_emission_var"] = tk.StringVar()
        part_total_emission_var = part_data["part_total_emission_var"]
        part_total_emission_var.set(part_data.get("carbon_emission", "0.0"))

        page_inner = tk.Frame(machine_process_page_frame, bg="#073B3A")
        page_inner.pack(fill="x", padx=5, pady=5)

        # Part label
        ttk.Label(
            page_inner,
            text=f"Part {current_part_index+1} of {len(machine_parts_data)}",
            style="Custom.TLabel"
        ).grid(row=0, column=0, columnspan=4, sticky="w", pady=5)

        # Machine Row Quantity Entry
        ttk.Label(page_inner, text="Number of Machine Processes", style="Custom.TLabel").grid(row=1, column=0, sticky="w", padx=5, pady=2)
        machine_row_qty_var = tk.StringVar(value=str(part_data["machine_row_qty"]))
        qty_entry_frame = tk.Frame(page_inner, bg=page_inner.cget("bg"))
        qty_entry_frame.grid(row=1, column=1, sticky="w", padx=5, pady=2)

        info_icon_machine_qty = tk.Label(
            qty_entry_frame,
            text="ℹ️:",
            font=("Arial", 12),
            bg="#073B3A",
            fg="white"
        )
        info_icon_machine_qty.pack(side="left", padx=(0, 2))
        ToolTip(
            info_icon_machine_qty,
            "Enter the number of machine processes for this part. \nEach process will have its own inputs of machine type, time(mins) and outputs below. \nIf no inputs, simply leave it empty and continue."
        )

        machine_row_qty_entry = ttk.Entry(qty_entry_frame, width=5, textvariable=machine_row_qty_var)
        machine_row_qty_entry.pack(side="left")

        def on_machine_row_qty_change(*args):
            try:
                qty = int(machine_row_qty_var.get())
                if qty < 1:
                    qty = 1
                part_data["machine_row_qty"] = qty
                while len(part_data["rows"]) < qty:
                    part_data["rows"].append({"machine_type": "", "energy_use": "0.0", "machining_time": "", "carbon_emission": "0.0"})
                while len(part_data["rows"]) > qty:
                    part_data["rows"].pop()
                update_machine_process_page()
            except ValueError:
                pass

        machine_row_qty_var.trace_add("write", lambda *a: on_machine_row_qty_change())

        # Table headers
        ttk.Label(page_inner, text="Machine Type", style="Custom.TLabel").grid(row=2, column=0, padx=5)
        ttk.Label(page_inner, text="Machining Time (min) w/o Setup", style="Custom.TLabel").grid(row=2, column=1, padx=5)
        ttk.Label(page_inner, text="Energy (kWh/kg)", style="Custom.TLabel").grid(row=2, column=2, padx=5)
        ttk.Label(page_inner, text="Carbon Emission", style="Custom.TLabel").grid(row=2, column=3, padx=5)

        row_vars = []
        for i, row in enumerate(part_data["rows"]):
            machine_type_values = [""] + machine_value_list
            machine_type_value = (row.get("machine_type", "") or "").strip()
            if machine_type_value not in machine_type_values:
                machine_type_value = ""
            machine_type_var = tk.StringVar(value=machine_type_value)
            machine_type_combo = ttk.Combobox(page_inner, width=18, values=machine_type_values, textvariable=machine_type_var, state="readonly")
            machine_type_combo.grid(row=3+i, column=0, padx=5, pady=2)
            machine_type_combo.set(machine_type_value)
            print(f"[DEBUG] Set machine_type_combo to: '{machine_type_value}' (values: {machine_type_values})")
            machining_time_var = tk.StringVar(value=row.get("machining_time", ""))
            print(f"[DEBUG] update_machine_process_page: Row {i} machining_time = {row.get('machining_time', '')}")
            machining_time_entry = ttk.Entry(page_inner, width=10, textvariable=machining_time_var)
            machining_time_entry.grid(row=3+i, column=1, padx=5, pady=2)
            energy_use_var = tk.StringVar(value=row.get("energy_use", "0.0"))
            energy_use_label = ttk.Label(page_inner, textvariable=energy_use_var, style="Custom.TLabel")
            energy_use_label.grid(row=3+i, column=2, padx=5, pady=2)
            carbon_emission_var = tk.StringVar(value=row.get("carbon_emission", "0.0"))
            carbon_emission_label = ttk.Label(page_inner, textvariable=carbon_emission_var, style="Custom.TLabel")
            carbon_emission_label.grid(row=3+i, column=3, padx=5, pady=2)
            row_vars.append({
                "machine_type": machine_type_var,
                "energy_use": energy_use_var,
                "machining_time": machining_time_var,
                "carbon_emission": carbon_emission_var
            })

            def on_machine_type_change(*args, idx=i):
                if restoring_profile:
                    return
                row = part_data["rows"][idx]
                row["machine_type"] = row_vars[idx]["machine_type"].get()
                if row["machine_type"] in machine_value_list:
                    m_idx = machine_value_list.index(row["machine_type"])
                    energy = specific_machine_energy_use_list[m_idx]
                    row["energy_use"] = energy
                    row_vars[idx]["energy_use"].set(energy)
                else:
                    row["energy_use"] = "0.0"
                    row_vars[idx]["energy_use"].set("0.0")
                make_row_updater(idx)()

            def on_machining_time_change(*args, idx=i):
                if restoring_profile:
                    return
                row = part_data["rows"][idx]
                row["machining_time"] = row_vars[idx]["machining_time"].get()
                make_row_updater(idx)()

            machine_type_var.trace_add("write", on_machine_type_change)
            machining_time_var.trace_add("write", on_machining_time_change)

        def make_row_updater(row_idx):
            def update_row_emission(*args):
                row = part_data["rows"][row_idx]
                try:
                    machining_time = float(row["machining_time"]) if row["machining_time"] else 0.0
                    energy_use = float(row["energy_use"]) if row["energy_use"] else 0.0
                    weight_after = float(weight_after_entry.get()) if weight_after_entry.get() else 0.0
                    weight_after_quantity = float(weight_after_quantity_entry.get()) if weight_after_quantity_entry.get() else 1.0
                    selected_country = manufacturing_country_combo.get()
                    if selected_country in country_list:
                        idx = country_list.index(selected_country)
                        electricity_emission = float(electricity_list[idx])
                    else:
                        electricity_emission = 0.0
                    carbon_emission = energy_use * weight_after * weight_after_quantity * (machining_time / 60) * electricity_emission
                    row["carbon_emission"] = f"{carbon_emission:.2f}"
                    row_vars[row_idx]["carbon_emission"].set(f"{carbon_emission:.2f}")
                except Exception:
                    row["carbon_emission"] = "0.0"
                    row_vars[row_idx]["carbon_emission"].set("0.0")
                # --- Always recalculate part and update UI ---
                recalculate_all_machine_part_emissions()
                # Update the StringVar for this part
                part_total_emission_var.set(part_data.get("carbon_emission", "0.0"))
                update_sum_of_emissions()
            return update_row_emission

        def update_all_rows_emission(*args):
            for idx in range(len(part_data["rows"])):
                make_row_updater(idx)()

        weight_after_entry.bind("<KeyRelease>", lambda event: update_all_rows_emission())
        weight_after_quantity_entry.bind("<KeyRelease>", lambda event: update_all_rows_emission())
        manufacturing_country_combo.bind("<<ComboboxSelected>>", lambda event: update_all_rows_emission())

        # Show total for this part
        ttk.Label(page_inner, text="Total Carbon Emission for this Part (CO2eKg):", style="Custom.TLabel").grid(
            row=3+len(part_data["rows"]), column=0, columnspan=2, sticky="e", padx=5, pady=5
        )
        ttk.Label(page_inner, textvariable=part_total_emission_var, style="Custom.TLabel").grid(
            row=3+len(part_data["rows"]), column=2, columnspan=2, sticky="w", padx=5, pady=5
        )
        def bind_mouse_wheel_to_right_canvas(widget):
            widget.bind("<MouseWheel>", lambda event: right_canvas.yview_scroll(-1 * (event.delta // 120), "units"))
            widget.bind("<Button-4>", lambda event: right_canvas.yview_scroll(-1, "units"))  # Linux scroll up
            widget.bind("<Button-5>", lambda event: right_canvas.yview_scroll(1, "units"))   # Linux scroll down
            for child in widget.winfo_children():
                bind_mouse_wheel_to_right_canvas(child)

        bind_mouse_wheel_to_right_canvas(machine_process_page_frame)

    def bind_mouse_wheel_to_right_canvas(widget):
        widget.bind("<MouseWheel>", lambda event: right_canvas.yview_scroll(-1 * (event.delta // 120), "units"))
        widget.bind("<Button-4>", lambda event: right_canvas.yview_scroll(-1, "units")) 
        widget.bind("<Button-5>", lambda event: right_canvas.yview_scroll(1, "units"))   
        for child in widget.winfo_children():
            bind_mouse_wheel_to_right_canvas(child)

    def sync_all_parts_ui_to_data():
        global current_part_index
        if not machine_parts_data or not machine_process_row_widgets:
            return
        original_index = current_part_index
        for idx in range(len(machine_parts_data)):
            current_part_index = idx
            update_machine_process_page()
            root.update_idletasks()  # Ensure UI is updated before syncing
            sync_current_part_ui_to_data()
        current_part_index = original_index
        update_machine_process_page()

    def go_to_prev_part():
        global current_part_index
        sync_current_part_ui_to_data()
        if current_part_index > 0:
            current_part_index -= 1
            scroll_pos = right_canvas.yview()
            update_machine_process_page()
            right_canvas.yview_moveto(scroll_pos[0])

    def go_to_next_part():
        global current_part_index
        sync_current_part_ui_to_data()
        if current_part_index < len(machine_parts_data) - 1:
            current_part_index += 1
            scroll_pos = right_canvas.yview()
            update_machine_process_page()
            right_canvas.yview_moveto(scroll_pos[0])

    # Debounce for parts quantity entry
    parts_debounce_id = None
    def debounced_on_parts_quantity_change(event=None):
        global parts_debounce_id
        if parts_debounce_id and root and root.winfo_exists():
            root.after_cancel(parts_debounce_id)
        if root and root.winfo_exists():
            parts_debounce_id = root.after(500, on_parts_quantity_change)

    parts_quantity_entry.bind("<KeyRelease>", debounced_on_parts_quantity_change)
    parts_quantity_entry.bind("<FocusOut>", lambda event: on_parts_quantity_change())

    def on_parts_quantity_change(*args):
        global restoring_profile
        if restoring_profile:
            return
        global current_part_index
        try:
            qty_str = parts_quantity_entry.get().strip()
            qty = int(qty_str) if qty_str else 1
            if qty < 1:
                qty = 1
            # Only regenerate if quantity actually changed
            if qty != len(machine_parts_data):
                while len(machine_parts_data) < qty:
                    machine_parts_data.append({})
                while len(machine_parts_data) > qty:
                    machine_parts_data.pop()
                if current_part_index >= qty:
                    current_part_index = qty - 1
                update_machine_process_page()
        except ValueError:
            pass

    def save_profile_inputs(profile_data, widgets):
        sync_current_part_ui_to_data()
        # Save static widgets
        for key in [
            'transport_country_combo', 'material_combo', 'weight_entry', 'quantity_entry_weight',
            'weight_after_entry', 'weight_after_quantity_entry', 'packaging_dropdown',
            'metal_recycling_dropdown', 'machine_process_quantity_entry', 'manufacturing_country_combo', 'quantity_entry'
        ]:
            widget = widgets.get(key)
            if widget and widget.winfo_exists() and hasattr(widget, 'get'):
                profile_data[key] = widget.get()
        print("Saved static widget values:", {k: profile_data[k] for k in profile_data if k not in ['transport_rows', 'machine_process_rows']})

        # Save dynamic transport rows
        profile_data['transport_rows'] = []
        for idx, row in enumerate(transport_row_widgets):
            transport_combo = row.get('transport_combo')
            if transport_combo and transport_combo.winfo_exists():
                row_data = {
                    'transport': transport_combo.get(),
                    'variation': row['variation_combo'].get() if row['variation_combo'].winfo_exists() else '',
                    'fuel': row['fuel_combo'].get() if row['fuel_combo'].winfo_exists() else '',
                    'origin': row['origin_entry'].get() if row.get('origin_entry') and row['origin_entry'].winfo_exists() else '',
                    'destination': row['destination_entry'].get() if row.get('destination_entry') and row['destination_entry'].winfo_exists() else '',
                    'mode': transport_combo.mode_var.get() if hasattr(transport_combo, 'mode_var') else '',
                    'value_label': row['value_label'].cget('text') if row.get('value_label') and row['value_label'].winfo_exists() else '',
                }
                print(f"[DEBUG][save_profile_inputs] Row {idx}: transport={row_data['transport']}, variation={row_data['variation']}, fuel={row_data['fuel']}, origin={row_data['origin']}, destination={row_data['destination']}, mode={row_data['mode']}, value_label={row_data['value_label']}")
                profile_data['transport_rows'].append(row_data)

        # Only update the current part (book-style navigation)
        if machine_parts_data and machine_process_row_widgets:
            part_data = machine_parts_data[current_part_index]
            for row_idx, row_widgets in enumerate(machine_process_row_widgets):
                if row_idx < len(part_data["rows"]):
                    row = part_data["rows"][row_idx]
                    row["machine_type"] = row_widgets["machine_type_combo"].get()
                    row["energy_use"] = row_widgets["energy_use_value_label"].cget("text")
                    row["machining_time"] = row_widgets["machining_time_entry"].get()
                    row["carbon_emission"] = row_widgets["carbon_emission_value_label"].cget("text")

        # Save dynamic machine process rows (book-style)
        profile_data['machine_process_rows'] = []
        for part_idx, part_data in enumerate(machine_parts_data, 1):
            if "rows" in part_data:
                for row_idx, row in enumerate(part_data["rows"], 1):
                    profile_data['machine_process_rows'].append({
                        "part": f"Part {part_idx}",
                        "row": row_idx,
                        "machine_type": row.get("machine_type", ""),
                        "energy_use": row.get("energy_use", "0.0"),
                        "machining_time": row.get("machining_time", ""),
                        "carbon_emission": row.get("carbon_emission", "0.0")
                    })
        print(f"[DEBUG][save_profile_inputs] Saved {len(profile_data['machine_process_rows'])} machine process rows (book-style).")
        prev_btn.config(command=go_to_prev_part)
        next_btn.config(command=go_to_next_part)
        parts_quantity_entry.bind("<KeyRelease>", lambda event: on_parts_quantity_change())

    # Rebind events for transportation country, material selection, and weight entry
    def rebind_events():
        """
        Rebinds events for transportation country, material selection, and weight entry.
        """
        transport_country_combo.bind("<<ComboboxSelected>>", lambda event: [
            update_transportation_country_values(transport_country_combo.get()),
            recalculate_all()
        ])
        material_combo.bind("<<ComboboxSelected>>", lambda event: calculate_material_value())

    # Function to update the processes output dynamically
    def update_processes_output(machine_processes_entry, processes_output_value_label):
        try:
            # Get the value from the input field
            quantity = float(machine_processes_entry.get()) if machine_processes_entry.get() else 0.0
            processes_output_value_label.config(text=f"{quantity:.2f}")
        except ValueError:
            processes_output_value_label.config(text="Invalid")

    def update_recycled_weight(weight_entry, weight_after_entry, recycled_weight_value_label):
        try:
            # Get the values from the input fields
            weight = float(weight_entry.get()) if weight_entry.get() else 0.0
            weight_after = float(weight_after_entry.get()) if weight_after_entry.get() else 0.0

            # Calculate the recycled weight
            recycled_weight = weight - weight_after

            # Update the label with the calculated value
            recycled_weight_value_label.config(text=f"{recycled_weight:.2f}")
            update_recycled_emission_value()
        except ValueError:
            # Handle invalid input gracefully
            recycled_weight_value_label.config(text="Invalid")
            update_recycled_emission_value()
    right_frame.pack_propagate(False)  

    def update_energy_use(event, machine_type_combo, energy_use_value_label):
        machine_type = machine_type_combo.get()
        if machine_type in machine_value_list:
            idx = machine_value_list.index(machine_type)
            energy = specific_machine_energy_use_list[idx]
            energy_use_value_label.config(text=energy)
        else:
            energy_use_value_label.config(text="0.0")    

    # Function to handle mouse wheel scrolling
    def on_mouse_wheel(event):
        if event.num == 4:  # Linux scroll up
            canvas.yview_scroll(-1, "units")
        elif event.num == 5:  # Linux scroll down
            canvas.yview_scroll(1, "units")
        else:  # Windows and macOS
            canvas.yview_scroll(-1 * (event.delta // 120), "units")

    # Bind the mouse wheel event to the canvas
    canvas.bind("<MouseWheel>", on_mouse_wheel) 
    canvas.bind("<Button-4>", on_mouse_wheel)    
    canvas.bind("<Button-5>", on_mouse_wheel)   

    # After creating canvas and scrollbar:
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    scrollbar.config(command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    def update_scroll_region(event):
        canvas.configure(scrollregion=canvas.bbox("all"))

    scrollable_frame.bind("<Configure>", update_scroll_region)

    # Create a custom style for frames and labels
    style = ttk.Style()
    style.configure("Eco.TFrame", background="#A8D5BA")  # Set eco-design background for frames
    style.configure("Eco.TLabel", background="#A8D5BA")  # Set eco-design background for labels
    style.configure("Eco.TLabelframe", background="#A8D5BA", borderwidth=0)  # Set eco-design background for Labelframes
    style.configure("Eco.TLabelframe.Label", background="#A8D5BA", font=("TkDefaultFont", 10, "bold"))  # Set frame text to blend in

    # Add Transportation Country Selection
    transport_country_frame = ttk.Labelframe(
        scrollable_frame, text="Transportation Country of Choice", style="Eco.TLabelframe"
    )
    transport_country_frame.grid(row=0, column=0, sticky="ew", padx=15, pady=5)
    transport_country_label = ttk.Label(
        transport_country_frame, text="Select Country:", style="Eco.TLabel"
    )
    transport_country_label.grid(row=0, column=0, sticky="w")
    transport_country_combo = ttk.Combobox(
        transport_country_frame, width=25, state="readonly"
    )
    transport_country_combo.grid(row=0, column=1, sticky="ew")
    transport_country_combo["values"] = country_list

    # Info icon 
    info_icon_transport = tk.Label(
        transport_country_frame,
        text="ℹ️",
        bg="#A8D5BA",
        font=("Arial", 12)
    )
    info_icon_transport.grid(row=0, column=2, sticky="w", padx=(5, 0))
    ToolTip(
        info_icon_transport,
        "Select the country from which the raw material will be provided from."
    )

    def on_transport_country_selected(event):
        if restoring_profile:
            return
        try:
            recalculate_all()
            auto_save_inputs()
        except Exception as e:
            print(f"[ERROR] Exception in on_transport_country_selected: {e}")
    transport_country_combo.bind("<<ComboboxSelected>>", on_transport_country_selected)

    # Create a custom style for the outlined frame
    style.configure("Black.TLabelframe", background="#A8D5BA", bordercolor="black", borderwidth=2)
    style.configure("Black.TLabelframe.Label", background="#A8D5BA", font=("TkDefaultFont", 10, "bold"))
    style.configure("Outlined.TFrame", background="#FA8072", borderwidth=2) #Material Emission Border Colour

    # Create a parent frame to simulate the border
    border_frame = tk.Frame(scrollable_frame, background="#A52A2A", padx=2, pady=2)  # Black border
    border_frame.grid(row=0, column=1, columnspan=2, rowspan=3, sticky="nsew", padx=10, pady=10)

    # Place the ttk.Labelframe inside the border frame
    materials_group_frame = ttk.Labelframe(border_frame, text="", style="Black.TLabelframe")
    materials_group_frame.pack(fill="both", expand=True)

    # Configure the parent frame to expand with its content
    materials_group_frame.grid_rowconfigure(0, weight=1)
    materials_group_frame.grid_rowconfigure(1, weight=1)
    materials_group_frame.grid_rowconfigure(2, weight=1)
    materials_group_frame.grid_columnconfigure(0, weight=1)
    materials_group_frame.grid_columnconfigure(1, weight=1)

    # Steel
    steel_frame = ttk.Labelframe(scrollable_frame, text="Steel (CO2eKg)", style="Eco.TLabelframe")
    steel_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10, in_=materials_group_frame)
    steel_frame.grid_rowconfigure(0, weight=1)  # Allow vertical centering
    steel_frame.grid_columnconfigure(0, weight=1)  # Allow horizontal centering
    steel_label = ttk.Label(steel_frame, text="0.0", style="Eco.TLabel")
    steel_label.grid(row=0, column=0, sticky="")

    # Aluminium
    aluminium_frame = ttk.Labelframe(scrollable_frame, text="Aluminium (CO2eKg)", style="Eco.TLabelframe")
    aluminium_frame.grid(row=0, column=1, sticky="nsew", padx=10, pady=10, in_=materials_group_frame)
    aluminium_frame.grid_rowconfigure(0, weight=1)  # Allow vertical centering
    aluminium_frame.grid_columnconfigure(0, weight=1)  # Allow horizontal centering
    aluminium_label = ttk.Label(aluminium_frame, text="0.0", style="Eco.TLabel")
    aluminium_label.grid(row=0, column=0, sticky="")

    # Cement
    cement_frame = ttk.Labelframe(scrollable_frame, text="Cement (CO2eKg)", style="Eco.TLabelframe")
    cement_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=10, in_=materials_group_frame)
    cement_frame.grid_rowconfigure(0, weight=1)  # Allow vertical centering
    cement_frame.grid_columnconfigure(0, weight=1)  # Allow horizontal centering
    cement_label = ttk.Label(cement_frame, text="0.0", style="Eco.TLabel")
    cement_label.grid(row=0, column=0, sticky="")

    # Electricity
    electricity_frame = ttk.Labelframe(scrollable_frame, text="Electricity (CO2eKg/kWh)", style="Eco.TLabelframe")
    electricity_frame.grid(row=2, column=1, sticky="nsew", padx=10, pady=10, in_=materials_group_frame)
    electricity_frame.grid_rowconfigure(0, weight=1)  # Allow vertical centering
    electricity_frame.grid_columnconfigure(0, weight=1)  # Allow horizontal centering
    electricity_label = ttk.Label(electricity_frame, text="0.0", style="Eco.TLabel")
    electricity_label.grid(row=0, column=0, sticky="")

    # Plastic
    plastic_frame = ttk.Labelframe(scrollable_frame, text="Global Plastic (CO2eKg)", style="Eco.TLabelframe")
    plastic_frame.grid(row=2, column=0, sticky="nsew", padx=10, pady=10, in_=materials_group_frame)
    plastic_frame.grid_rowconfigure(0, weight=1)  # Allow vertical centering
    plastic_frame.grid_columnconfigure(0, weight=1)  # Allow horizontal centering
    plastic_label = ttk.Label(plastic_frame, text="0.0", style="Eco.TLabel")
    plastic_label.grid(row=0, column=0, sticky="")

    # Carbon Fiber
    carbon_fiber_frame = ttk.Labelframe(scrollable_frame, text="Global Carbon Fiber (CO2eKg)", style="Eco.TLabelframe")
    carbon_fiber_frame.grid(row=1, column=1, sticky="nsew", padx=10, pady=10, in_=materials_group_frame)
    carbon_fiber_frame.grid_rowconfigure(0, weight=1)  # Allow vertical centering
    carbon_fiber_frame.grid_columnconfigure(0, weight=1)  # Allow horizontal centering
    carbon_fiber_label = ttk.Label(carbon_fiber_frame, text="0.0", style="Eco.TLabel")
    carbon_fiber_label.grid(row=0, column=0, sticky="")

    # Quantity Input
    quantity_frame = ttk.Labelframe(scrollable_frame, text="Number of Additional Transport Options", style="Eco.TLabelframe")
    quantity_frame.grid(row=4, column=0, sticky="ew", padx=15, pady=1)
    quantity_label = ttk.Label(quantity_frame, text="Enter Quantity:", style="Eco.TLabel")
    quantity_label.grid(row=0, column=0, sticky="w")
    quantity_entry = ttk.Entry(quantity_frame, width=5)
    quantity_entry.grid(row=0, column=1, sticky="w")

    # Info icon with tooltip
    info_icon_quantity = tk.Label(
        quantity_frame,
        text="ℹ️",
        bg="#A8D5BA",
        font=("Arial", 12)
    )
    info_icon_quantity.grid(row=0, column=2, sticky="w", padx=(5, 0))
    ToolTip(
        info_icon_quantity,
        "Enter the number of additional transport options. Each option will generate a new row for input."
    )

    # Bind the <KeyRelease> event to trigger automatic row generation
    def debounced_on_quantity_change(event=None):
        global debounce_id
        if debounce_id:
            root.after_cancel(debounce_id)
        debounce_id = root.after(500, on_quantity_change)
    quantity_entry.bind("<KeyRelease>", debounced_on_quantity_change)
    quantity_entry.bind("<FocusOut>", lambda event: on_quantity_change())

     # --- Business Travel (UI) ---  # NEW
    bt_parent = scrollable_frame  # same parent as Material/Transport sections

    bt_frame = ttk.Labelframe(bt_parent, text="Business Travel", style="Eco.TLabelframe", padding=10)
    bt_frame.grid(row=2, column=0, sticky="ew", padx=15, pady=5)  # put under Material (row=1)

    bt_qty_var = tk.StringVar(value=str(len(profile_data.get("business_travel_rows", [])) or 1))
    ttk.Label(bt_frame, text="Number of Business Travel Rows", style="Eco.TLabel").grid(row=0, column=0, sticky="w", padx=5, pady=5)
    bt_qty_entry = ttk.Entry(bt_frame, textvariable=bt_qty_var, width=6)
    bt_qty_entry.grid(row=0, column=1, sticky="w", padx=5, pady=5)
    ttk.Label(bt_frame, text="← Enter quantity here", style="Eco.TLabel").grid(row=0, column=2, sticky="w", padx=5)

    bt_rows_holder = tk.Frame(bt_frame, bg="#A8D5BA")
    bt_rows_holder.grid(row=1, column=0, columnspan=3, sticky="nsew", pady=(6, 0))
    bt_frame.grid_columnconfigure(0, weight=1)
    bt_frame.grid_rowconfigure(1, weight=1)

    def generate_business_travel_rows(quantity: int):
        for w in bt_rows_holder.winfo_children():
            w.destroy()
        business_travel_row_widgets.clear()
        headers = ["Travel Type", "Variation", "Fuel", "Origin", "Destination", "Mode"]
        for c, h in enumerate(headers):
            ttk.Label(bt_rows_holder, text=h, style="Eco.TLabel").grid(row=0, column=c, padx=5, pady=(0,4), sticky="w")

        existing = profile_data.get("business_travel_rows", [])
        for i in range(max(0, quantity)):
            row_state = existing[i] if i < len(existing) else {}
            type_var = tk.StringVar(value=row_state.get("transport",""))
            type_combo = ttk.Combobox(bt_rows_holder, width=18, textvariable=type_var, values=["Van"], state="readonly")
            type_combo.grid(row=i+1, column=0, padx=5, pady=3, sticky="w")

            var_var = tk.StringVar(value=row_state.get("variation",""))
            var_combo = ttk.Combobox(bt_rows_holder, width=18, textvariable=var_var, state="readonly")
            var_combo.grid(row=i+1, column=1, padx=5, pady=3, sticky="w")

            fuel_var = tk.StringVar(value=row_state.get("fuel",""))
            fuel_combo = ttk.Combobox(bt_rows_holder, width=14, textvariable=fuel_var, state="readonly")
            fuel_combo.grid(row=i+1, column=2, padx=5, pady=3, sticky="w")

            origin_var = tk.StringVar(value=row_state.get("origin",""))
            dest_var   = tk.StringVar(value=row_state.get("destination",""))
            origin_entry = ttk.Entry(bt_rows_holder, width=16, textvariable=origin_var)
            dest_entry   = ttk.Entry(bt_rows_holder, width=16, textvariable=dest_var)
            origin_entry.grid(row=i+1, column=3, padx=5, pady=3, sticky="w")
            dest_entry.grid(row=i+1, column=4, padx=5, pady=3, sticky="w")

            mode_var = tk.StringVar(value=row_state.get("mode","None"))
            mode_combo = ttk.Combobox(bt_rows_holder, width=12, textvariable=mode_var, values=["None","driving","plane","ship"], state="readonly")
            mode_combo.grid(row=i+1, column=5, padx=5, pady=3, sticky="w")

            value_label = ttk.Label(bt_rows_holder, text="Total Emission: 0.00 CO2eKg", style="Eco.TLabel")
            value_label.grid(row=i+1, column=6, padx=5, pady=3, sticky="w")
            business_travel_row_widgets.append({
                "type_combo": type_combo,
                "variation_combo": var_combo,
                "fuel_combo": fuel_combo,
                "origin_entry": origin_entry,
                "destination_entry": dest_entry,
                "mode_combo": mode_combo,
                "value_label": value_label,
            })
            total_emission_labels.append(value_label)

            def update_dependent(i=i):
                t = business_travel_row_widgets[i]["type_combo"].get()

                # Always reset both combos first
                business_travel_row_widgets[i]["variation_combo"]["values"] = []
                business_travel_row_widgets[i]["fuel_combo"]["values"] = []
                business_travel_row_widgets[i]["variation_combo"].set("")
                business_travel_row_widgets[i]["fuel_combo"].set("")

                # Default: disable both until we know it's Van
                business_travel_row_widgets[i]["variation_combo"].config(state="readonly")
                business_travel_row_widgets[i]["fuel_combo"].config(state="readonly")

                if t.lower() == "van":
                    # Pull variation and fuel data for Van from your Excel-based lists
                    business_travel_row_widgets[i]["variation_combo"]["values"] = van_list
                    if not business_travel_row_widgets[i]["variation_combo"].get() and van_list:
                        business_travel_row_widgets[i]["variation_combo"].set(van_list[0])

                    # Variation options for Van (from your emission data)
                    business_travel_row_widgets[i]["variation_combo"]["values"] = van_list
                    if not business_travel_row_widgets[i]["variation_combo"].get() and van_list:
                        business_travel_row_widgets[i]["variation_combo"].set(van_list[0])

                    # Fuel options for Van (fixed 3 choices)
                    business_travel_row_widgets[i]["fuel_combo"]["values"] = ["Diesel", "Petrol", "Battery"]
                    if not business_travel_row_widgets[i]["fuel_combo"].get():
                        business_travel_row_widgets[i]["fuel_combo"].set("Diesel")
                else:
                    # For future expansion — other travel types can go here
                    business_travel_row_widgets[i]["variation_combo"]["values"] = []
                    business_travel_row_widgets[i]["fuel_combo"]["values"] = []


    def on_bt_qty_change(*_):
        try:
            n = max(0, int(bt_qty_var.get()))
        except Exception:
            n = 0
        generate_business_travel_rows(n)

    bt_qty_var.trace_add("write", on_bt_qty_change)
    initial_bt_n = len(profile_data.get("business_travel_rows", [])) or 1
    bt_qty_var.set(str(initial_bt_n))
    generate_business_travel_rows(initial_bt_n)

 
    # Material Selection
    material_frame = ttk.Labelframe(scrollable_frame, text="Material Selection", style="Eco.TLabelframe")
    material_frame.grid(row=1, column=0, sticky="ew", padx=15, pady=0)  
    material_label = ttk.Label(material_frame, text="Select Material:", style="Eco.TLabel")
    material_label.grid(row=0, column=0, sticky="w")
    material_combo = ttk.Combobox(material_frame, width=25, state="readonly")
    material_combo.grid(row=0, column=1, sticky="ew")
    materials = ["Steel", "Aluminium", "Cement", "Plastic", "Carbon Fiber"]
    material_combo["values"] = materials

    # Info icon with tooltip
    info_icon_material = tk.Label(
        material_frame,
        text="ℹ️",
        bg="#A8D5BA",
        font=("Arial", 12)
    )
    info_icon_material.grid(row=0, column=2, sticky="w", padx=(5, 0))
    ToolTip(
        info_icon_material,
        "Select the raw material for your product."
    )

    # Bind the <<ComboboxSelected>> event to trigger automatic calculation
    def on_material_selected(event):
        if restoring_profile:
            return
        try:
            auto_save_inputs()
            recalculate_all()
        except Exception as e:
            print(f"[ERROR] Exception in on_material_selected: {e}")
    material_combo.bind("<<ComboboxSelected>>", on_material_selected)

    # Calculated Value Output
    calculated_value_frame = ttk.Labelframe(scrollable_frame, text="Material Emission [CO2eKg]", style="Eco.TLabelframe")
    calculated_value_frame.grid(row=3, column=0, sticky="ew", padx=15, pady=5)  

    # Create an outlined box for the label
    outlined_frame = ttk.Frame(calculated_value_frame, style="Outlined.TFrame", padding=3)  
    outlined_frame.grid(row=0, column=0, sticky="ew", padx=5, pady=5)

    # Add the label inside the outlined box
    calculated_value_label = ttk.Label(outlined_frame, text="Value: ", style="Eco.TLabel")
    calculated_value_label.grid(row=0, column=0, sticky="w")

    # Create a frame to group the title and value output
    sum_emission_frame = tk.Frame(header_frame, bg="#FFFFFF", highlightbackground="#FFFFFF", highlightthickness=2)
    sum_emission_frame.pack(side="left", padx=40, pady=10)  # Place it slightly to the right of the logo

    # Create a sub-frame for the dark green background inside the white outline
    sum_emission_inner_frame = tk.Frame(sum_emission_frame, bg="#006400", padx=10, pady=10)  # Dark green background
    sum_emission_inner_frame.pack(fill="both", expand=True)

    # Sum of Emissions Output
    sum_emission_label = ttk.Label(
        sum_emission_inner_frame,
        text="Total Sum of Emission: 0.0 CO2eKg",
        style="Eco.TLabel",
        anchor="center"  # Center-align the text
    )
    sum_emission_label.pack(anchor="center", pady=(5, 0)) 

    # Adjust font size and make it bold
    sum_emission_label.config(font=("Arial", 14, "bold"), background="#006400", foreground="#FFFFFF")

    # --- Threshold Section ---
    # Variable to track if threshold is enabled
    threshold_enabled_var = tk.BooleanVar(value=False)
    # Variable to store the threshold value
    threshold_value_var = tk.StringVar(value="")

    # Add "Set Carbon Emission Threshold:" label
    threshold_label = tk.Label(
        sum_emission_inner_frame,
        text="Set Carbon Emission Threshold:",
        bg="#006400",
        fg="white",
        font=("Arial", 10, "bold")
    )
    threshold_label.pack(anchor="w", pady=(10, 0))

    # Add the checkbox
    threshold_checkbox = tk.Checkbutton(
        sum_emission_inner_frame,
        variable=threshold_enabled_var,
        command=lambda: on_threshold_toggle(),
        bg="#006400",
        fg="white",
        selectcolor="#006400",
        activebackground="#006400",
        activeforeground="white"
    )
    threshold_checkbox.pack(anchor="w", pady=(0, 0))

    # Add the threshold entry
    threshold_entry = tk.Entry(
        sum_emission_inner_frame,
        textvariable=threshold_value_var,
        width=10,
        bg="white",
        validate="key",
        validatecommand=(validate_float, '%P')
    )
    def on_threshold_toggle():
        if threshold_enabled_var.get():
            threshold_entry.pack(anchor="w", pady=(0, 5))
        else:
            threshold_entry.pack_forget()
            threshold_value_var.set("")
        update_sum_of_emissions()

    def on_threshold_value_change(*args):
        update_sum_of_emissions()

    # Bind changes in the entry to update highlight
    threshold_value_var.trace_add("write", lambda *args: on_threshold_value_change())
    on_threshold_toggle()

    # Create a canvas to hold the outlined_box_frame
    outlined_canvas = tk.Canvas(scrollable_frame, bg="#A8D5BA", highlightthickness=0)
    outlined_canvas.grid(row=5, column=0, columnspan=3, sticky="nsew", padx=15, pady=10)

    # Function to update the scroll region dynamically
    def update_scroll_region(event):
        canvas.configure(scrollregion=canvas.bbox("all"))

    # Function to forward mouse wheel events from child frames to the canvas
    def forward_mouse_wheel_to_canvas(event):
        canvas.yview_scroll(-1 * (event.delta // 120), "units") 
        return "break"  

    # Bind mouse wheel events for "Transportation Country of Choice" frame
    transport_country_frame.bind("<MouseWheel>", forward_mouse_wheel_to_canvas) 
    transport_country_frame.bind("<Button-4>", lambda event: canvas.yview_scroll(-1, "units"))  
    transport_country_frame.bind("<Button-5>", lambda event: canvas.yview_scroll(1, "units"))   

    # Bind mouse wheel events for "Material Selection" frame
    material_frame.bind("<MouseWheel>", forward_mouse_wheel_to_canvas) 
    material_frame.bind("<Button-4>", lambda event: canvas.yview_scroll(-1, "units")) 
    material_frame.bind("<Button-5>", lambda event: canvas.yview_scroll(1, "units"))   

    # Bind mouse wheel events for "Weight Input" frame
    weight_frame.bind("<MouseWheel>", forward_mouse_wheel_to_canvas)  
    weight_frame.bind("<Button-4>", lambda event: canvas.yview_scroll(-1, "units"))  
    weight_frame.bind("<Button-5>", lambda event: canvas.yview_scroll(1, "units"))   

    # Bind mouse wheel events for "Material Emission" frame
    calculated_value_frame.bind("<MouseWheel>", forward_mouse_wheel_to_canvas) 
    calculated_value_frame.bind("<Button-4>", lambda event: canvas.yview_scroll(-1, "units"))  
    calculated_value_frame.bind("<Button-5>", lambda event: canvas.yview_scroll(1, "units"))  

    # Bind mouse wheel events for "Number of Additional Transport Options" frame
    quantity_frame.bind("<MouseWheel>", forward_mouse_wheel_to_canvas)  
    quantity_frame.bind("<Button-4>", lambda event: canvas.yview_scroll(-1, "units"))  
    quantity_frame.bind("<Button-5>", lambda event: canvas.yview_scroll(1, "units"))  

    # Create the outlined_box_frame inside the canvas
    outlined_box_frame = tk.Frame(outlined_canvas, bg="black", padx=2, pady=2)
    outlined_canvas.create_window((0, 0), window=outlined_box_frame, anchor="nw")
    outlined_canvas.grid(row=5, column=0, columnspan=3, sticky="nsew", padx=15, pady=10)
    scrollable_frame.grid_rowconfigure(5, weight=1)  # Allow row 5 to expand

    # Bind mouse wheel events for outlined_box_frame
    outlined_box_frame.bind("<MouseWheel>", forward_mouse_wheel_to_canvas) 
    outlined_box_frame.bind("<Button-4>", lambda event: canvas.yview_scroll(-1, "units"))  
    outlined_box_frame.bind("<Button-5>", lambda event: canvas.yview_scroll(1, "units"))  

    # Create the inner frame with a background color
    inner_box_frame = tk.Frame(outlined_box_frame, bg="#A8D5BA", width=400, height=500, padx=10, pady=10)
    inner_box_frame.pack(fill="both", expand=True)

    outlined_box_frame.grid_remove()
    inner_box_frame.pack_forget()
    outlined_canvas.configure(scrollregion=(0, 0, 0, 0))
    scrollbar.configure(command=None)
    canvas.unbind("<MouseWheel>")
    canvas.unbind("<Button-4>")
    canvas.unbind("<Button-5>")

    machine_parts_data = [{}]
    update_machine_process_page()

    bind_mouse_wheel_to_right_canvas(right_scrollable_frame)

    # After all widgets and dynamic rows are created in show_main_page:
    machine_process_quantity_entry = parts_quantity_entry
    widgets = {
        'transport_country_combo': transport_country_combo,
        'material_combo': material_combo,
        'weight_entry': weight_entry,
        'quantity_entry_weight': quantity_entry_weight,
        'weight_after_entry': weight_after_entry,
        'weight_after_quantity_entry': weight_after_quantity_entry,
        'packaging_dropdown': packaging_dropdown,
        'metal_recycling_dropdown': metal_recycling_dropdown,
        'machine_process_quantity_entry': machine_process_quantity_entry,
        'manufacturing_country_combo': manufacturing_country_combo,
        'quantity_entry': quantity_entry,
    }
    root.widgets = widgets
    store_all_widget_values(widgets)
    
    # --- Place the binding code here ---
    for key in [
       'transport_country_combo', 'material_combo', 'weight_entry', 'quantity_entry_weight',
       'weight_after_entry', 'weight_after_quantity_entry', 'packaging_dropdown',
       'metal_recycling_dropdown', 'manufacturing_country_combo'
    ]:
        widget = widgets.get(key)
        if widget:
            def make_store_func(w=widget):
                return lambda event: store_all_widget_values(widgets)
            if isinstance(widget, ttk.Combobox):
                widget.bind("<Button-1>", make_store_func())  # Mouse click before selection
                widget.bind("<FocusIn>", make_store_func())   # Focus in
            else:
                widget.bind("<FocusIn>", make_store_func())

    save_status_label = tk.Label(header_frame, text="", font=("Arial", 9), bg="#A8D5BA", fg="green")
    save_status_label.pack(side="right", padx=10, pady=10)

    def debounced_auto_save_inputs(event=None):
        global restoring_profile
        if restoring_profile:
            return
        nonlocal auto_save_after_id
        if auto_save_after_id:
            root.after_cancel(auto_save_after_id)
        auto_save_after_id = root.after(1000, auto_save_inputs)  # 1 second debounce
   
    def auto_save_inputs(*args):
        global restoring_profile, is_navigating_back
        if restoring_profile or is_navigating_back:
            return
        print("[DEBUG][auto_save_inputs] Triggered, machine_process_row_widgets length:", len(machine_process_row_widgets))
        for idx, row in enumerate(transport_row_widgets):
            try:
                print(f"[DEBUG][auto_save_inputs] Row {idx}: transport={row['transport_combo'].get()}, variation={row['variation_combo'].get()}, fuel={row['fuel_combo'].get()}")
            except Exception as e:
                print(f"[DEBUG][auto_save_inputs] Row {idx}: Error reading values: {e}")

        for idx, row in enumerate(machine_process_row_widgets):
            try:
                print(f"[DEBUG][auto_save_inputs] Machine Row {idx}: machine_type={row['machine_type_combo'].get()}, energy_use={row['energy_use_value_label'].cget('text')}, machining_time={row['machining_time_entry'].get()}, carbon_emission={row['carbon_emission_value_label'].cget('text')}")
            except Exception as e:
                print(f"[DEBUG][auto_save_inputs] Machine Row {idx}: Error reading values: {e}")

        print("Auto-saving quantity_entry:", widgets['quantity_entry'].get())
        save_profile_inputs(profile_data, widgets)
        if active_profile_name and hasattr(cover_page, "profiles"):
            cover_page.profiles[active_profile_name] = dict(profile_data)
            save_profiles(cover_page.profiles)
            save_status_label.config(text="Profile auto-saved.") 
            def clear_status():
                if save_status_label.winfo_exists():
                    save_status_label.config(text="")
            save_status_label.after(1500, clear_status)

    # Bind auto-save to static widgets
    for key in [
       'transport_country_combo', 'material_combo', 'weight_entry', 'quantity_entry_weight',
       'weight_after_entry', 'weight_after_quantity_entry', 'packaging_dropdown',
       'metal_recycling_dropdown', 'manufacturing_country_combo'
    ]:
        widget = widgets.get(key)
        if widget:
            if isinstance(widget, ttk.Combobox):
                # Only bind <<ComboboxSelected>> for Comboboxes, not <KeyRelease>
                if key == 'packaging_dropdown':
                    widget.bind("<<ComboboxSelected>>", on_packaging_selected)
                elif key == 'metal_recycling_dropdown':
                    widget.bind("<<ComboboxSelected>>", on_metal_recycling_selected)
                elif key == 'transport_country_combo':
                    widget.bind("<<ComboboxSelected>>", on_transport_country_selected)
                elif key == 'material_combo':
                    widget.bind("<<ComboboxSelected>>", on_material_selected)
                elif key == 'manufacturing_country_combo':
                    widget.bind("<<ComboboxSelected>>", on_manufacturing_country_selected)
            else:
                widget.bind("<FocusOut>", lambda event: auto_save_inputs())
                widget.bind("<KeyRelease>", debounced_auto_save_inputs)
   
    def restore_profile_inputs(profile_data, widgets, generate_transport_rows, recalculate_all):
        print("[DEBUG] restore_profile_inputs called")
        global restoring_profile
        restoring_profile = True
        print("Restoring profile data:", profile_data)

        # Restore static widgets
        for key in [
            'transport_country_combo', 'material_combo', 'weight_entry', 'quantity_entry_weight',
            'weight_after_entry', 'weight_after_quantity_entry', 'packaging_dropdown',
            'metal_recycling_dropdown', 'machine_process_quantity_entry', 'manufacturing_country_combo', 'quantity_entry'
        ]:
            widget = widgets.get(key)
            value = profile_data.get(key, '')
            if widget and hasattr(widget, 'set'):
                widget.set(value)
            elif widget and hasattr(widget, 'delete') and hasattr(widget, 'insert'):
                widget.delete(0, tk.END)
                widget.insert(0, value)
        update_transportation_country_values(widgets['transport_country_combo'].get())

        print("[DEBUG] After restoring static widgets:")
        print("  weight_entry:", widgets['weight_entry'].get())
        print("  quantity_entry_weight:", widgets['quantity_entry_weight'].get())

        # --- Set quantity fields before generating rows ---
        transport_rows = profile_data.get('transport_rows', [])
        quantity_str = profile_data.get('quantity_entry', '')
        if 'quantity_entry' in widgets:
            widgets['quantity_entry'].delete(0, tk.END)
            widgets['quantity_entry'].insert(0, quantity_str)

        machine_rows = profile_data.get('machine_process_rows', [])
        machine_quantity_str = profile_data.get('machine_process_quantity_entry', '')
        if 'machine_process_quantity_entry' in widgets:
            widgets['machine_process_quantity_entry'].delete(0, tk.END)
            widgets['machine_process_quantity_entry'].insert(0, machine_quantity_str)

        # Generate the correct number of machine process rows
        try:
            machine_quantity = int(machine_quantity_str) if machine_quantity_str else len(machine_rows)
        except Exception:
            machine_quantity = len(machine_rows)
        update_machine_process_page()
        print(f"[DEBUG] After generate_machine_process_rows: {len(machine_process_row_widgets)} rows")

        # Generate the correct number of transport rows
        try:
            transport_quantity = int(quantity_str) if quantity_str else len(transport_rows)
        except Exception:
            transport_quantity = len(transport_rows)
        generate_transport_rows(transport_quantity, None)

        if transport_quantity > 0:
            outlined_box_frame.grid()
            inner_box_frame.pack(fill="both", expand=True)
            outlined_canvas.update_idletasks()
            outlined_canvas.configure(scrollregion=outlined_canvas.bbox("all"))

        # Restore values for each transport row (delayed to ensure widgets exist)
        def restore_transport_rows():
            for i, row_data in enumerate(transport_rows):
                if i < len(transport_row_widgets):
                    row = transport_row_widgets[i]
                    if row.get('transport_combo') and row['transport_combo'].winfo_exists():
                        row['transport_combo'].set(row_data.get('transport', ''))
                        update_transport_types(type('Event', (object,), {'widget': row['transport_combo']})())
                        def delayed_restore(row=row, row_data=row_data):
                            if (row['variation_combo'] and row['variation_combo'].winfo_exists()):
                                row['variation_combo'].set(row_data.get('variation', ''))
                            if (row['fuel_combo'] and row['fuel_combo'].winfo_exists()):
                                row['fuel_combo'].set(row_data.get('fuel', ''))
                            if (row['origin_entry'] and row['origin_entry'].winfo_exists()):
                                row['origin_entry'].delete(0, tk.END)
                                row['origin_entry'].insert(0, row_data.get('origin', ''))
                            if (row['destination_entry'] and row['destination_entry'].winfo_exists()):
                                row['destination_entry'].delete(0, tk.END)
                                row['destination_entry'].insert(0, row_data.get('destination', ''))
                            if (row['transport_combo'] and row['transport_combo'].winfo_exists()):
                                row['transport_combo'].mode_var.set(row_data.get('mode', 'None'))
                                def after_distance():
                                    if (row['value_label'] and row['value_label'].winfo_exists() and
                                        row['result_label'] and row['result_label'].winfo_exists() and
                                        row['total_emission_label'] and row['total_emission_label'].winfo_exists() and
                                        row['transport_combo'] and row['transport_combo'].winfo_exists()):
                                        update_value(type('Event', (object,), {'widget': row['transport_combo']})(), row['transport_combo'])
                                        calculate_total_emission(
                                            value_label=row['value_label'],
                                            result_label=row['result_label'],
                                            total_emission_label=row['total_emission_label'],
                                            weight_entry=widgets['weight_entry'],
                                            transport_combo=row['transport_combo']
                                        )
                                calculate_distance_dynamic_with_callback(
                                    row['origin_entry'],
                                    row['destination_entry'],
                                    row['transport_combo'].mode_var,
                                    row['result_label'],
                                    after_distance
                                )
                        root.after(100, delayed_restore)
            # After all transport rows are restored, continue with the rest
            restore_machine_rows_and_finalize()

        def restore_machine_rows_and_finalize():
            print("[DEBUG] restore_machine_rows_and_finalize called")
            # Restore values for each machine process row
            def restore_machine_row(row, row_data):
                row['machine_type_combo'].set(row_data.get('machine_type', ''))
                row['energy_use_value_label'].config(text=row_data.get('energy_use', ''))
                update_energy_use(None, row['machine_type_combo'], row['energy_use_value_label'])
                row['machining_time_entry'].delete(0, tk.END)
                row['machining_time_entry'].insert(0, row_data.get('machining_time', ''))
                row['carbon_emission_value_label'].config(text=row_data.get('carbon_emission', ''))
                update_carbon_emission(
                    row['machining_time_entry'],
                    row['energy_use_value_label'],
                    row['carbon_emission_value_label'],
                    widgets.get('weight_after_entry')
                )

            widget = widgets.get('metal_recycling_dropdown')
            value = profile_data.get('metal_recycling_dropdown', '')
            if widget and hasattr(widget, 'set'):
                widget.set(value)
                update_metal_recycling_value(None)
                update_recycled_emission_value()
            widget = widgets.get('packaging_dropdown')
            value = profile_data.get('packaging_dropdown', '')
            if widget and hasattr(widget, 'set'):
                widget.set(value)
                widget.after(50, lambda: [update_packaging_emission_value(None), update_packaging_emission_total()])
            # --- Ensure event handlers are triggered after restoring dropdowns ---
            if widgets.get('packaging_dropdown'):
                root.after(100, lambda: widgets['packaging_dropdown'].event_generate('<<ComboboxSelected>>'))
            if widgets.get('metal_recycling_dropdown'):
                root.after(100, lambda: widgets['metal_recycling_dropdown'].event_generate('<<ComboboxSelected>>'))
            print(f"[DEBUG] Restoring packaging dropdown to: '{value}'")

            # --- Book-style machine process restoration ---
            if machine_rows:
                parts = {}
                for row in machine_rows:
                    if isinstance(row, dict):
                        part_info = row.get('part', '')
                        machine_type = row.get('machine_type', '')
                        energy_use = row.get('energy_use', '0.0')
                        machining_time = row.get('machining_time', '')
                        carbon_emission = row.get('carbon_emission', '0.0')
                    else:
                        part_info = row[0]
                        machine_type = row[1] if len(row) > 1 else ""
                        energy_use = row[2] if len(row) > 2 else "0.0"
                        machining_time = row[3] if len(row) > 3 else ""
                        carbon_emission = row[4] if len(row) > 4 else "0.0"
                    if "Part" in str(part_info):
                        try:
                            part_num = int(str(part_info).split()[1])
                        except Exception:
                            continue
                        if part_num not in parts:
                            parts[part_num] = []
                        parts[part_num].append({
                            "machine_type": machine_type,
                            "energy_use": energy_use,
                            "machining_time": machining_time,
                            "carbon_emission": carbon_emission
                        })
                machine_qty = str(len(parts))
                widgets['machine_process_quantity_entry'].delete(0, tk.END)
                widgets['machine_process_quantity_entry'].insert(0, machine_qty)
                on_parts_quantity_change()
                # Rebuild machine_parts_data to match the number of parts in the profile
                machine_parts_data.clear()
                for part_num in sorted(parts.keys()):
                    rows = parts[part_num]
                    # Fill missing fields with defaults if needed
                    for r in rows:
                        if not r.get("machine_type"):
                            r["machine_type"] = machine_value_list[0] if machine_value_list else ""
                        if not r.get("energy_use"):
                            r["energy_use"] = specific_machine_energy_use_list[0] if specific_machine_energy_use_list else "0.0"
                        if not r.get("machining_time"):
                            r["machining_time"] = ""
                        if not r.get("carbon_emission"):
                            r["carbon_emission"] = "0.0"
                    machine_parts_data.append({
                        "machine_row_qty": len(rows),
                        "rows": rows
                    })
                # If you want to always have at least one part
                if not machine_parts_data:
                    machine_parts_data.append({
                        "machine_row_qty": 1,
                        "rows": [{
                            "machine_type": "",
                            "energy_use": "0.0",
                            "machining_time": "",
                            "carbon_emission": "0.0"
                        }]
                    })
                global current_part_index
                current_part_index = 0
                update_machine_process_page()
                part_data = machine_parts_data[current_part_index]
                for row_idx, row_widgets in enumerate(machine_process_row_widgets):
                    if row_idx < len(part_data["rows"]):
                        row = part_data["rows"][row_idx]
                        row_widgets["machine_type_combo"].set(row.get("machine_type", ""))
                        row_widgets["energy_use_value_label"].config(text=row.get("energy_use", "0.0"))
                        row_widgets["machining_time_entry"].delete(0, tk.END)
                        row_widgets["machining_time_entry"].insert(0, row.get("machining_time", ""))
                        row_widgets["carbon_emission_value_label"].config(text=row.get("carbon_emission", "0.0"))
                        update_carbon_emission(
                            row_widgets["machining_time_entry"],
                            row_widgets["energy_use_value_label"],
                            row_widgets["carbon_emission_value_label"],
                            widgets.get('weight_after_entry')
                        )
                recalculate_all_machine_part_emissions()
                if "part_total_emission_var" in part_data:
                    part_data["part_total_emission_var"].set(part_data.get("carbon_emission", "0.0"))
                from pprint import pprint
                print("[DEBUG][reload] machine_parts_data after reload:")
                pprint(machine_parts_data)

            # IMPORTANT: Force update of carbon emission overview BEFORE clearing the restoring_profile flag
            transport_country = widgets.get('transport_country_combo').get()
            rebuild_carbon_emission_overview(transport_country)

            global restoring_profile
            restoring_profile = False

            # Now recalculate everything with restoring_profile == False
            update_total_weight_and_recalculate()
            update_weight_after_total_and_recycled()
            recalculate_all_machine_part_emissions()
            update_sum_of_emissions()
            update_pie_chart_mainthread()  # <-- Force pie chart update after all recalculations
            print("[DEBUG] update_pie_chart_mainthread called from restore_machine_rows_and_finalize")

            # Ensure all parts' data is synced and displayed
            sync_all_parts_after_restore()

        # Start the restoration process
        root.after(200, restore_transport_rows)

        weight_entry.bind("<KeyRelease>", update_total_weight_and_recalculate)
        weight_entry.bind("<FocusOut>", update_total_weight_and_recalculate)
        quantity_entry_weight.bind("<KeyRelease>", update_total_weight_and_recalculate)
        quantity_entry_weight.bind("<FocusOut>", update_total_weight_and_recalculate)
        weight_after_entry.bind("<KeyRelease>", update_weight_after_total_and_recycled)
        weight_after_entry.bind("<FocusOut>", update_weight_after_total_and_recycled)
        weight_after_quantity_entry.bind("<KeyRelease>", update_weight_after_total_and_recycled)
        weight_after_quantity_entry.bind("<FocusOut>", update_weight_after_total_and_recycled)

    def sync_all_parts_after_restore():
        global current_part_index
        original_index = current_part_index
        for idx in range(len(machine_parts_data)):
            current_part_index = idx
            update_machine_process_page()
            root.update_idletasks()
            sync_current_part_ui_to_data()
        current_part_index = original_index
        update_machine_process_page()

    # Restore all widget values and dynamic rows
    restore_profile_inputs(
        profile_data,
        widgets,
        generate_transport_rows,
        recalculate_all
    )
    sync_all_parts_after_restore()  
    store_full_state(widgets)
    
def main():
    ensure_default_files()
    print(f"Running app version: {APP_VERSION}") 
    root = tk.Tk()
    root.configure(bg="#A8D5BA")
    root.title("SPHERE Carbon Emission Calculator")
    check_for_update_and_prompt(root)
    menubar = tk.Menu(root)
    settings_menu = tk.Menu(menubar, tearoff=0)
    settings_menu.add_command(
        label="Edit Emission Factors",
        command=lambda: show_custom_emission_factors_dialog(
            root, materials, on_update=None  # No main page to update from cover
        )
    )
    settings_menu.add_command(
    label="Check for Updates",
    command=lambda: check_for_update_and_prompt(root)
    )
    menubar.add_cascade(label="Settings", menu=settings_menu)
    root.config(menu=menubar)

    # Create the cover page once and keep it alive
    cover_page = CoverPage(root, show_main_page_callback=None)  # Callback set later

    def show_main_page_callback(profile_data):
        cover_page.pack_forget()
        def back_to_cover():
            for widget in root.winfo_children():
                if widget is not cover_page:
                    widget.destroy()
            cover_page.pack(fill="both", expand=True)
        show_main_page(
            root, profile_data, back_to_cover, cover_page, profile_data.get("product_name"), settings_menu
        )
    cover_page.show_main_page_callback = show_main_page_callback
    cover_page.pack(fill="both", expand=True)
    root.mainloop()

if __name__ == "__main__":
    main()